"""
run_cam_soil_constrained_calibration.py

CAM calibration runner for the clean green-roof model.

Goal:
- Improve the CAM temperature graph against measured NI data.
- Minimize RMSE/MAE within bounded, physically interpretable parameters.
- Avoid unrealistic CAM daytime evapotranspiration by adding a soft penalty
  based on the observed CAM soil-moisture behavior.

Why this runner exists:
CAM soil-moisture data did not show a clean repeated daytime drydown, while C3 did.
So CAM tuning should not simply increase daytime latent cooling until the temperature
curve looks good. This script lets the optimizer tune shading/solar absorption/thermal
inertia while keeping CAM daytime plant-cooling flux bounded.

Default raw-file workflow expects these files in BASE_DIR:
    weatherfile mar-april.xlsx
    Pengukuran 30_1 Maret 2026.xlsx
    Pengukuran 30_2 Maret 2026.xlsx
    Pengukuran 3 April 2026.xlsx
    sensor 1 COM5_CAM.csv
    sensor 2 COM6_CAM.csv

Simpler workflow:
    python run_cam_soil_constrained_calibration.py --input merged_cam.csv

Required merged columns for --input:
    T_air_C, T_in_C, solar_W_m2, target column (default T1Tb or T_target_C)
Optional merged columns:
    soil_moisture, RH_pct, theta_s1_pct, theta_s2_pct

Run:
    python run_cam_soil_constrained_calibration.py
    python run_cam_soil_constrained_calibration.py --target-col T2A2
    python run_cam_soil_constrained_calibration.py --soil-constraint none
"""

from __future__ import annotations

import argparse
import json
import math
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import asdict, replace
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

import green_roof_model_clean as gr

try:
    from scipy.optimize import differential_evolution, minimize
    SCIPY_AVAILABLE = True
except Exception:  # pragma: no cover
    SCIPY_AVAILABLE = False


# ==============================================================================
# 01 — DEFAULT CONFIGURATION
# ==============================================================================

DEFAULT_CAM_WINDOW = ("2026-03-31 11:58:00", "2026-04-02 21:42:00")
DEFAULT_NI_FILES = [
    "Pengukuran 30_1 Maret 2026.xlsx",
    "Pengukuran 30_2 Maret 2026.xlsx",
    "Pengukuran 3 April 2026.xlsx",
]

# T1Tb is the safer default for indoor/underside CAM roof temperature.
# T2A2 was treated as uncertain/TBD in earlier code, so keep it as an option.
DEFAULT_TARGET_COL = "T1Tb"
DEFAULT_TIN_COL = "T1Ka"

# Parameters are intentionally bounded. The optimizer can improve fit, but should
# not be allowed to invent unphysical CAM daytime ET or extreme material properties.
PARAM_BOUNDS = {
    "cam_amplitude_scale": (0.05, 1.20),
    "cam_day_max_W_m2": (0.0, 10.0),
    "cam_night_max_W_m2": (0.0, 25.0),
    "lai": (0.40, 2.50),
    "cover_fraction": (0.65, 0.98),
    "solar_absorptivity": (0.40, 0.95),
    "h_out_W_m2K": (5.0, 22.0),
    "h_in_W_m2K": (2.0, 10.0),
    "soil_k_W_mK": (0.08, 0.95),
    "soil_rho_kg_m3": (350.0, 1400.0),
    "soil_cp_J_kgK": (900.0, 2200.0),
    "concrete_k_W_mK": (0.80, 2.20),
    "initial_temp_C": (20.0, 45.0),
}

BASELINE_PARAMS = {
    "cam_amplitude_scale": 0.55,
    "cam_day_max_W_m2": 8.0,
    "cam_night_max_W_m2": 14.0,
    "lai": 0.80,
    "cover_fraction": 0.95,
    "solar_absorptivity": 0.68,
    "h_out_W_m2K": 12.0,
    "h_in_W_m2K": 8.0,
    "soil_k_W_mK": 0.55,
    "soil_rho_kg_m3": 1300.0,
    "soil_cp_J_kgK": 1450.0,
    "concrete_k_W_mK": 1.40,
    "initial_temp_C": None,
}


# ==============================================================================
# 02 — GENERAL UTILITIES
# ==============================================================================

def _to_float(x) -> float:
    if x is None:
        return np.nan
    if isinstance(x, str):
        x = x.strip().replace(",", ".")
        if x in {"", "---", "--", "nan", "NaN"}:
            return np.nan
    return pd.to_numeric(x, errors="coerce")


def _clean_temperature_channels(df: pd.DataFrame, cols: Iterable[str]) -> pd.DataFrame:
    out = df.copy()
    for col in cols:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
            bad = (out[col] < -10) | (out[col] > 80)
            if bad.any():
                out.loc[bad, col] = np.nan
                out[col] = out[col].interpolate("time", limit=30, limit_direction="both")
    return out


def _window_df(df: pd.DataFrame, start: str, end: str) -> pd.DataFrame:
    start_ts = pd.Timestamp(start)
    end_ts = pd.Timestamp(end)
    return df[(df.index >= start_ts) & (df.index <= end_ts)].copy()


def _safe_metric_dict(metrics: Dict[str, float]) -> Dict[str, Optional[float]]:
    clean = {}
    for key, value in metrics.items():
        if value is None:
            clean[key] = None
        elif isinstance(value, (float, np.floating)) and not np.isfinite(value):
            clean[key] = None
        else:
            clean[key] = float(value) if isinstance(value, (float, np.floating)) else value
    return clean


# ==============================================================================
# 03 — RAW DATA LOADERS
# ==============================================================================

def load_weather_excel(filepath: Path, sheet_name: str = "3-24april") -> pd.DataFrame:
    """Load Davis weather Excel/TXT into model columns."""
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Weather file not found: {filepath}")

    if filepath.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            ts = pd.to_datetime(f"{row[0]} {row[1]}", dayfirst=True, errors="coerce")
            if pd.isna(ts):
                continue
            rows.append({
                "datetime": ts,
                "T_air_C": _to_float(row[2]),
                "RH_pct": _to_float(row[5]),
                "wind": _to_float(row[7]),
                "rain_mm": _to_float(row[17]),
                "solar_W_m2": _to_float(row[19]),
            })
        df = pd.DataFrame(rows)
        if df.empty:
            raise ValueError(f"Weather file parsed empty: {filepath}")
        df = df.set_index("datetime").sort_index()
    else:
        raw = pd.read_csv(filepath, sep="\t", skiprows=2, header=None, low_memory=False)
        ts = pd.to_datetime(raw.iloc[:, 0].astype(str).str.strip() + " " + raw.iloc[:, 1].astype(str).str.strip(),
                            dayfirst=True, errors="coerce")
        df = pd.DataFrame({
            "T_air_C": pd.to_numeric(raw.iloc[:, 2], errors="coerce"),
            "RH_pct": pd.to_numeric(raw.iloc[:, 5], errors="coerce"),
            "wind": pd.to_numeric(raw.iloc[:, 7], errors="coerce"),
            "rain_mm": pd.to_numeric(raw.iloc[:, 17], errors="coerce"),
            "solar_W_m2": pd.to_numeric(raw.iloc[:, 19], errors="coerce"),
        }, index=ts).dropna(axis=0, how="all")
        df.index.name = "datetime"

    df = df[~df.index.duplicated(keep="first")].sort_index()
    for col, lo, hi in [("solar_W_m2", 0, None), ("RH_pct", 0, 100), ("wind", 0.1, None), ("rain_mm", 0, None)]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df[col] = df[col].clip(lower=lo, upper=hi)
    for col in ["T_air_C", "RH_pct", "solar_W_m2", "wind"]:
        if col in df.columns:
            df[col] = df[col].interpolate("time", limit=30, limit_direction="both")
    return df[[c for c in ["T_air_C", "RH_pct", "solar_W_m2", "wind", "rain_mm"] if c in df.columns]]


def load_ni_excel(filepath: Path) -> pd.DataFrame:
    """Load LabVIEW/NI Excel file using direct XML parsing."""
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"NI file not found: {filepath}")

    names = [
        "timestamp_serial",
        "T1Kd", "T1Kb", "T1Ke", "T1Ka", "T1Kc",
        "T3Kb", "T3Kd", "T3Ke", "T3Kc", "T3Ka",
        "T2Ka", "T2Kd", "T2Kc", "2Ke",  "T1A",
        "T2A",  "T2A2", "T1Tb", "T1Ta", "T1Ta2",
    ]

    with zipfile.ZipFile(filepath, "r") as z:
        sheet_names = [n for n in z.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml", n)]
        if not sheet_names:
            raise ValueError(f"Cannot find worksheet XML in {filepath}")
        with z.open(sheet_names[0]) as f:
            root = ET.fromstring(f.read())

    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    data = []
    for row in root.findall(f".//{ns}row")[1:]:
        vals = []
        for cell in row.findall(f"{ns}c"):
            v = cell.find(f"{ns}v")
            vals.append(float(v.text) if v is not None else np.nan)
        if len(vals) >= 21:
            data.append(vals[:21])

    df = pd.DataFrame(data, columns=names)
    if df.empty:
        raise ValueError(f"NI file parsed empty: {filepath}")
    labview_epoch = pd.Timestamp("1904-01-01")
    df["datetime"] = labview_epoch + pd.to_timedelta(df["timestamp_serial"], unit="D")
    df = df.drop(columns=["timestamp_serial"]).set_index("datetime").sort_index()
    df = _clean_temperature_channels(df, [c for c in df.columns if c.startswith("T")])
    return df


def load_multiple_ni(files: Sequence[Path]) -> pd.DataFrame:
    frames = []
    missing = []
    for fp in files:
        fp = Path(fp)
        if fp.exists():
            frames.append(load_ni_excel(fp))
        else:
            missing.append(str(fp))
    if not frames:
        raise FileNotFoundError("No NI files found. Missing: " + ", ".join(missing))
    df = pd.concat(frames).sort_index()
    df = df[~df.index.duplicated(keep="last")]
    return df


def _read_text_with_encoding(filepath: Path) -> str:
    data = Path(filepath).read_bytes()
    if data[:200].count(b"\x00") > 40:
        return data.decode("utf-16le", errors="ignore").replace("\x00", "")
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("latin1", errors="ignore")


def load_soil_sensor_csv(filepath: Path, label: str, timestamp_mode: str = "utc_to_wib") -> pd.DataFrame:
    """Load one RIKA CSV exported with GMT timestamp."""
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Soil sensor file not found: {filepath}")
    txt = _read_text_with_encoding(filepath)
    from io import StringIO
    df = pd.read_csv(StringIO(txt), sep=";")
    if "Timestamp" not in df.columns:
        raise ValueError(f"Timestamp column not found in {filepath}. Columns: {df.columns.tolist()}")

    if timestamp_mode == "utc_to_wib":
        ts = pd.to_datetime(df["Timestamp"], utc=True, errors="coerce").dt.tz_convert("Asia/Jakarta").dt.tz_localize(None)
    elif timestamp_mode == "local":
        ts = pd.to_datetime(df["Timestamp"].astype(str).str.replace(" GMT", "", regex=False), errors="coerce")
    else:
        raise ValueError("timestamp_mode must be 'utc_to_wib' or 'local'.")

    # IMPORTANT:
    # Do not pass pandas Series directly while also passing index=ts.
    # Pandas aligns Series by their original RangeIndex to the DatetimeIndex,
    # which turns every value into NaN and makes the soil plot say no_soil_data.
    # Convert both timestamp and values to arrays first.
    ts_index = pd.DatetimeIndex(ts)
    out = pd.DataFrame({
        f"soil_temp_{label}_C": pd.to_numeric(df.get("Temperature"), errors="coerce").to_numpy(),
        f"moisture_{label}_pct": pd.to_numeric(df.get("Moisture (%)"), errors="coerce").to_numpy(),
    }, index=ts_index)
    out.index.name = "datetime"
    out = out.dropna(how="all").sort_index()
    return out


def load_cam_soil_pair(s1_path: Path, s2_path: Path, timestamp_mode: str = "utc_to_wib", swap_depths: bool = False) -> pd.DataFrame:
    s1 = load_soil_sensor_csv(s1_path, "s1", timestamp_mode=timestamp_mode)
    s2 = load_soil_sensor_csv(s2_path, "s2", timestamp_mode=timestamp_mode)
    soil = pd.concat([s1, s2], axis=1).sort_index()
    soil = soil[~soil.index.duplicated(keep="first")]

    # Flag impossible glitches but do not over-clean long trends.
    for col in ["moisture_s1_pct", "moisture_s2_pct"]:
        if col in soil:
            bad = (soil[col] < 3) | (soil[col] > 95)
            soil.loc[bad, col] = np.nan
            soil[col] = soil[col].interpolate("time", limit=15, limit_direction="both")

    if swap_depths:
        soil["theta_shallow_pct"] = soil["moisture_s2_pct"]
        soil["theta_deep_pct"] = soil["moisture_s1_pct"]
    else:
        soil["theta_shallow_pct"] = soil["moisture_s1_pct"]
        soil["theta_deep_pct"] = soil["moisture_s2_pct"]

    soil["soil_moisture"] = soil[["theta_shallow_pct", "theta_deep_pct"]].mean(axis=1) / 100.0
    return soil


# ==============================================================================
# 04 — DATA PREPARATION
# ==============================================================================

def load_merged_input(path: Path, target_col: Optional[str] = None) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)

    # Datetime detection.
    dt_candidates = ["datetime", "timestamp", "Timestamp", "time", "DateTime", "Date Time"]
    for col in dt_candidates:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
            df = df.dropna(subset=[col]).set_index(col)
            break
    df = df.sort_index()

    aliases = {
        "T_air_C": ["T_air_C", "T_a", "T_out", "outdoor", "temperature_2m", "Temp Out", "T_ambient"],
        "T_in_C": ["T_in_C", "T_in", "T1Ka", "indoor", "inside", "T_in_CAM"],
        "solar_W_m2": ["solar_W_m2", "G_sol", "Solar Rad.", "shortwave_radiation", "solar", "radiation"],
        "RH_pct": ["RH_pct", "RH", "Out Hum", "relative_humidity_2m"],
        "soil_moisture": ["soil_moisture", "theta", "theta_mean", "VWC", "moisture", "Moisture (%)"],
    }
    for canonical, opts in aliases.items():
        if canonical not in df.columns:
            for opt in opts:
                if opt in df.columns:
                    df[canonical] = pd.to_numeric(df[opt], errors="coerce")
                    break

    if target_col and target_col in df.columns:
        df["T_target_C"] = pd.to_numeric(df[target_col], errors="coerce")
    elif "T_target_C" not in df.columns:
        for opt in ["T1Tb", "T2A2", "CAM", "T_s_in_CAM", "target"]:
            if opt in df.columns:
                df["T_target_C"] = pd.to_numeric(df[opt], errors="coerce")
                break

    required = ["T_air_C", "T_in_C", "solar_W_m2", "T_target_C"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Merged input missing columns {missing}. Existing columns: {df.columns.tolist()}")

    return df[~df.index.duplicated(keep="first")].sort_index()


def build_input_from_raw_files(args: argparse.Namespace) -> Tuple[pd.DataFrame, pd.DataFrame]:
    base = Path(args.base_dir)
    window_start, window_end = args.window_start, args.window_end

    weather = load_weather_excel(base / args.weather_file, sheet_name=args.weather_sheet)
    ni_files = [base / f for f in args.ni_files]
    ni = load_multiple_ni(ni_files)
    soil = load_cam_soil_pair(base / args.soil_s1, base / args.soil_s2,
                              timestamp_mode=args.timestamp_mode, swap_depths=args.swap_depths)

    weather = _window_df(weather, window_start, window_end).resample("1min").mean().interpolate("time")
    ni = _window_df(ni, window_start, window_end).resample("1min").mean().interpolate("time")
    soil = _window_df(soil, window_start, window_end).resample("1min").mean().interpolate("time", limit=15)

    if weather.empty:
        raise ValueError(f"Weather has no data in selected window {window_start} -> {window_end}")
    if ni.empty:
        raise ValueError(f"NI has no data in selected window {window_start} -> {window_end}")
    if soil.empty:
        print("WARNING: soil data empty in selected window. Calibration will run without soil-moisture summary.")

    if args.target_col not in ni.columns:
        raise ValueError(f"Target column {args.target_col!r} not found in NI columns: {ni.columns.tolist()}")
    if args.tin_col not in ni.columns:
        raise ValueError(f"Indoor boundary column {args.tin_col!r} not found in NI columns: {ni.columns.tolist()}")

    merged = pd.concat([
        weather[["T_air_C", "RH_pct", "solar_W_m2"]],
        ni[[args.tin_col, args.target_col]].rename(columns={args.tin_col: "T_in_C", args.target_col: "T_target_C"}),
        soil[[c for c in ["soil_moisture", "theta_shallow_pct", "theta_deep_pct", "moisture_s1_pct", "moisture_s2_pct"] if c in soil.columns]],
    ], axis=1).sort_index()

    merged = merged.interpolate("time", limit=30, limit_direction="both")
    merged = merged.dropna(subset=["T_air_C", "T_in_C", "solar_W_m2", "T_target_C"])
    return merged, soil


# ==============================================================================
# 05 — SOIL-MOISTURE CONSTRAINT
# ==============================================================================

def slope_per_hour(series: pd.Series) -> float:
    s = series.dropna()
    if len(s) < 3:
        return np.nan
    x = (s.index - s.index[0]).total_seconds().to_numpy() / 3600.0
    y = s.to_numpy(dtype=float)
    if np.ptp(x) <= 0:
        return np.nan
    return float(np.polyfit(x, y, 1)[0])


def summarize_soil_constraint(df: pd.DataFrame) -> Dict[str, float | str | None]:
    """Summarize daytime CAM moisture behavior for physical constraint."""
    out: Dict[str, float | str | None] = {
        "soil_available": False,
        "shallow_day_slope_pctpt_per_h": None,
        "deep_day_slope_pctpt_per_h": None,
        "classification": "no_soil_data",
        "recommended_max_day_q_plant_W_m2": 5.0,
    }
    if df.empty or not isinstance(df.index, pd.DatetimeIndex):
        return out

    work = df.copy()
    if "solar_W_m2" in work.columns:
        day = work[work["solar_W_m2"] > 50]
    else:
        hour = work.index.hour + work.index.minute / 60.0
        day = work[(hour >= 7.0) & (hour <= 17.5)]

    if day.empty:
        return out

    shallow_col = "theta_shallow_pct" if "theta_shallow_pct" in day.columns else None
    deep_col = "theta_deep_pct" if "theta_deep_pct" in day.columns else None
    shallow_slope = slope_per_hour(day[shallow_col]) if shallow_col else np.nan
    deep_slope = slope_per_hour(day[deep_col]) if deep_col else np.nan

    out["soil_available"] = True
    out["shallow_day_slope_pctpt_per_h"] = None if not np.isfinite(shallow_slope) else float(shallow_slope)
    out["deep_day_slope_pctpt_per_h"] = None if not np.isfinite(deep_slope) else float(deep_slope)

    # For CAM, flat or ascending moisture means do not allow high daytime plant cooling.
    # A strong drydown would relax this penalty slightly.
    ref_slope = shallow_slope if np.isfinite(shallow_slope) else deep_slope
    if not np.isfinite(ref_slope):
        out["classification"] = "unknown"
        out["recommended_max_day_q_plant_W_m2"] = 5.0
    elif ref_slope <= -1.0:
        out["classification"] = "descending"
        out["recommended_max_day_q_plant_W_m2"] = 8.0
    elif ref_slope >= 0.3:
        out["classification"] = "ascending_or_recharged"
        out["recommended_max_day_q_plant_W_m2"] = 3.5
    else:
        out["classification"] = "flat_or_weak_drydown"
        out["recommended_max_day_q_plant_W_m2"] = 4.5
    return out


# ==============================================================================
# 06 — MODEL CONFIG AND OPTIMIZATION
# ==============================================================================

def config_from_params(params: Dict[str, float], target_col: str = "T_target_C") -> gr.SimulationConfig:
    initial_temp = params.get("initial_temp_C")
    if initial_temp is None or not np.isfinite(initial_temp):
        initial_temp = None

    config = gr.build_default_config(
        plant="cam",
        soil_thickness_m=0.10,
        concrete_thickness_m=0.10,
        lai=params["lai"],
        cover_fraction=params["cover_fraction"],
        cam_amplitude_scale=params["cam_amplitude_scale"],
        h_in_W_m2K=params["h_in_W_m2K"],
        h_out_W_m2K=params["h_out_W_m2K"],
        solar_absorptivity=params["solar_absorptivity"],
        target_col=target_col,
    )
    config = gr.with_updated_plant(
        config,
        cam_day_max_W_m2=params["cam_day_max_W_m2"],
        cam_night_max_W_m2=params["cam_night_max_W_m2"],
        q_plant_cap_W_m2=35.0,
    )
    config = gr.with_updated_boundary(config, initial_temp_C=initial_temp)

    layers = list(config.layers)
    layers[0] = replace(
        layers[0],
        thickness_m=0.10,
        k_W_mK=params["soil_k_W_mK"],
        rho_kg_m3=params["soil_rho_kg_m3"],
        cp_J_kgK=params["soil_cp_J_kgK"],
    )
    layers[1] = replace(layers[1], thickness_m=0.10, k_W_mK=params["concrete_k_W_mK"])
    config = replace(config, layers=tuple(layers), spinup_steps=0)
    config.validate()
    return config


def run_model(df: pd.DataFrame, params: Dict[str, float]) -> pd.DataFrame:
    config = config_from_params(params, target_col="T_target_C")
    return gr.green_roof_model_clean(df, config=config, diagnostics=False)


def compute_extra_metrics(out: pd.DataFrame) -> Dict[str, float]:
    metrics = gr.compute_metrics(out["T_target_C"], out["T_pred_C"])
    obs = pd.to_numeric(out["T_target_C"], errors="coerce")
    pred = pd.to_numeric(out["T_pred_C"], errors="coerce")
    mask = np.isfinite(obs) & np.isfinite(pred)
    if mask.any():
        metrics["amp_obs_C"] = float(obs[mask].max() - obs[mask].min())
        metrics["amp_pred_C"] = float(pred[mask].max() - pred[mask].min())
        metrics["amp_error_C"] = metrics["amp_pred_C"] - metrics["amp_obs_C"]
        metrics["peak_error_C"] = float(pred[mask].max() - obs[mask].max())
        metrics["min_error_C"] = float(pred[mask].min() - obs[mask].min())
    return metrics


def params_to_vector(params: Dict[str, float], names: List[str]) -> np.ndarray:
    return np.asarray([params[n] for n in names], dtype=float)


def vector_to_params(x: Iterable[float], names: List[str], base: Dict[str, float]) -> Dict[str, float]:
    params = dict(base)
    for name, value in zip(names, x):
        lo, hi = PARAM_BOUNDS[name]
        params[name] = float(np.clip(value, lo, hi))
    return params


def make_objective(df: pd.DataFrame, soil_summary: Dict[str, float | str | None], args: argparse.Namespace):
    names = list(PARAM_BOUNDS.keys())
    bounds = [PARAM_BOUNDS[n] for n in names]
    max_day_q = float(args.max_day_q_plant if args.max_day_q_plant is not None else soil_summary.get("recommended_max_day_q_plant_W_m2", 4.5))

    trials: List[Dict[str, float]] = []

    def objective(x: np.ndarray) -> float:
        params = vector_to_params(x, names, BASELINE_PARAMS)
        try:
            out = run_model(df, params)
            # Target copied by model when target_col exists.
            if "T_target_C" not in out.columns:
                out["T_target_C"] = df["T_target_C"]
            metrics = compute_extra_metrics(out)
            rmse = metrics.get("rmse_C", np.inf)
            amp_error = abs(metrics.get("amp_error_C", 0.0))
            if not np.isfinite(rmse):
                return 1e6

            solar = df["solar_W_m2"].reindex(out.index).fillna(0)
            day_mask = solar > 50
            day_q_mean = float(out.loc[day_mask, "q_plant_W_m2"].mean()) if day_mask.any() else 0.0
            night_q_mean = float(out.loc[~day_mask, "q_plant_W_m2"].mean()) if (~day_mask).any() else 0.0

            soil_penalty = 0.0
            if args.soil_constraint != "none":
                excess = max(0.0, day_q_mean - max_day_q)
                soil_penalty = args.soil_penalty_weight * (excess ** 2)

            # Prefer CAM daytime plant cooling not to exceed nocturnal cooling by too much,
            # unless the soil data support drydown.
            ratio_penalty = 0.0
            if args.soil_constraint == "strict" and night_q_mean > 0.25:
                ratio_excess = max(0.0, day_q_mean / night_q_mean - 1.25)
                ratio_penalty = args.soil_penalty_weight * ratio_excess

            score = rmse + args.amplitude_weight * amp_error + soil_penalty + ratio_penalty

            row = {"score": float(score), "day_q_mean_W_m2": day_q_mean, "night_q_mean_W_m2": night_q_mean, **metrics, **params}
            trials.append(row)
            return float(score)
        except Exception as exc:
            trials.append({"score": 1e6, "error": str(exc)})
            return 1e6

    return objective, names, bounds, trials


def calibrate(df: pd.DataFrame, soil_summary: Dict[str, float | str | None], args: argparse.Namespace) -> Tuple[Dict[str, float], pd.DataFrame, pd.DataFrame]:
    objective, names, bounds, trials = make_objective(df, soil_summary, args)

    # Baseline initial temperature follows first measured target for less startup bias.
    base = dict(BASELINE_PARAMS)
    if len(df) and np.isfinite(df["T_target_C"].iloc[0]):
        base["initial_temp_C"] = float(df["T_target_C"].iloc[0])

    x0 = params_to_vector(base, names)
    baseline_score = objective(x0)
    print(f"Baseline objective score: {baseline_score:.4f}")

    if SCIPY_AVAILABLE:
        print("Running bounded global search...")
        result_de = differential_evolution(
            objective,
            bounds=bounds,
            maxiter=args.maxiter,
            popsize=args.popsize,
            seed=args.seed,
            polish=False,
            updating="immediate",
            workers=1,
            tol=0.01,
        )
        print(f"Global search best score: {result_de.fun:.4f}")

        print("Polishing with L-BFGS-B...")
        result_local = minimize(
            objective,
            result_de.x,
            method="L-BFGS-B",
            bounds=bounds,
            options={"maxiter": args.local_maxiter, "ftol": 1e-7},
        )
        best_x = result_local.x if result_local.fun <= result_de.fun else result_de.x
    else:
        print("SciPy not available. Using random bounded search fallback.")
        rng = np.random.default_rng(args.seed)
        best_x = x0.copy()
        best_score = baseline_score
        for _ in range(args.random_trials):
            x = np.asarray([rng.uniform(lo, hi) for lo, hi in bounds], dtype=float)
            score = objective(x)
            if score < best_score:
                best_score = score
                best_x = x

    best_params = vector_to_params(best_x, names, base)
    best_out = run_model(df, best_params)
    if "T_target_C" not in best_out.columns:
        best_out["T_target_C"] = df["T_target_C"]

    baseline_out = run_model(df, base)
    if "T_target_C" not in baseline_out.columns:
        baseline_out["T_target_C"] = df["T_target_C"]

    trials_df = pd.DataFrame(trials).sort_values("score") if trials else pd.DataFrame()
    return best_params, best_out, baseline_out, trials_df


# ==============================================================================
# 07 — OUTPUTS
# ==============================================================================

def save_outputs(df: pd.DataFrame, best_params: Dict[str, float], best_out: pd.DataFrame,
                 baseline_out: pd.DataFrame, trials: pd.DataFrame,
                 soil_summary: Dict[str, float | str | None], args: argparse.Namespace) -> None:
    outdir = Path(args.output_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    best_metrics = compute_extra_metrics(best_out)
    baseline_metrics = compute_extra_metrics(baseline_out)

    # Add q summaries.
    for label, out, metrics in [("calibrated", best_out, best_metrics), ("baseline", baseline_out, baseline_metrics)]:
        solar = df["solar_W_m2"].reindex(out.index).fillna(0)
        day_mask = solar > 50
        metrics["day_q_plant_mean_W_m2"] = float(out.loc[day_mask, "q_plant_W_m2"].mean()) if day_mask.any() else np.nan
        metrics["night_q_plant_mean_W_m2"] = float(out.loc[~day_mask, "q_plant_W_m2"].mean()) if (~day_mask).any() else np.nan

    best_out.to_csv(outdir / "cam_prediction_calibrated.csv", index_label="datetime")
    baseline_out.to_csv(outdir / "cam_prediction_baseline.csv", index_label="datetime")
    if not trials.empty:
        trials.to_csv(outdir / "cam_calibration_trials.csv", index=False)

    with open(outdir / "cam_best_params.json", "w", encoding="utf-8") as f:
        json.dump(best_params, f, indent=2)
    with open(outdir / "cam_metrics_calibrated.json", "w", encoding="utf-8") as f:
        json.dump(_safe_metric_dict(best_metrics), f, indent=2)
    with open(outdir / "cam_metrics_baseline.json", "w", encoding="utf-8") as f:
        json.dump(_safe_metric_dict(baseline_metrics), f, indent=2)
    with open(outdir / "cam_soil_constraint_summary.json", "w", encoding="utf-8") as f:
        json.dump(soil_summary, f, indent=2)

    plot_calibration(df, best_out, baseline_out, soil_summary, outdir / "cam_calibrated_vs_measured.png")
    plot_fluxes(df, best_out, baseline_out, outdir / "cam_calibrated_flux_check.png")

    print("\n=== BASELINE METRICS ===")
    print(json.dumps(_safe_metric_dict(baseline_metrics), indent=2))
    print("\n=== CALIBRATED METRICS ===")
    print(json.dumps(_safe_metric_dict(best_metrics), indent=2))
    print("\n=== BEST PARAMS ===")
    print(json.dumps(best_params, indent=2))
    print(f"\nSaved outputs to: {outdir.resolve()}")


def plot_calibration(df: pd.DataFrame, best_out: pd.DataFrame, baseline_out: pd.DataFrame,
                     soil_summary: Dict[str, float | str | None], path: Path) -> None:
    fig, axes = plt.subplots(4, 1, figsize=(14, 12), sharex=True)

    target = df["T_target_C"].reindex(best_out.index)
    axes[0].plot(target.index, target.values, label="Measured CAM target", linewidth=2.0)
    axes[0].plot(baseline_out.index, baseline_out["T_pred_C"], label="Baseline model", linestyle="--", linewidth=1.6)
    axes[0].plot(best_out.index, best_out["T_pred_C"], label="Calibrated model", linewidth=1.8)
    axes[0].set_ylabel("T_s,in (°C)")
    axes[0].set_title("CAM calibration: measured vs baseline vs calibrated")
    axes[0].legend()
    axes[0].grid(True, alpha=0.3)

    err_base = baseline_out["T_pred_C"] - target
    err_best = best_out["T_pred_C"] - target
    axes[1].plot(err_base.index, err_base.values, label="Baseline error", linestyle="--")
    axes[1].plot(err_best.index, err_best.values, label="Calibrated error")
    axes[1].axhline(0, linestyle=":", linewidth=1)
    axes[1].set_ylabel("Model - measured (°C)")
    axes[1].legend()
    axes[1].grid(True, alpha=0.3)

    axes[2].plot(best_out.index, df["T_air_C"].reindex(best_out.index), label="T_air_C")
    axes[2].plot(best_out.index, df["T_in_C"].reindex(best_out.index), label="T_in_C")
    axes[2].set_ylabel("Temperature (°C)")
    axes[2].legend(loc="upper left")
    axes[2].grid(True, alpha=0.3)
    ax2 = axes[2].twinx()
    ax2.plot(best_out.index, df["solar_W_m2"].reindex(best_out.index), label="solar_W_m2", linestyle="--", alpha=0.7)
    ax2.set_ylabel("Solar (W/m²)")

    for col in ["theta_shallow_pct", "theta_deep_pct"]:
        if col in df.columns:
            axes[3].plot(df.index, df[col], label=col)
    axes[3].set_ylabel("Moisture (%)")
    axes[3].set_xlabel("Datetime")
    axes[3].grid(True, alpha=0.3)
    axes[3].legend()
    axes[3].set_title(f"CAM soil constraint: {soil_summary.get('classification')}")

    fig.tight_layout()
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def plot_fluxes(df: pd.DataFrame, best_out: pd.DataFrame, baseline_out: pd.DataFrame, path: Path) -> None:
    fig, axes = plt.subplots(3, 1, figsize=(14, 9), sharex=True)
    axes[0].plot(best_out.index, baseline_out["q_plant_W_m2"], label="Baseline q_plant", linestyle="--")
    axes[0].plot(best_out.index, best_out["q_plant_W_m2"], label="Calibrated q_plant")
    axes[0].set_ylabel("q_plant (W/m²)")
    axes[0].set_title("CAM latent/plant cooling flux check")
    axes[0].legend()
    axes[0].grid(True, alpha=0.3)

    axes[1].plot(best_out.index, baseline_out["q_solar_W_m2"], label="Baseline absorbed solar", linestyle="--")
    axes[1].plot(best_out.index, best_out["q_solar_W_m2"], label="Calibrated absorbed solar")
    axes[1].set_ylabel("q_solar (W/m²)")
    axes[1].legend()
    axes[1].grid(True, alpha=0.3)

    axes[2].plot(best_out.index, baseline_out["q_bottom_W_m2"], label="Baseline q_bottom", linestyle="--")
    axes[2].plot(best_out.index, best_out["q_bottom_W_m2"], label="Calibrated q_bottom")
    axes[2].axhline(0, linestyle=":", linewidth=1)
    axes[2].set_ylabel("q_bottom (W/m²)")
    axes[2].set_xlabel("Datetime")
    axes[2].legend()
    axes[2].grid(True, alpha=0.3)

    fig.tight_layout()
    fig.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(fig)




def print_column_validity(df: pd.DataFrame, label: str = "input") -> None:
    """Print non-NaN counts and ranges for key columns to diagnose NaN windows."""
    print(f"\n=== {label.upper()} COLUMN VALIDITY ===")
    for col in [
        "T_air_C", "T_in_C", "solar_W_m2", "T_target_C",
        "soil_moisture", "theta_shallow_pct", "theta_deep_pct",
    ]:
        if col not in df.columns:
            continue
        s = pd.to_numeric(df[col], errors="coerce")
        n_valid = int(s.notna().sum())
        msg = f"{col:18s}: valid {n_valid:6d} / {len(df):6d}"
        if n_valid > 0:
            msg += f" | min={s.min():.3g}, max={s.max():.3g}"
        else:
            msg += " | all NaN"
        print(msg)


# ==============================================================================
# 08 — CLI
# ==============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Soil-moisture-constrained CAM calibration runner")
    parser.add_argument("--input", type=str, default=None, help="Optional merged CSV/XLSX input file")
    parser.add_argument("--base-dir", type=str, default=".")
    parser.add_argument("--output-dir", type=str, default="outputs_cam_calibrated")

    parser.add_argument("--weather-file", type=str, default="weatherfile mar-april.xlsx")
    parser.add_argument("--weather-sheet", type=str, default="3-24april")
    parser.add_argument("--ni-files", nargs="+", default=DEFAULT_NI_FILES)
    parser.add_argument("--soil-s1", type=str, default="sensor 1 COM5_CAM.csv")
    parser.add_argument("--soil-s2", type=str, default="sensor 2 COM6_CAM.csv")
    parser.add_argument("--timestamp-mode", choices=["utc_to_wib", "local"], default="utc_to_wib")
    parser.add_argument("--swap-depths", action="store_true", help="Use S2 as shallow and S1 as deep")

    parser.add_argument("--window-start", type=str, default=DEFAULT_CAM_WINDOW[0])
    parser.add_argument("--window-end", type=str, default=DEFAULT_CAM_WINDOW[1])
    parser.add_argument("--target-col", type=str, default=DEFAULT_TARGET_COL)
    parser.add_argument("--tin-col", type=str, default=DEFAULT_TIN_COL)

    parser.add_argument("--soil-constraint", choices=["none", "soft", "strict"], default="soft")
    parser.add_argument("--max-day-q-plant", type=float, default=None,
                        help="Override recommended max mean daytime q_plant W/m2")
    parser.add_argument("--soil-penalty-weight", type=float, default=0.08)
    parser.add_argument("--amplitude-weight", type=float, default=0.15)

    parser.add_argument("--maxiter", type=int, default=18, help="Differential evolution max iterations")
    parser.add_argument("--popsize", type=int, default=8)
    parser.add_argument("--local-maxiter", type=int, default=80)
    parser.add_argument("--random-trials", type=int, default=250)
    parser.add_argument("--seed", type=int, default=42)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    outdir = Path(args.output_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    if args.input:
        print(f"Loading merged input: {args.input}")
        df = load_merged_input(Path(args.input), target_col=args.target_col)
        df = _window_df(df, args.window_start, args.window_end)
        soil_like = df
    else:
        print("Loading raw weather + NI + CAM soil files...")
        df, soil_like = build_input_from_raw_files(args)

    if df.empty:
        raise ValueError("Prepared calibration input is empty. Check timestamp mode/window/files.")

    # Normalize soil_moisture if given as %.
    if "soil_moisture" in df.columns and df["soil_moisture"].max(skipna=True) > 1.5:
        df["soil_moisture"] = df["soil_moisture"] / 100.0

    df.to_csv(outdir / "cam_calibration_input_used.csv", index_label="datetime")

    # Build a soil-summary frame without duplicating columns when --input already
    # contains both solar and soil-moisture columns.
    if soil_like is df:
        soil_summary_frame = df.copy()
    else:
        soil_summary_frame = soil_like.copy()
        if "solar_W_m2" not in soil_summary_frame.columns and "solar_W_m2" in df.columns:
            soil_summary_frame = soil_summary_frame.join(df[["solar_W_m2"]], how="left")
    soil_summary = summarize_soil_constraint(soil_summary_frame)

    print("\n=== INPUT SUMMARY ===")
    print(f"Time range : {df.index.min()} -> {df.index.max()}")
    print(f"Rows       : {len(df)}")
    print(f"Target     : {args.target_col} -> T_target_C")
    print(f"T_in       : {args.tin_col} -> T_in_C")
    print_column_validity(df, label="calibration input")
    print("\n=== SOIL CONSTRAINT SUMMARY ===")
    print(json.dumps(soil_summary, indent=2))

    best_params, best_out, baseline_out, trials = calibrate(df, soil_summary, args)
    save_outputs(df, best_params, best_out, baseline_out, trials, soil_summary, args)


if __name__ == "__main__":
    main()
