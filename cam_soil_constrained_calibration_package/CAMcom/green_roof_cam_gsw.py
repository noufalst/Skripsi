"""
================================================================================
GREEN ROOF CAM-ONLY MODEL — DATA-DRIVEN LI-COR gsw / r_s VERSION
================================================================================
Reference model:
    Chagolla-Aranda et al. (2025), Journal of Building Engineering 103, 112053

Purpose:
    CAM-only physical simulation for Bromelia green roof.
    This version explicitly uses LI-COR stomatal conductance (gsw) to build
    a data-driven hourly stomatal resistance profile. It does not force a
    hard day/night CAM multiplier as the default.

Core difference from reduced calibration versions:
    LI-COR gsw -> hourly gsw profile -> r_s(t) -> h_eva_f -> j_eva_f

Main outputs separated for diagnosis:
    j_eva_f      : foliage / plant transpiration contribution [kg/m2/s]
    j_eva_g      : substrate evaporation contribution [kg/m2/s]
    j_eva_total  : total evapotranspiration [kg/m2/s]
    r_s_s_m      : CAM stomatal resistance used in the timestep [s/m]
    gsw_equiv    : equivalent conductance used by model [mol/m2/s]

Notes:
    - This file intentionally follows the readable section structure used in the
      previous green_roof_model.py workflow.
    - Numerical constants are kept visible instead of hidden in many shortcuts.
================================================================================
"""

from __future__ import annotations

# ==============================================================================
# 00 — IMPORTS
# ==============================================================================

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple, Union
import io
import re
import zipfile
import xml.etree.ElementTree as ET

import numpy as np
import pandas as pd

def read_csv_with_encoding_fallback(filepath: Union[str, Path], **kwargs) -> pd.DataFrame:
    """Read CSV robustly for logger exports with non-UTF8 characters.

    Important: some RIKA/soil logger CSV files are UTF-16LE without BOM.
    Pandas can *appear* to read them as UTF-8 but returns broken columns such as
    ["N", "Unnamed: ..."]. Therefore this function validates the decoded
    columns before accepting a result.
    """
    filepath = Path(filepath)
    raw_head = filepath.read_bytes()[:4096]
    null_ratio = raw_head.count(b"\x00") / max(len(raw_head), 1)

    # If many NUL bytes are present, this is almost certainly UTF-16-like.
    if null_ratio > 0.20:
        encodings = ["utf-16le", "utf-16", "utf-16be", "utf-8", "utf-8-sig", "cp1252", "latin1"]
    else:
        encodings = ["utf-8", "utf-8-sig", "cp1252", "latin1", "utf-16le", "utf-16", "utf-16be"]

    expected_any = {
        "timestamp", "moisture (%)", "temperature",
        "time", "date", "gsw", "a", "e", "qin",
    }

    def looks_valid(df: pd.DataFrame) -> bool:
        if df.empty and len(df.columns) == 0:
            return False
        cols = [str(c).strip().lower().replace("\ufeff", "") for c in df.columns]
        if any(c in expected_any for c in cols):
            return True
        # Reject classic wrong-encoding symptom.
        if all((c.startswith("unnamed") or c in {"n", "nan"}) for c in cols):
            return False
        # Accept non-logger CSVs only if there are multiple meaningful columns.
        meaningful = [c for c in cols if c and not c.startswith("unnamed")]
        return len(meaningful) >= 2

    last_error = None
    attempted = []
    for enc in encodings:
        try:
            df = pd.read_csv(filepath, encoding=enc, **kwargs)
            attempted.append((enc, list(df.columns)))
            if looks_valid(df):
                # Normalize BOM/extra spaces from headers.
                df.columns = [str(c).strip().replace("\ufeff", "") for c in df.columns]
                return df
            last_error = ValueError(f"Decoded with {enc}, but columns look invalid: {list(df.columns)}")
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
        except Exception as exc:
            last_error = exc
            continue

    msg = f"Could not read CSV with valid columns: {filepath}. Attempts: {attempted[:4]}"
    raise ValueError(msg) from last_error


# ==============================================================================
# 01 — DATA CLASSES
# ==============================================================================

@dataclass
class PlantParameters:
    """Parameter spesifik tanaman CAM/Bromelia."""
    name: str = "Bromelia (CAM)"
    plant_type: str = "CAM"

    # Canopy geometry
    H_f: float = 0.098             # m, canopy height
    d_f: float = 0.045             # m, representative leaf width
    LAI: float = 1.95              # m2/m2
    cover_fraction: float = 0.95   # projected horizontal cover fraction, 0-1

    # Leaf optical properties
    rho_f: float = 0.390           # reflectivity
    tau_f: float = 0.070           # transmissivity
    alpha_f: float = 0.540         # absorptivity = 1-rho-tau
    epsilon_f: float = 0.95        # emissivity


@dataclass
class SubstrateParameters:
    """Parameter substrat tanah + sekam padi."""
    lambda_dry: float = 0.12       # W/mK, dry substrate
    lambda_sat: float = 0.66       # W/mK, saturated substrate; default recomputed if None
    rho_g: float = 400.0           # kg/m3, bulk density scientific guess
    cp_g: float = 1300.0           # J/kgK

    theta_sat: float = 0.90        # m3/m3, high porosity soil-rice husk guess
    theta_min: float = 0.05        # m3/m3
    k_theta_sat: float = 5e-6      # m/s
    psi_sat: float = 0.35          # m, positive magnitude
    b: float = 5.5                 # Brooks-Corey exponent

    rho_g_rad: float = 0.15        # substrate reflectivity
    epsilon_g: float = 0.95        # substrate emissivity
    lambda_water: float = 0.60     # W/mK
    cp_water: float = 4180.0       # J/kgK
    l_fg: float = 2.45e6           # J/kg


@dataclass
class SlabParameters:
    """Concrete slab properties."""
    lambda_s: float = 1.74         # W/mK
    rho_s: float = 2300.0          # kg/m3
    cp_s: float = 840.0            # J/kgK
    H_slab: float = 0.10           # m


@dataclass
class GeometryParameters:
    """Geometry and boundary parameters."""
    H_g: float = 0.10              # m, substrate thickness used by current experiment
    A_roof: float = 1.0            # m2, optional total area scaling
    T_in_default: float = 29.5 + 273.15
    h_in: float = 8.0              # W/m2K
    dynamic_h_in: bool = False     # keep False unless doing sensitivity


@dataclass
class NumericalParameters:
    """Numerical settings matching the reference paper style."""
    dt: float = 1.0                # s
    Nz_substrate: int = 107
    Nz_slab: int = 67
    save_every_s: int = 60


@dataclass
class CAMRsProfile:
    """
    Data-driven CAM stomatal conductance/resistance profile.

    Scientific meaning:
    - r_stoma_min_s_m is ONE minimum stomatal resistance, derived from the
      highest reliable LI-COR gsw values.
    - hourly_gsw_median is the measured/aggregated CAM conductance profile.
    - The model uses r_s(t) from hourly_gsw_median when available.
    - No hard CAM phase multiplier is forced by default. This matches the
      finding that CAM day/night 95th-percentile gsw may be similar, while
      median/integrated behavior can still differ.
    """
    r_stoma_min_s_m: float = 90.0
    hourly_gsw_median: Dict[int, float] = field(default_factory=dict)
    hourly_gsw_p25: Dict[int, float] = field(default_factory=dict)
    hourly_gsw_p75: Dict[int, float] = field(default_factory=dict)
    hourly_gsw_n: Dict[int, int] = field(default_factory=dict)
    fallback_gsw_mol_m2_s: float = 0.05
    r_s_min_limit_s_m: float = 40.0
    r_s_max_limit_s_m: float = 8000.0
    source_summary: str = "defaults; replace with LI-COR-derived profile when available"

    # Backward-compatible properties used only for old print statements.
    @property
    def r_s_night_s_m(self) -> float:
        vals = [gsw_to_r_s_s_m(v) for h, v in self.hourly_gsw_median.items() if h >= 19 or h <= 5]
        vals = [v for v in vals if np.isfinite(v)]
        return float(np.nanmedian(vals)) if vals else float(gsw_to_r_s_s_m(self.fallback_gsw_mol_m2_s))

    @property
    def r_s_midday_s_m(self) -> float:
        vals = [gsw_to_r_s_s_m(v) for h, v in self.hourly_gsw_median.items() if 8 <= h <= 14]
        vals = [v for v in vals if np.isfinite(v)]
        return float(np.nanmedian(vals)) if vals else float(gsw_to_r_s_s_m(self.fallback_gsw_mol_m2_s))

    @property
    def r_s_late_afternoon_s_m(self) -> float:
        vals = [gsw_to_r_s_s_m(v) for h, v in self.hourly_gsw_median.items() if 15 <= h <= 18]
        vals = [v for v in vals if np.isfinite(v)]
        return float(np.nanmedian(vals)) if vals else float(gsw_to_r_s_s_m(self.fallback_gsw_mol_m2_s))


# Default global instances. Runner may copy/update these.
bromelia = PlantParameters()
substrat = SubstrateParameters()
slab = SlabParameters()
geom = GeometryParameters()
num = NumericalParameters()
cam_rs_profile = CAMRsProfile()


# ==============================================================================
# 02 — GENERAL SMALL HELPERS
# ==============================================================================

def _to_float(x) -> float:
    if x is None:
        return np.nan
    if isinstance(x, str):
        x = x.strip()
        if x in {"", "---", "--", "nan", "NaN", "None"}:
            return np.nan
    return pd.to_numeric(x, errors="coerce")


def _clip(value, low, high):
    return float(np.clip(value, low, high))


def _safe_series(series: pd.Series, name: str) -> pd.Series:
    out = pd.to_numeric(series, errors="coerce")
    out.name = name
    return out


def gsw_to_r_s_s_m(gsw_mol_m2_s: Union[float, np.ndarray, pd.Series]) -> Union[float, np.ndarray, pd.Series]:
    """
    Convert LI-COR stomatal conductance gsw [mol m-2 s-1] to stomatal resistance [s/m].

    Conversion used in the previous workflow:
        r_s = 1 / (gsw * 0.0224)
    """
    gsw = np.asarray(gsw_mol_m2_s, dtype=float)
    out = np.full_like(gsw, np.nan, dtype=float)
    mask = np.isfinite(gsw) & (gsw > 0)
    out[mask] = 1.0 / (gsw[mask] * 0.0224)
    if np.ndim(gsw_mol_m2_s) == 0:
        return float(out)
    return out


def r_s_to_gsw_mol_m2_s(r_s_s_m: Union[float, np.ndarray, pd.Series]) -> Union[float, np.ndarray, pd.Series]:
    """Inverse of gsw_to_r_s_s_m()."""
    r_s = np.asarray(r_s_s_m, dtype=float)
    out = np.full_like(r_s, np.nan, dtype=float)
    mask = np.isfinite(r_s) & (r_s > 0)
    out[mask] = 1.0 / (r_s[mask] * 0.0224)
    if np.ndim(r_s_s_m) == 0:
        return float(out)
    return out


# ==============================================================================
# 03 — DATA LOADERS: WEATHER, NI, SOIL MOISTURE, LI-COR
# ==============================================================================

def load_weather_data(
    filepath: Union[str, Path],
    date_start: Optional[str] = None,
    date_end: Optional[str] = None,
    sheet_name: str = "3-24april",
) -> pd.DataFrame:
    """
    Load Davis weather data.

    Supports:
    1. Excel export: weatherfile mar-april.xlsx
    2. TXT/tab-separated WeatherLink export

    Output columns:
        T_a, RH, u, rain, G_sol, rain_flux
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Weather file not found: {filepath}")

    start = pd.Timestamp(date_start) if date_start else None
    end = pd.Timestamp(date_end) if date_end else None

    print(f"Loading weather data: {filepath}")

    if filepath.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            ts = pd.to_datetime(f"{row[0]} {row[1]}", dayfirst=True, errors="coerce")
            if pd.isna(ts):
                continue
            if start is not None and ts < start:
                continue
            if end is not None and ts > end:
                continue
            rows.append({
                "timestamp": ts,
                "T_a": _to_float(row[2]),
                "RH": _to_float(row[5]),
                "u": _to_float(row[7]),
                "rain": _to_float(row[17]),
                "G_sol": _to_float(row[19]),
            })
        df = pd.DataFrame(rows)
        if df.empty:
            raise ValueError(f"Weather data empty in requested period: {start} -> {end}")
        df = df.set_index("timestamp").sort_index()
    else:
        raw = pd.read_csv(filepath, sep="\t", skiprows=2, header=None,
                          na_values=["---", "  ---", " ---", "---  "], low_memory=False)
        df = pd.DataFrame({
            "date": raw.iloc[:, 0].astype(str).str.strip(),
            "time": raw.iloc[:, 1].astype(str).str.strip(),
            "T_a": pd.to_numeric(raw.iloc[:, 2], errors="coerce"),
            "RH": pd.to_numeric(raw.iloc[:, 5], errors="coerce"),
            "u": pd.to_numeric(raw.iloc[:, 7], errors="coerce"),
            "rain": pd.to_numeric(raw.iloc[:, 17], errors="coerce"),
            "G_sol": pd.to_numeric(raw.iloc[:, 19], errors="coerce"),
        })
        df["timestamp"] = pd.to_datetime(df["date"] + " " + df["time"],
                                          dayfirst=True, errors="coerce")
        df = df.dropna(subset=["timestamp"]).set_index("timestamp").drop(columns=["date", "time"]).sort_index()
        if start is not None:
            df = df[df.index >= start]
        if end is not None:
            df = df[df.index <= end]

    df = df[~df.index.duplicated(keep="first")]
    df["G_sol"] = _safe_series(df["G_sol"], "G_sol").fillna(0).clip(lower=0)
    df["u"] = _safe_series(df["u"], "u").clip(lower=0.1)
    df["RH"] = _safe_series(df["RH"], "RH").clip(0, 100)
    df["T_a"] = _safe_series(df["T_a"], "T_a")
    df["rain"] = _safe_series(df["rain"], "rain").fillna(0).clip(lower=0)

    for col in ["T_a", "RH", "u", "G_sol"]:
        df[col] = df[col].interpolate("time", limit=60, limit_direction="both")

    # Rain in WeatherLink is commonly mm/min at 1-min interval; 1 mm water = 1 kg/m2.
    df["rain_flux"] = df["rain"] / 60.0
    df = df.dropna(subset=["T_a", "RH", "u", "G_sol"])

    print(f"  Weather range : {df.index[0]} -> {df.index[-1]} | {len(df)} rows")
    return df


def load_NI_sensor_data(filepath: Union[str, Path]) -> pd.DataFrame:
    """
    Load NI/LabVIEW Excel file using XML parser.

    CAM channels used:
        T1Ka : indoor air / T_in_CAM
        T1Tb : indoor roof / T_s_in_CAM
        T1Ta : outdoor roof CAM
        T1Ke : soil top CAM
        T2A  : soil bottom CAM, with anomaly filtering
        T2A2 : optional target candidate
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"NI file not found: {filepath}")

    print(f"Loading NI sensor data: {filepath}")
    with zipfile.ZipFile(filepath, "r") as z:
        with z.open("xl/worksheets/sheet1.xml") as f:
            xml_content = f.read()

    tree = ET.fromstring(xml_content)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rows_xml = tree.findall(f".//{ns}row")

    names = [
        "timestamp_serial",
        "T1Kd", "T1Kb", "T1Ke", "T1Ka", "T1Kc",
        "T3Kb", "T3Kd", "T3Ke", "T3Kc", "T3Ka",
        "T2Ka", "T2Kd", "T2Kc", "2Ke", "T1A",
        "T2A", "T2A2", "T1Tb", "T1Ta", "T1Ta2",
    ]

    data = []
    for row in rows_xml[1:]:
        vals = []
        for cell in row.findall(f"{ns}c"):
            v = cell.find(f"{ns}v")
            vals.append(float(v.text) if v is not None else np.nan)
        if len(vals) == len(names):
            data.append(vals)

    df = pd.DataFrame(data, columns=names)
    labview_epoch = pd.Timestamp("1904-01-01")
    df["timestamp"] = labview_epoch + pd.to_timedelta(df["timestamp_serial"], unit="D")
    df = df.set_index("timestamp").drop(columns=["timestamp_serial"]).sort_index()

    # General temperature anomaly filtering.
    for col in [c for c in df.columns if c.startswith("T")]:
        mask_bad = (df[col] < -10) | (df[col] > 80)
        if mask_bad.sum() > 0:
            print(f"  Cleaning anomaly {col}: {int(mask_bad.sum())} points")
            df.loc[mask_bad, col] = np.nan
            df[col] = df[col].interpolate("time", limit=30, limit_direction="both")

    df["T_g_top_CAM"] = df["T1Ke"]
    df["T_g_bot_CAM"] = df["T2A"]
    df["T_s_in_CAM"] = df["T1Tb"]
    df["T_s_ext_CAM"] = df["T1Ta"]
    df["T_in_CAM"] = df["T1Ka"]

    print(f"  NI range      : {df.index[0]} -> {df.index[-1]} | {len(df)} rows")
    return df


def load_multiple_NI_sensor_data(filepaths: Sequence[Union[str, Path]]) -> pd.DataFrame:
    """Concatenate multiple NI files and remove duplicate timestamps."""
    dfs = []
    for fp in filepaths:
        fp = Path(fp)
        if fp.exists():
            dfs.append(load_NI_sensor_data(fp))
        else:
            print(f"WARNING: NI file not found, skipped: {fp}")
    if not dfs:
        raise ValueError("No NI files were loaded.")
    df = pd.concat(dfs).sort_index()
    df = df[~df.index.duplicated(keep="last")]
    print(f"Combined NI     : {df.index[0]} -> {df.index[-1]} | {len(df)} rows")
    return df


def load_soil_sensor_csv(
    filepath: Union[str, Path],
    timestamp_mode: str = "gmt_to_wib",
) -> pd.DataFrame:
    """
    Load one RIKA soil sensor CSV.

    timestamp_mode:
        gmt_to_wib : parse Timestamp as UTC/GMT then convert to Asia/Jakarta
        local      : parse Timestamp as already-local naive time
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Soil sensor file not found: {filepath}")

    df = read_csv_with_encoding_fallback(filepath, sep=";")
    if "Timestamp" not in df.columns:
        raise ValueError(f"No Timestamp column in {filepath}")

    if timestamp_mode == "gmt_to_wib":
        ts = pd.to_datetime(df["Timestamp"], utc=True, errors="coerce")
        ts = ts.dt.tz_convert("Asia/Jakarta").dt.tz_localize(None)
    elif timestamp_mode == "local":
        ts = pd.to_datetime(df["Timestamp"], errors="coerce")
    else:
        raise ValueError("timestamp_mode must be 'gmt_to_wib' or 'local'")

    moisture = pd.to_numeric(df.get("Moisture (%)"), errors="coerce")
    soil_temp = pd.to_numeric(df.get("Temperature"), errors="coerce")

    # Use arrays to avoid pandas index alignment bugs.
    out = pd.DataFrame({
        "timestamp": ts.to_numpy(),
        "moisture_pct": moisture.to_numpy(),
        "theta": (moisture / 100.0).to_numpy(),
        "soil_temp_C": soil_temp.to_numpy(),
    })
    out = out.dropna(subset=["timestamp"]).set_index("timestamp").sort_index()
    out = out[~out.index.duplicated(keep="last")]

    # Flag obviously impossible VWC values as NaN.
    bad = (out["theta"] < 0.03) | (out["theta"] > 0.95)
    if bad.sum() > 0:
        out.loc[bad, ["moisture_pct", "theta"]] = np.nan
        out[["moisture_pct", "theta"]] = out[["moisture_pct", "theta"]].interpolate(
            "time", limit=15, limit_direction="both"
        )
    return out


def load_cam_soil_moisture(
    sensor1_file: Union[str, Path],
    sensor2_file: Union[str, Path],
    timestamp_mode: str = "gmt_to_wib",
    swap_depths: bool = False,
) -> pd.DataFrame:
    """
    Load CAM soil moisture from two sensor CSVs.

    Default assumption:
        Sensor 1 COM5_CAM = shallow
        Sensor 2 COM6_CAM = deep
    """
    s1 = load_soil_sensor_csv(sensor1_file, timestamp_mode=timestamp_mode)
    s2 = load_soil_sensor_csv(sensor2_file, timestamp_mode=timestamp_mode)

    if not swap_depths:
        shallow, deep = s1, s2
        mapping = "S1=shallow, S2=deep"
    else:
        shallow, deep = s2, s1
        mapping = "S2=shallow, S1=deep"

    df = pd.concat([
        shallow.rename(columns={
            "moisture_pct": "theta_shallow_pct",
            "theta": "theta_shallow",
            "soil_temp_C": "soil_temp_shallow_C",
        }),
        deep.rename(columns={
            "moisture_pct": "theta_deep_pct",
            "theta": "theta_deep",
            "soil_temp_C": "soil_temp_deep_C",
        }),
    ], axis=1).sort_index()

    print(f"CAM soil mapping: {mapping}")
    print(f"  Soil range    : {df.index[0]} -> {df.index[-1]} | {len(df)} rows")
    return df


def _scan_table_for_header(df_raw: pd.DataFrame, required_hint: str = "gsw") -> Optional[pd.DataFrame]:
    """Find a row containing required_hint and promote it to header."""
    hint = required_hint.lower()
    max_rows = min(len(df_raw), 80)
    for i in range(max_rows):
        row = df_raw.iloc[i].astype(str).str.strip().str.lower()
        if any(cell == hint for cell in row):
            header = df_raw.iloc[i].astype(str).str.strip().to_list()
            data = df_raw.iloc[i+1:].copy()
            data.columns = header
            # Drop likely unit row if gsw cell is nonnumeric after header.
            if "gsw" in data.columns:
                first_gsw = pd.to_numeric(pd.Series(data["gsw"]).head(3), errors="coerce")
                if first_gsw.isna().iloc[0]:
                    data = data.iloc[1:].copy()
            return data.reset_index(drop=True)
    return None


def load_licor_file(filepath: Union[str, Path]) -> pd.DataFrame:
    """
    Robust LI-COR loader for CSV/XLSX.

    It scans for a header row containing 'gsw', then converts useful columns.
    CSV exports from the LI-6800 often contain a long tab-separated preamble,
    so text files are parsed manually before falling back to pandas.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"LI-COR file not found: {filepath}")

    candidates = []
    if filepath.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        xls = pd.ExcelFile(filepath)
        for sheet in xls.sheet_names:
            try:
                raw = pd.read_excel(filepath, sheet_name=sheet, header=None)
                table = _scan_table_for_header(raw, required_hint="gsw")
                if table is not None:
                    table["_sheet"] = sheet
                    candidates.append(table)
            except Exception:
                continue
    else:
        # Preferred path for LI-COR CSV/TXT: manual tab parsing after the preamble.
        try:
            text = filepath.read_text(encoding="utf-8", errors="ignore")
            lines = text.splitlines()
            header_i = None
            for i, line in enumerate(lines):
                parts = [p.strip() for p in line.split("\t")]
                if "gsw" in [p.lower() for p in parts]:
                    header_i = i
                    break
            if header_i is not None:
                header = [p.strip() for p in lines[header_i].split("\t")]
                data_rows = []
                for line in lines[header_i + 1:]:
                    if not line.strip():
                        continue
                    parts = [p.strip() for p in line.split("\t")]
                    # skip units row and malformed metadata lines
                    if len(parts) < len(header) * 0.5:
                        continue
                    if parts[0].lower() in {"", "obs"}:
                        continue
                    if len(parts) < len(header):
                        parts = parts + [""] * (len(header) - len(parts))
                    data_rows.append(parts[:len(header)])
                if data_rows:
                    candidates.append(pd.DataFrame(data_rows, columns=header))
        except Exception:
            pass

        # Fallbacks for simpler CSVs.
        if not candidates:
            for sep in ["\t", ",", ";"]:
                try:
                    raw = read_csv_with_encoding_fallback(filepath, header=None, sep=sep, engine="python", on_bad_lines="skip")
                    table = _scan_table_for_header(raw, required_hint="gsw")
                    if table is not None:
                        candidates.append(table)
                        break
                except Exception:
                    continue
        if not candidates:
            try:
                table = read_csv_with_encoding_fallback(filepath)
                if any(str(c).strip().lower() == "gsw" for c in table.columns):
                    candidates.append(table)
            except Exception:
                pass

    if not candidates:
        raise ValueError(f"Could not find LI-COR table with gsw column in {filepath}")

    df = pd.concat(candidates, ignore_index=True)
    df.columns = [str(c).strip() for c in df.columns]
    for col in ["A", "E", "gsw", "TleafEB", "VPDleaf", "Qin", "Rabs", "Tair"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Attach session start from filename if possible.
    m = re.search(r"(20\d{2})-(\d{2})-(\d{2})-(\d{4})", filepath.name)
    if m:
        y, mo, d, hm = m.groups()
        hour, minute = int(hm[:2]), int(hm[2:])
        session_start = pd.Timestamp(year=int(y), month=int(mo), day=int(d), hour=hour, minute=minute)
    else:
        session_start = pd.NaT
    df["session_start"] = session_start
    df["source_file"] = filepath.name
    return df


def summarize_licor_gsw(filepath: Union[str, Path]) -> Dict[str, float]:
    """Return a compact summary for one LI-COR file."""
    df = load_licor_file(filepath)
    gsw = pd.to_numeric(df.get("gsw"), errors="coerce")
    gsw_valid = gsw[(gsw > 0.001) & (gsw < 5.0)]
    if len(gsw_valid) == 0:
        return {"file": Path(filepath).name, "n": 0}
    r_s = gsw_to_r_s_s_m(gsw_valid.to_numpy())
    return {
        "file": Path(filepath).name,
        "n": int(len(gsw_valid)),
        "gsw_mean": float(gsw_valid.mean()),
        "gsw_p05": float(np.percentile(gsw_valid, 5)),
        "gsw_p50": float(np.percentile(gsw_valid, 50)),
        "gsw_p95": float(np.percentile(gsw_valid, 95)),
        "r_s_mean_s_m": float(np.nanmean(r_s)),
        "r_s_p05_s_m": float(np.nanpercentile(r_s, 5)),
        "r_s_p50_s_m": float(np.nanpercentile(r_s, 50)),
        "r_s_p95_s_m": float(np.nanpercentile(r_s, 95)),
    }


def _session_hour_from_licor_dataframe(df: pd.DataFrame) -> Optional[int]:
    """Return the session hour from parsed LI-COR metadata/filename."""
    if "session_start" in df.columns:
        vals = df["session_start"].dropna()
        if len(vals):
            ts = pd.Timestamp(vals.iloc[0])
            return int(ts.hour)
    return None


def _clean_licor_gsw_series(
    gsw: pd.Series,
    min_positive: float = 0.0005,
    max_reasonable: float = 3.0,
) -> pd.Series:
    """
    Clean LI-COR gsw for resistance modeling.

    Negative gsw is treated as invalid measurement/noise, not real conductance.
    Very small positive values are retained because CAM daytime closure can be near zero.
    """
    out = pd.to_numeric(gsw, errors="coerce")
    out = out.where((out > min_positive) & (out < max_reasonable), np.nan)
    return out


def build_cam_rs_profile_from_licor(
    licor_files: Sequence[Union[str, Path]],
    fallback: CAMRsProfile = CAMRsProfile(),
    aggregation: str = "median",
) -> CAMRsProfile:
    """
    Build a data-driven hourly CAM gsw/r_s profile from one or more LI-COR files.

    Important modeling decision:
    - We derive ONE r_stoma_min from the robust high-end conductance data.
    - We derive r_s(t) from the hourly LI-COR gsw profile.
    - We do not impose night=open and day=closed as fixed multipliers.

    Hour assignment currently uses the timestamp in the filename/session metadata.
    This is appropriate for your non-continuous LI-COR sessions, where each file
    represents a measurement block rather than 24/7 logging.
    """
    rows = []
    summaries = []

    for fp in licor_files:
        fp = Path(fp)
        if not fp.exists():
            summaries.append(f"{fp.name}: file not found")
            continue
        try:
            df = load_licor_file(fp)
            if "gsw" not in df.columns:
                summaries.append(f"{fp.name}: no gsw column")
                continue

            hour = _session_hour_from_licor_dataframe(df)
            if hour is None:
                summaries.append(f"{fp.name}: usable table but no session hour")
                continue

            gsw_clean = _clean_licor_gsw_series(df["gsw"])
            valid = gsw_clean.dropna()
            if len(valid) == 0:
                summaries.append(f"{fp.name}: no valid positive gsw after cleaning")
                continue

            for val in valid:
                rows.append({"hour": int(hour), "gsw": float(val), "source_file": fp.name})

            rs_vals = gsw_to_r_s_s_m(valid.to_numpy())
            summaries.append(
                f"{fp.name}: hour={hour:02d}, n={len(valid)}, "
                f"gsw_med={valid.median():.4f}, gsw_p95={np.nanpercentile(valid,95):.4f}, "
                f"r_s_med={np.nanmedian(rs_vals):.1f} s/m"
            )
        except Exception as exc:
            summaries.append(f"{fp.name}: skipped ({exc})")

    if not rows:
        return CAMRsProfile(
            r_stoma_min_s_m=fallback.r_stoma_min_s_m,
            hourly_gsw_median=fallback.hourly_gsw_median,
            hourly_gsw_p25=fallback.hourly_gsw_p25,
            hourly_gsw_p75=fallback.hourly_gsw_p75,
            hourly_gsw_n=fallback.hourly_gsw_n,
            fallback_gsw_mol_m2_s=fallback.fallback_gsw_mol_m2_s,
            r_s_min_limit_s_m=fallback.r_s_min_limit_s_m,
            r_s_max_limit_s_m=fallback.r_s_max_limit_s_m,
            source_summary="No usable LI-COR gsw files; using fallback profile.\n" + "\n".join(summaries),
        )

    all_df = pd.DataFrame(rows)
    grouped = all_df.groupby("hour")["gsw"]
    hourly_med = grouped.median().to_dict()
    hourly_p25 = grouped.quantile(0.25).to_dict()
    hourly_p75 = grouped.quantile(0.75).to_dict()
    hourly_n = grouped.count().astype(int).to_dict()

    # Robust minimum resistance from high-end gsw. Avoid using a single absolute max.
    gsw_all = all_df["gsw"].to_numpy(dtype=float)
    gsw_high = float(np.nanpercentile(gsw_all, 95))
    r_stoma_min = float(gsw_to_r_s_s_m(gsw_high))
    r_stoma_min = float(np.clip(r_stoma_min, fallback.r_s_min_limit_s_m, fallback.r_s_max_limit_s_m))

    fallback_gsw = float(np.nanmedian(gsw_all)) if np.isfinite(np.nanmedian(gsw_all)) else fallback.fallback_gsw_mol_m2_s

    # Interpolate missing hours circularly so simulation always has an r_s value.
    hourly_med_full = fill_hourly_gsw_profile(hourly_med, fallback_value=fallback_gsw)

    phase_note = compare_day_night_gsw_distribution(all_df)
    summaries.append(phase_note)
    summaries.append(f"r_stoma_min_CAM from 95th percentile gsw ({gsw_high:.4f}) = {r_stoma_min:.1f} s/m")

    return CAMRsProfile(
        r_stoma_min_s_m=r_stoma_min,
        hourly_gsw_median={int(k): float(v) for k, v in hourly_med_full.items()},
        hourly_gsw_p25={int(k): float(v) for k, v in hourly_p25.items()},
        hourly_gsw_p75={int(k): float(v) for k, v in hourly_p75.items()},
        hourly_gsw_n={int(k): int(v) for k, v in hourly_n.items()},
        fallback_gsw_mol_m2_s=fallback_gsw,
        r_s_min_limit_s_m=fallback.r_s_min_limit_s_m,
        r_s_max_limit_s_m=fallback.r_s_max_limit_s_m,
        source_summary="\n".join(summaries),
    )


def fill_hourly_gsw_profile(hourly_gsw: Dict[int, float], fallback_value: float = 0.05) -> Dict[int, float]:
    """Fill missing 0..23 hourly gsw values using circular interpolation."""
    clean = {int(k): float(v) for k, v in hourly_gsw.items() if np.isfinite(v) and v > 0}
    if not clean:
        return {h: float(fallback_value) for h in range(24)}

    x = np.array(sorted(clean.keys()), dtype=float)
    y = np.array([clean[int(h)] for h in x], dtype=float)
    if len(x) == 1:
        return {h: float(y[0]) for h in range(24)}

    # Circular extension for interpolation across midnight.
    x_ext = np.r_[x - 24, x, x + 24]
    y_ext = np.r_[y, y, y]
    full = {}
    for h in range(24):
        full[h] = float(np.interp(h, x_ext, y_ext))
    return full


def load_hourly_gsw_profile_csv(filepath: Union[str, Path]) -> Dict[int, float]:
    """
    Load manual/hourly CAM gsw profile from CSV/XLSX.

    Accepted columns:
        hour, gsw
        hour, gsw_median
        hour, gsw_mol_m2_s

    Example:
        hour,gsw
        10,0.03
        16,0.2261
        21,0.444
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(filepath)
    if filepath.suffix.lower() in {".xlsx", ".xls", ".xlsm"}:
        df = pd.read_excel(filepath)
    else:
        df = read_csv_with_encoding_fallback(filepath)
    cols = {str(c).strip().lower(): c for c in df.columns}
    if "hour" not in cols:
        raise ValueError("hourly gsw profile must contain an 'hour' column")
    gsw_col = None
    for candidate in ["gsw", "gsw_median", "gsw_mol_m2_s", "gsw_mean"]:
        if candidate in cols:
            gsw_col = cols[candidate]
            break
    if gsw_col is None:
        raise ValueError("hourly gsw profile must contain gsw/gsw_median/gsw_mol_m2_s")
    out = {}
    for _, row in df.iterrows():
        h = int(row[cols["hour"]])
        g = float(row[gsw_col])
        if 0 <= h <= 23 and np.isfinite(g) and g > 0:
            out[h] = g
    if not out:
        raise ValueError("hourly gsw profile contains no valid positive values")
    return fill_hourly_gsw_profile(out, fallback_value=float(np.nanmedian(list(out.values()))))


def profile_from_manual_hourly_gsw(
    pairs: Sequence[str],
    base_profile: Optional[CAMRsProfile] = None,
) -> CAMRsProfile:
    """
    Update/create profile from strings like '10:0.03' or '16=0.2261'.
    Missing hours are circularly interpolated.
    """
    profile = base_profile or CAMRsProfile()
    manual = dict(profile.hourly_gsw_median)
    for item in pairs:
        if not item:
            continue
        if ":" in item:
            h_s, g_s = item.split(":", 1)
        elif "=" in item:
            h_s, g_s = item.split("=", 1)
        else:
            raise ValueError(f"manual gsw pair must be hour:gsw, got {item!r}")
        h = int(float(h_s.strip()))
        g = float(g_s.strip())
        if not (0 <= h <= 23):
            raise ValueError(f"hour out of range in {item!r}")
        if not (np.isfinite(g) and g > 0):
            raise ValueError(f"gsw must be positive in {item!r}")
        manual[h] = g
    filled = fill_hourly_gsw_profile(manual, fallback_value=profile.fallback_gsw_mol_m2_s)
    gvals = np.array(list(filled.values()), dtype=float)
    rmin = float(gsw_to_r_s_s_m(np.nanpercentile(gvals, 95)))
    rmin = float(np.clip(rmin, profile.r_s_min_limit_s_m, profile.r_s_max_limit_s_m))
    return CAMRsProfile(
        r_stoma_min_s_m=rmin,
        hourly_gsw_median=filled,
        hourly_gsw_p25=profile.hourly_gsw_p25,
        hourly_gsw_p75=profile.hourly_gsw_p75,
        hourly_gsw_n=profile.hourly_gsw_n,
        fallback_gsw_mol_m2_s=profile.fallback_gsw_mol_m2_s,
        r_s_min_limit_s_m=profile.r_s_min_limit_s_m,
        r_s_max_limit_s_m=profile.r_s_max_limit_s_m,
        source_summary=profile.source_summary + "\nmanual hourly gsw overrides: " + ", ".join(pairs),
    )


def compare_day_night_gsw_distribution(all_df: pd.DataFrame) -> str:
    """Summarize whether CAM day/night upper tails are actually different."""
    if all_df.empty:
        return "day/night gsw comparison: no data"
    df = all_df.copy()
    df["period"] = np.where((df["hour"] >= 19) | (df["hour"] <= 5), "night", "day")
    lines = ["day/night gsw comparison:"]
    stats = {}
    for period, grp in df.groupby("period"):
        vals = grp["gsw"].dropna().to_numpy(dtype=float)
        if len(vals) == 0:
            continue
        stats[period] = {
            "n": len(vals),
            "median": float(np.nanmedian(vals)),
            "p95": float(np.nanpercentile(vals, 95)),
            "mean": float(np.nanmean(vals)),
        }
        lines.append(
            f"  {period}: n={len(vals)}, mean={stats[period]['mean']:.4f}, "
            f"median={stats[period]['median']:.4f}, p95={stats[period]['p95']:.4f}"
        )
    if "day" in stats and "night" in stats:
        ratio = stats["day"]["p95"] / max(stats["night"]["p95"], 1e-9)
        lines.append(f"  p95 day/night ratio = {ratio:.2f}; default model uses hourly data, not hard phase multiplier.")
    return "\n".join(lines)


# ==============================================================================
# 04 — PRE-PROCESSING AND VALIDATION DATASET BUILDER
# ==============================================================================

def prepare_weather_1s(weather_df: pd.DataFrame, dt: float = 1.0) -> pd.DataFrame:
    """Resample weather to dt seconds; rain flux is forward-filled, not linearly interpolated."""
    rule = f"{int(dt)}s"
    states = weather_df[["T_a", "RH", "u", "G_sol"]].resample(rule).interpolate("time")
    rain = weather_df["rain_flux"].fillna(0).resample(rule).ffill().fillna(0)
    states["rain_flux"] = rain
    return states


def align_series_to_index(series: pd.Series, index: pd.DatetimeIndex, dt: float = 1.0) -> pd.Series:
    """Resample a series and align to the simulation index."""
    rule = f"{int(dt)}s"
    s = series.sort_index().resample(rule).interpolate("time")
    s = s.reindex(index).interpolate("time").ffill().bfill()
    return s


def make_theta_initial_profile(
    theta_initial: Optional[Union[float, Sequence[float], Dict[str, Sequence[float]]]],
    H_g: float,
    Nz: int,
    substrate: SubstrateParameters,
) -> np.ndarray:
    """Create initial VWC profile."""
    z = np.linspace(0, H_g, Nz)
    if theta_initial is None:
        return np.full(Nz, substrate.theta_sat * 0.80)
    if isinstance(theta_initial, dict):
        depths = np.asarray(theta_initial.get("depths", [0.02, 0.07]), dtype=float)
        values = np.asarray(theta_initial.get("values"), dtype=float)
        return np.clip(np.interp(z, depths, values, left=values[0], right=values[-1]),
                       substrate.theta_min, substrate.theta_sat)
    if isinstance(theta_initial, (list, tuple, np.ndarray)) and len(theta_initial) == 2:
        values = np.asarray(theta_initial, dtype=float)
        return np.clip(np.interp(z, [0.02, 0.07], values, left=values[0], right=values[-1]),
                       substrate.theta_min, substrate.theta_sat)
    return np.full(Nz, float(theta_initial))


def get_theta_initial_from_soil(soil_df: Optional[pd.DataFrame], start_time: pd.Timestamp,
                                substrate: SubstrateParameters) -> Optional[Dict[str, Sequence[float]]]:
    """Get initial shallow/deep theta near start_time."""
    if soil_df is None or soil_df.empty:
        return None
    r = soil_df[["theta_shallow", "theta_deep"]].resample("1min").mean().interpolate("time", limit=10)
    window = r[(r.index >= start_time - pd.Timedelta("15min")) &
               (r.index <= start_time + pd.Timedelta("15min"))].dropna()
    if window.empty:
        window = r[r.index >= start_time].dropna().head(30)
    if window.empty:
        return None
    vals = window.iloc[0]
    return {
        "depths": [0.02, 0.07],
        "values": [
            float(np.clip(vals["theta_shallow"], substrate.theta_min, substrate.theta_sat)),
            float(np.clip(vals["theta_deep"], substrate.theta_min, substrate.theta_sat)),
        ],
    }


# ==============================================================================
# 05 — THERMODYNAMIC HELPERS
# ==============================================================================

def saturation_pressure(T_K: float) -> float:
    """Magnus saturation vapor pressure [Pa]."""
    T_C = T_K - 273.15
    return float(610.78 * np.exp(17.269 * T_C / (T_C + 237.29)))


def dew_point_C(T_C: float, RH: float) -> float:
    """Dew point [C] via Magnus approximation."""
    RH = _clip(RH, 1.0, 100.0)
    a, b = 17.27, 237.7
    gamma = (a * T_C) / (b + T_C) + np.log(RH / 100.0)
    return float((b * gamma) / (a - gamma))


def sky_temperature(T_a_K: float, RH: float) -> float:
    """Effective sky temperature [K], using dew-point-based sky emissivity."""
    T_C = T_a_K - 273.15
    Tdp = dew_point_C(T_C, RH)
    epsilon_sky = 0.711 + 0.56 * (Tdp / 100.0) + 0.73 * (Tdp / 100.0) ** 2
    epsilon_sky = _clip(epsilon_sky, 0.60, 1.00)
    return float(epsilon_sky ** 0.25 * T_a_K)


def ambient_vapor_pressure(T_a_K: float, RH: float) -> float:
    """Actual ambient vapor pressure [Pa]."""
    return float((RH / 100.0) * saturation_pressure(T_a_K))


def psychrometric_constant() -> float:
    """Psychrometric constant gamma [Pa/K]."""
    return float(1005.0 * 101325.0 / (0.622 * 2.45e6))


def interior_ceiling_convective_h(T_surface_K: float, T_air_K: float) -> float:
    """Natural-convection fallback for underside ceiling; keep bounded."""
    dT = float(T_surface_K - T_air_K)
    dT_abs = max(abs(dT), 0.1)
    a = 0.76 if dT > 0 else 1.52
    return float(np.clip(a * dT_abs ** (1.0 / 3.0), 0.5, 4.0))


# ==============================================================================
# 06 — CAM STOMATAL RESISTANCE MODEL FROM LI-COR
# ==============================================================================

def hourly_profile_gsw(timestamp: pd.Timestamp, profile: CAMRsProfile) -> float:
    """Return hourly median gsw from the LI-COR-derived profile."""
    hour = int(pd.Timestamp(timestamp).hour)
    if profile.hourly_gsw_median:
        val = profile.hourly_gsw_median.get(hour, np.nan)
        if np.isfinite(val) and val > 0:
            return float(val)
    return float(profile.fallback_gsw_mol_m2_s)


def cam_profile_r_s(timestamp: pd.Timestamp, profile: CAMRsProfile) -> float:
    """Convert the hourly gsw profile into stomatal resistance r_s [s/m]."""
    gsw = hourly_profile_gsw(timestamp, profile)
    r_s = float(gsw_to_r_s_s_m(gsw))
    r_s = float(np.clip(r_s, profile.r_s_min_limit_s_m, profile.r_s_max_limit_s_m))
    return r_s


def compute_aerodynamic_resistance(u: float, plant: PlantParameters) -> float:
    """Aerodynamic resistance r_a [s/m], Eq. A.2-A.4 style."""
    k = 0.41
    d0 = 0.701 * plant.H_f ** 0.975
    Z0 = 0.131 * plant.H_f ** 0.997
    z_ref = plant.H_f + 2.0
    ratio = max((z_ref - d0) / max(Z0, 1e-6), 1.01)
    r_a = (np.log(ratio) ** 2) / (k ** 2 * max(u, 0.1))
    return float(max(r_a, 5.0))


def soil_moisture_stress(theta_avg: float, substrate: SubstrateParameters) -> float:
    """
    Soil moisture stress multiplier for r_s.

    This keeps LI-COR-derived phase behavior as the primary driver, while allowing
    drier substrate to increase stomatal resistance. It is bounded to avoid runaway.
    """
    theta_safe = max(float(theta_avg), substrate.theta_min + 1e-4)
    stress = (substrate.theta_sat - substrate.theta_min) / max(theta_safe - substrate.theta_min, 1e-4)
    return float(np.clip(stress, 1.0, 6.0))


def compute_stomatal_resistance_cam(
    timestamp: pd.Timestamp,
    theta_avg: float,
    plant: PlantParameters,
    substrate: SubstrateParameters,
    rs_profile: CAMRsProfile,
) -> Tuple[float, float, str]:
    """
    Actual CAM stomatal resistance r_s [s/m].

    Default behavior is data-driven:
        hourly LI-COR gsw profile -> r_s(t)

    A bounded soil-moisture stress can increase r_s when substrate is dry.
    No forced night/day CAM phase multiplier is applied here.
    """
    base_rs = cam_profile_r_s(timestamp, rs_profile)
    f_theta = soil_moisture_stress(theta_avg, substrate)
    r_s = float(np.clip(base_rs * f_theta, rs_profile.r_s_min_limit_s_m, rs_profile.r_s_max_limit_s_m))

    h = timestamp.hour + timestamp.minute / 60.0
    if h >= 19.0 or h < 5.5:
        label = "night_profile"
    elif 5.5 <= h < 11.5:
        label = "morning_profile"
    elif 11.5 <= h < 15.5:
        label = "midday_profile"
    elif 15.5 <= h < 18.5:
        label = "late_afternoon_profile"
    else:
        label = "evening_profile"

    gsw_equiv = r_s_to_gsw_mol_m2_s(r_s)
    return float(r_s), float(gsw_equiv), label


# ==============================================================================
# 07 — FOLIAGE MODEL: Eq. (1)–(3), A.1, A.5
# ==============================================================================

def solve_foliage_temperature(
    T_f_prev: float,
    T_a_K: float,
    T_g_surface: float,
    G_sol: float,
    RH: float,
    u: float,
    theta_avg: float,
    timestamp: pd.Timestamp,
    plant: PlantParameters,
    substrate: SubstrateParameters,
    rs_profile: CAMRsProfile,
    dt: float,
) -> Tuple[float, Dict[str, float]]:
    """Solve foliage temperature using implicit linearized energy balance."""
    sigma = 5.67e-8
    rho_air = 1.2
    cp_air = 1005.0
    gamma = psychrometric_constant()

    T_sky = sky_temperature(T_a_K, RH)
    P_a = ambient_vapor_pressure(T_a_K, RH)
    P_f_sat = saturation_pressure(T_f_prev)
    VPD = max(P_f_sat - P_a, 0.0)

    eps_fg = 1.0 / (1.0 / plant.epsilon_f + 1.0 / substrate.epsilon_g - 1.0)
    R_f_net = (
        plant.alpha_f * G_sol
        - plant.epsilon_f * sigma * (T_f_prev ** 4 - T_sky ** 4)
        - eps_fg * sigma * (T_f_prev ** 4 - T_g_surface ** 4)
    )

    r_a = compute_aerodynamic_resistance(u, plant)
    r_s, gsw_equiv, phase = compute_stomatal_resistance_cam(
        timestamp, theta_avg, plant, substrate, rs_profile
    )

    h_conv_f = plant.LAI * (rho_air * cp_air) / r_a
    h_eva_f = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_s)

    rho_cp_f = rho_air * cp_air
    d_eff = 0.001
    dPdT = saturation_pressure(T_f_prev + 0.5) - saturation_pressure(T_f_prev - 0.5)

    A = rho_cp_f * d_eff / dt + h_conv_f + h_eva_f * dPdT
    B = (
        rho_cp_f * d_eff / dt * T_f_prev
        + R_f_net
        + h_conv_f * T_a_K
        + h_eva_f * (P_a - P_f_sat + dPdT * T_f_prev)
    )

    T_f_new = B / max(A, 1e-12)
    diag = {
        "r_a_s_m": float(r_a),
        "r_s_s_m": float(r_s),
        "gsw_equiv_mol_m2_s": float(gsw_equiv),
        "cam_phase": phase,
        "VPD_Pa": float(VPD),
        "h_conv_f": float(h_conv_f),
        "h_eva_f": float(h_eva_f),
        "R_f_net": float(R_f_net),
    }
    return float(T_f_new), diag


# ==============================================================================
# 08 — SUBSTRATE MODEL: heat, moisture, ET separation
# ==============================================================================

def compute_substrate_properties(theta: np.ndarray, substrate: SubstrateParameters) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Moisture-dependent substrate properties."""
    theta = np.asarray(theta, dtype=float)
    theta_safe = np.maximum(theta, substrate.theta_min + 1e-6)
    S = np.clip(theta_safe / substrate.theta_sat, 0.0, 1.0)

    rho_cp_g = substrate.cp_g * (0.2 + theta_safe) * substrate.rho_g
    k_theta = substrate.k_theta_sat * S ** (2 * substrate.b + 3)
    lambda_sat = substrate.lambda_sat if substrate.lambda_sat is not None else substrate.lambda_dry + substrate.theta_sat * substrate.lambda_water
    lambda_g = substrate.lambda_dry + S * (lambda_sat - substrate.lambda_dry)
    D_theta = substrate.b * substrate.k_theta_sat * abs(substrate.psi_sat) / theta_safe * S ** (substrate.b + 3)
    return rho_cp_g, lambda_g, k_theta, D_theta


def substrate_vapor_resistance(theta_surface: float, substrate: SubstrateParameters) -> float:
    """Substrate surface vapor resistance r_vap [s/m], moisture dependent."""
    S = float(np.clip(theta_surface / substrate.theta_sat, 0.05, 1.0))
    return float(100.0 * S ** (-2.0))


def tdma_solver(a: np.ndarray, b: np.ndarray, c: np.ndarray, d: np.ndarray) -> np.ndarray:
    """Tridiagonal Matrix Algorithm. a,b,c are lower/main/upper diagonals."""
    n = len(d)
    cp = np.zeros(n)
    dp = np.zeros(n)
    x = np.zeros(n)

    denom0 = b[0] if abs(b[0]) > 1e-30 else 1e-30
    cp[0] = c[0] / denom0
    dp[0] = d[0] / denom0
    for i in range(1, n):
        denom = b[i] - a[i] * cp[i - 1]
        if abs(denom) < 1e-30:
            denom = 1e-30
        cp[i] = c[i] / denom
        dp[i] = (d[i] - a[i] * dp[i - 1]) / denom
    x[-1] = dp[-1]
    for i in range(n - 2, -1, -1):
        x[i] = dp[i] - cp[i] * x[i + 1]
    return x


def solve_substrate_heat(
    T_g: np.ndarray,
    T_f: float,
    T_slab_top: float,
    G_sol: float,
    T_a_K: float,
    RH: float,
    u: float,
    theta: np.ndarray,
    plant: PlantParameters,
    substrate: SubstrateParameters,
    H_g: float,
    dt: float,
) -> np.ndarray:
    """Solve substrate temperature profile with FVM + TDMA."""
    sigma = 5.67e-8
    rho_air = 1.2
    cp_air = 1005.0
    gamma = psychrometric_constant()

    Nz = len(T_g)
    dz = H_g / (Nz - 1)
    rho_cp_g, lambda_g, _, _ = compute_substrate_properties(theta, substrate)

    P_a = ambient_vapor_pressure(T_a_K, RH)
    P_g_sat = saturation_pressure(T_g[0])
    eps_fg = 1.0 / (1.0 / plant.epsilon_f + 1.0 / substrate.epsilon_g - 1.0)

    # Real boxes may not be fully covered. Exposed fraction receives direct solar.
    cover = float(np.clip(plant.cover_fraction, 0.0, 1.0))
    G_to_substrate = ((1.0 - cover) + cover * plant.tau_f) * G_sol

    R_g_net = (1.0 - substrate.rho_g_rad) * G_to_substrate + eps_fg * sigma * (T_f ** 4 - T_g[0] ** 4)

    d0 = 0.701 * plant.H_f ** 0.975
    Z_M = 0.131 * plant.H_f ** 0.997
    Z_u = plant.H_f + 2.0
    log_r = np.log(max((plant.H_f - d0) / max(Z_M, 1e-6), 1.01)) / np.log(max((Z_u - d0) / max(Z_M, 1e-6), 1.01))
    u_f = u * max(log_r, 0.01)
    a_drag = (0.28 * plant.LAI * plant.H_f * plant.d_f) ** 0.5
    u_g = max(u_f * np.exp(-a_drag * (1.0 - 0.05 / max(plant.H_f, 0.01))), 0.01)

    r_c = 1.0 / (0.004 + 0.012 * u_g)
    r_vap = substrate_vapor_resistance(float(theta[0]), substrate)
    r_a = compute_aerodynamic_resistance(u, plant)

    h_conv_g = plant.LAI * (rho_air * cp_air) / (r_a + r_c)
    h_eva_g = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_vap)
    q_eva_g = h_eva_g * max(P_g_sat - P_a, 0.0)

    q_top = R_g_net - h_conv_g * (T_g[0] - T_a_K) - q_eva_g

    aw = np.zeros(Nz); ap = np.zeros(Nz); ae = np.zeros(Nz); bv = np.zeros(Nz)
    for i in range(1, Nz - 1):
        lw = 0.5 * (lambda_g[i - 1] + lambda_g[i])
        le = 0.5 * (lambda_g[i] + lambda_g[i + 1])
        aw[i] = lw / dz ** 2
        ae[i] = le / dz ** 2
        ap[i] = rho_cp_g[i] / dt + aw[i] + ae[i]
        bv[i] = rho_cp_g[i] / dt * T_g[i]

    ae[0] = lambda_g[0] / dz ** 2
    ap[0] = rho_cp_g[0] / dt + ae[0]
    bv[0] = rho_cp_g[0] / dt * T_g[0] + q_top / dz

    # Dirichlet-like coupling to slab top boundary.
    lam_bc = lambda_g[-1] / dz ** 2
    aw[-1] = lambda_g[-1] / dz ** 2
    ap[-1] = rho_cp_g[-1] / dt + aw[-1] + lam_bc
    bv[-1] = rho_cp_g[-1] / dt * T_g[-1] + lam_bc * T_slab_top

    return tdma_solver(-aw, ap, -ae, bv)


def solve_substrate_moisture(
    theta: np.ndarray,
    T_f: float,
    T_g_surface: float,
    T_a_K: float,
    RH: float,
    u: float,
    j_irrigation: float,
    r_s_s_m: float,
    plant: PlantParameters,
    substrate: SubstrateParameters,
    H_g: float,
    dt: float,
) -> Tuple[np.ndarray, Dict[str, float]]:
    """Solve moisture transport and explicitly return foliage/substrate ET."""
    gamma = psychrometric_constant()
    rho_air = 1.2
    cp_air = 1005.0
    rho_water = 1000.0

    Nz = len(theta)
    dz = H_g / (Nz - 1)
    _, _, k_theta, D_theta = compute_substrate_properties(theta, substrate)

    P_a = ambient_vapor_pressure(T_a_K, RH)
    P_f_sat = saturation_pressure(T_f)
    P_g_sat = saturation_pressure(T_g_surface)
    r_a = compute_aerodynamic_resistance(u, plant)
    r_vap = substrate_vapor_resistance(float(theta[0]), substrate)

    h_eva_f = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_s_s_m)
    h_eva_g = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_vap)

    j_eva_f = max(0.0, h_eva_f * (P_f_sat - P_a) / substrate.l_fg)
    j_eva_g = max(0.0, h_eva_g * (P_g_sat - P_a) / substrate.l_fg)
    j_eva_total = j_eva_f + j_eva_g

    j_eva_ms = j_eva_total / rho_water
    j_irrig_ms = j_irrigation / rho_water
    j_net = j_irrig_ms - j_eva_ms

    aw = np.zeros(Nz); ap = np.zeros(Nz); ae = np.zeros(Nz); bv = np.zeros(Nz)
    for i in range(1, Nz - 1):
        Dw = 0.5 * (D_theta[i - 1] + D_theta[i])
        De = 0.5 * (D_theta[i] + D_theta[i + 1])
        kw = 0.5 * (k_theta[i - 1] + k_theta[i])
        ke = 0.5 * (k_theta[i] + k_theta[i + 1])
        aw[i] = Dw / dz ** 2
        ae[i] = De / dz ** 2
        ap[i] = 1.0 / dt + aw[i] + ae[i]
        bv[i] = theta[i] / dt + (ke - kw) / dz

    ae[0] = D_theta[0] / dz ** 2
    ap[0] = 1.0 / dt + ae[0]
    bv[0] = theta[0] / dt + (j_net - k_theta[0]) / dz

    # Free drainage at bottom.
    aw[-1] = D_theta[-1] / dz ** 2
    ap[-1] = 1.0 / dt + aw[-1]
    bv[-1] = theta[-1] / dt - k_theta[-1] / dz

    theta_new = tdma_solver(-aw, ap, -ae, bv)
    theta_new = np.clip(theta_new, substrate.theta_min, substrate.theta_sat)

    diag = {
        "j_eva_f": float(j_eva_f),
        "j_eva_g": float(j_eva_g),
        "j_eva_total": float(j_eva_total),
        "r_vap_s_m": float(r_vap),
        "h_eva_f_moisture": float(h_eva_f),
        "h_eva_g_moisture": float(h_eva_g),
    }
    return theta_new, diag


# ==============================================================================
# 09 — SLAB MODEL: Eq. (12)–(14)
# ==============================================================================

def solve_slab_heat(
    T_s: np.ndarray,
    T_g_bottom: float,
    lambda_g_bottom: float,
    slab: SlabParameters,
    geom: GeometryParameters,
    dt: float,
    T_in_K: Optional[float] = None,
) -> Tuple[np.ndarray, float, float]:
    """Solve slab temperature profile and indoor heat flux."""
    Nz = len(T_s)
    dz = slab.H_slab / (Nz - 1)
    rho_cp_s = slab.rho_s * slab.cp_s
    T_in = T_in_K if T_in_K is not None else geom.T_in_default

    aw = np.zeros(Nz); ap = np.zeros(Nz); ae = np.zeros(Nz); bv = np.zeros(Nz)
    for i in range(1, Nz - 1):
        aw[i] = slab.lambda_s / dz ** 2
        ae[i] = slab.lambda_s / dz ** 2
        ap[i] = rho_cp_s / dt + aw[i] + ae[i]
        bv[i] = rho_cp_s / dt * T_s[i]

    lam_top_bc = lambda_g_bottom / dz ** 2
    ae[0] = slab.lambda_s / dz ** 2
    ap[0] = rho_cp_s / dt + ae[0] + lam_top_bc
    bv[0] = rho_cp_s / dt * T_s[0] + lam_top_bc * T_g_bottom

    if geom.dynamic_h_in:
        h_in = interior_ceiling_convective_h(T_s[-1], T_in)
    else:
        h_in = geom.h_in

    aw[-1] = slab.lambda_s / dz ** 2
    ap[-1] = rho_cp_s / dt + aw[-1] + h_in / dz
    bv[-1] = rho_cp_s / dt * T_s[-1] + h_in * T_in / dz

    T_s_new = tdma_solver(-aw, ap, -ae, bv)
    q_s_in = h_in * (T_s_new[-1] - T_in)
    return T_s_new, float(q_s_in), float(h_in)


# ==============================================================================
# 10 — CAM MAIN SIMULATION LOOP
# ==============================================================================

def run_cam_simulation(
    weather_df: pd.DataFrame,
    plant: PlantParameters = bromelia,
    substrate: SubstrateParameters = substrat,
    slab_params: SlabParameters = slab,
    geom_params: GeometryParameters = geom,
    num_params: NumericalParameters = num,
    rs_profile: CAMRsProfile = cam_rs_profile,
    theta_initial: Optional[Union[float, Sequence[float], Dict[str, Sequence[float]]]] = None,
    T_in_series: Optional[pd.Series] = None,
    T_s_in_initial_C: Optional[float] = None,
    T_g_top_initial_C: Optional[float] = None,
    j_irrigation: float = 0.0,
    verbose: bool = True,
) -> Dict[str, list]:
    """Run the CAM-only green-roof physical simulation."""
    dt = num_params.dt
    save_every_s = num_params.save_every_s
    Nz_g = num_params.Nz_substrate
    Nz_s = num_params.Nz_slab

    if verbose:
        print("\n" + "=" * 72)
        print("CAM PHYSICAL SIMULATION — DATA-DRIVEN LI-COR gsw/r_s VERSION")
        print("=" * 72)
        print(f"r_stoma_min_CAM    : {rs_profile.r_stoma_min_s_m:.1f} s/m")
        print(f"profile r_s night  : {rs_profile.r_s_night_s_m:.1f} s/m")
        print(f"profile r_s midday : {rs_profile.r_s_midday_s_m:.1f} s/m")
        print(f"profile r_s late   : {rs_profile.r_s_late_afternoon_s_m:.1f} s/m")

    weather_1s = prepare_weather_1s(weather_df, dt=dt)
    if len(weather_1s) == 0:
        raise ValueError("weather_df is empty after resampling")

    T_in_1s = None
    if T_in_series is not None:
        T_in_1s = align_series_to_index(T_in_series, weather_1s.index, dt=dt)

    T_a_init = float(weather_1s["T_a"].iloc[0]) + 273.15
    T_g_top_init = T_a_init if T_g_top_initial_C is None else T_g_top_initial_C + 273.15
    T_s_in_init = T_a_init if T_s_in_initial_C is None else T_s_in_initial_C + 273.15

    T_f = T_a_init
    T_g = np.linspace(T_g_top_init, T_a_init, Nz_g)
    T_s = np.linspace(T_a_init, T_s_in_init, Nz_s)
    theta = make_theta_initial_profile(theta_initial, geom_params.H_g, Nz_g, substrate)

    results = {
        "datetime": [],
        "time_s": [],
        "T_a": [], "RH": [], "G_sol": [], "u": [],
        "T_in_used": [],
        "T_f": [], "T_g_top": [], "T_g_mid": [], "T_g_bot": [], "T_s_top": [], "T_s_mid": [], "T_s_in": [],
        "theta_top": [], "theta_mid": [], "theta_bot": [], "theta_mean": [],
        "q_s_in": [],
        "j_pr": [], "j_eva_f": [], "j_eva_g": [], "j_eva_total": [],
        "r_s_s_m": [], "gsw_equiv_mol_m2_s": [], "r_a_s_m": [], "VPD_Pa": [],
        "h_eva_f": [], "h_eva_g_moisture": [], "cam_phase": [],
        "h_in_used": [],
    }

    for step, (ts, row) in enumerate(weather_1s.iterrows()):
        T_a_K = float(row["T_a"]) + 273.15
        RH = float(np.clip(row["RH"], 1.0, 99.0))
        u = max(float(row["u"]), 0.1)
        G_sol = max(float(row["G_sol"]), 0.0)
        j_pr = float(j_irrigation) + max(float(row.get("rain_flux", 0.0)), 0.0)

        if T_in_1s is not None and ts in T_in_1s.index and pd.notna(T_in_1s.loc[ts]):
            T_in_current = float(T_in_1s.loc[ts]) + 273.15
        else:
            T_in_current = geom_params.T_in_default

        theta_avg = float(np.mean(theta))
        T_f, fol_diag = solve_foliage_temperature(
            T_f_prev=T_f,
            T_a_K=T_a_K,
            T_g_surface=T_g[0],
            G_sol=G_sol,
            RH=RH,
            u=u,
            theta_avg=theta_avg,
            timestamp=ts,
            plant=plant,
            substrate=substrate,
            rs_profile=rs_profile,
            dt=dt,
        )

        T_g = solve_substrate_heat(
            T_g=T_g,
            T_f=T_f,
            T_slab_top=T_s[0],
            G_sol=G_sol,
            T_a_K=T_a_K,
            RH=RH,
            u=u,
            theta=theta,
            plant=plant,
            substrate=substrate,
            H_g=geom_params.H_g,
            dt=dt,
        )

        theta, moist_diag = solve_substrate_moisture(
            theta=theta,
            T_f=T_f,
            T_g_surface=T_g[0],
            T_a_K=T_a_K,
            RH=RH,
            u=u,
            j_irrigation=j_pr,
            r_s_s_m=fol_diag["r_s_s_m"],
            plant=plant,
            substrate=substrate,
            H_g=geom_params.H_g,
            dt=dt,
        )

        _, lambda_g, _, _ = compute_substrate_properties(theta, substrate)
        T_s, q_s_in, h_in_used = solve_slab_heat(
            T_s=T_s,
            T_g_bottom=T_g[-1],
            lambda_g_bottom=lambda_g[-1],
            slab=slab_params,
            geom=geom_params,
            dt=dt,
            T_in_K=T_in_current,
        )

        if step % max(1, int(save_every_s / dt)) == 0:
            results["datetime"].append(ts)
            results["time_s"].append(step * dt)
            results["T_a"].append(T_a_K - 273.15)
            results["RH"].append(RH)
            results["G_sol"].append(G_sol)
            results["u"].append(u)
            results["T_in_used"].append(T_in_current - 273.15)
            results["T_f"].append(T_f - 273.15)
            results["T_g_top"].append(T_g[0] - 273.15)
            results["T_g_mid"].append(T_g[Nz_g // 2] - 273.15)
            results["T_g_bot"].append(T_g[-1] - 273.15)
            # Slab nodes exported for stack-aware validation:
            # T_s_top = top slab surface directly under substrate ("atap outdoor")
            # T_s_mid = middle slab temperature
            # T_s_in  = bottom/inside slab surface ("atap indoor")
            results["T_s_top"].append(T_s[0] - 273.15)
            results["T_s_mid"].append(T_s[Nz_s // 2] - 273.15)
            results["T_s_in"].append(T_s[-1] - 273.15)
            results["theta_top"].append(float(theta[0]))
            results["theta_mid"].append(float(theta[Nz_g // 2]))
            results["theta_bot"].append(float(theta[-1]))
            results["theta_mean"].append(float(np.mean(theta)))
            results["q_s_in"].append(float(q_s_in))
            results["j_pr"].append(float(j_pr))
            results["j_eva_f"].append(moist_diag["j_eva_f"])
            results["j_eva_g"].append(moist_diag["j_eva_g"])
            results["j_eva_total"].append(moist_diag["j_eva_total"])
            results["r_s_s_m"].append(fol_diag["r_s_s_m"])
            results["gsw_equiv_mol_m2_s"].append(fol_diag["gsw_equiv_mol_m2_s"])
            results["r_a_s_m"].append(fol_diag["r_a_s_m"])
            results["VPD_Pa"].append(fol_diag["VPD_Pa"])
            results["h_eva_f"].append(fol_diag["h_eva_f"])
            results["h_eva_g_moisture"].append(moist_diag["h_eva_g_moisture"])
            results["cam_phase"].append(fol_diag["cam_phase"])
            results["h_in_used"].append(h_in_used)

        if verbose and step % int(6 * 3600 / dt) == 0:
            print(
                f"  t={step*dt/3600:5.1f} h | "
                f"T_s_in={T_s[-1]-273.15:5.2f} C | "
                f"T_f={T_f-273.15:5.2f} C | "
                f"theta={np.mean(theta):.3f} | "
                f"r_s={fol_diag['r_s_s_m']:6.1f} s/m | "
                f"ETf={moist_diag['j_eva_f']*3600:.4f} kg/m2h | "
                f"ETg={moist_diag['j_eva_g']*3600:.4f} kg/m2h"
            )

    q = np.asarray(results["q_s_in"], dtype=float)
    results["Q_gain_J_m2"] = float(np.trapezoid(q, dx=save_every_s)) if len(q) else np.nan
    return results


# ==============================================================================
# 11 — VALIDATION METRICS AND RESULT EXPORT
# ==============================================================================

def results_to_dataframe(results: Dict[str, list]) -> pd.DataFrame:
    """Convert result dict to timestamp-indexed DataFrame."""
    df = pd.DataFrame({k: v for k, v in results.items() if isinstance(v, list)})
    if "datetime" in df.columns:
        df["datetime"] = pd.to_datetime(df["datetime"])
        df = df.set_index("datetime").sort_index()
    return df


def validation_metrics(sim: pd.Series, obs: pd.Series) -> Dict[str, float]:
    """Calculate validation metrics on common timestamps."""
    sim = sim.sort_index().resample("1min").mean()
    obs = obs.sort_index().resample("1min").mean()
    common = sim.index.intersection(obs.index)
    diff = sim.loc[common] - obs.loc[common]
    diff = diff.dropna()
    if diff.empty:
        return {"n": 0, "bias_C": np.nan, "mae_C": np.nan, "rmse_C": np.nan}
    sim_c = sim.loc[diff.index]
    obs_c = obs.loc[diff.index]
    return {
        "n": int(diff.count()),
        "bias_C": float(diff.mean()),
        "mae_C": float(diff.abs().mean()),
        "rmse_C": float(np.sqrt((diff ** 2).mean())),
        "amp_model_C": float(sim_c.max() - sim_c.min()),
        "amp_measured_C": float(obs_c.max() - obs_c.min()),
        "amp_error_C": float((sim_c.max() - sim_c.min()) - (obs_c.max() - obs_c.min())),
        "peak_error_C": float(sim_c.max() - obs_c.max()),
        "min_error_C": float(sim_c.min() - obs_c.min()),
    }


# ==============================================================================
# 12 — PARAMETER UPDATE HELPER
# ==============================================================================

def apply_cam_parameters(
    H_g: float = 0.10,
    H_slab: float = 0.10,
    h_in: float = 8.0,
    rho_g: float = 400.0,
    theta_sat: float = 0.90,
    k_theta_sat: float = 5e-6,
    lambda_dry: float = 0.12,
    lambda_sat: Optional[float] = None,
    LAI: float = 1.95,
    cover_fraction: float = 0.95,
    tau_f: float = 0.07,
) -> None:
    """Update global CAM parameters in one place, runner-style."""
    geom.H_g = float(H_g)
    geom.h_in = float(h_in)
    slab.H_slab = float(H_slab)
    substrat.rho_g = float(rho_g)
    substrat.theta_sat = float(theta_sat)
    substrat.k_theta_sat = float(k_theta_sat)
    substrat.lambda_dry = float(lambda_dry)
    if lambda_sat is None:
        lambda_sat = substrat.lambda_dry + substrat.theta_sat * substrat.lambda_water
    substrat.lambda_sat = float(lambda_sat)
    bromelia.LAI = float(LAI)
    bromelia.cover_fraction = float(cover_fraction)
    bromelia.tau_f = float(tau_f)
    bromelia.alpha_f = max(0.0, 1.0 - bromelia.rho_f - bromelia.tau_f)


# ==============================================================================
# 13 — SELF TEST
# ==============================================================================

if __name__ == "__main__":
    # Tiny synthetic test so the module can be sanity-checked without user files.
    idx = pd.date_range("2026-04-01 00:00:00", periods=10, freq="1min")
    weather = pd.DataFrame({
        "T_a": np.linspace(25, 30, len(idx)),
        "RH": np.full(len(idx), 80.0),
        "u": np.full(len(idx), 0.5),
        "G_sol": np.maximum(0, 400 * np.sin(np.linspace(0, np.pi, len(idx)))),
        "rain": np.zeros(len(idx)),
        "rain_flux": np.zeros(len(idx)),
    }, index=idx)
    apply_cam_parameters()
    res = run_cam_simulation(weather, verbose=False)
    print(results_to_dataframe(res).head())
