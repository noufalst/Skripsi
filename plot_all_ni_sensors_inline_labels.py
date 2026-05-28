"""
Plot all NI temperature sensor channels with labels at the end of each line.

Use:
    python plot_all_ni_sensors_inline_labels.py

Edit the FILES list below to match your NI Excel files.
"""

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# =============================================================================
# CONFIG
# =============================================================================

FILES = [
    "Pengukuran 30_1 Maret 2026.xlsx",
    "Pengukuran 30_2 Maret 2026.xlsx",
    "Pengukuran 3 April 2026.xlsx",
]

OUTPUT_DIR = Path("outputs_sensor_check")
OUTPUT_DIR.mkdir(exist_ok=True)

# Keep names as used in the March/April NI files.
NI_CHANNEL_NAMES = [
    "timestamp_serial",
    "T1Kd", "T1Kb", "T1Ke", "T1Ka", "T1Kc",
    "T3Kb", "T3Kd", "T3Ke", "T3Kc", "T3Ka",
    "T2Ka", "T2Kd", "T2Kc", "2Ke",  "T1A",
    "T2A",  "T2A2", "T1Tb", "T1Ta", "T1Ta2",
]

SENSOR_COLS = NI_CHANNEL_NAMES[1:]

# Optional time filter. Set to None if you want full data.
START_TIME = None
END_TIME = None

# Example:
# START_TIME = "2026-03-31 00:00:00"
# END_TIME   = "2026-04-03 00:00:00"

# Raw plot is recommended for sensor audit.
AUTO_CONVERT_F_TO_C = False

# If True, values outside this range are masked as NaN.
# For raw sensor audit, keep it False first.
MASK_UNPHYSICAL = False
PHYSICAL_MIN_C = -10
PHYSICAL_MAX_C = 120


# =============================================================================
# LOADER
# =============================================================================

def load_ni_excel(filepath):
    """Load one NI LabVIEW Excel file.

    Expected structure:
    col 0 = LabVIEW timestamp serial, col 1..20 = temperature channels.
    Timestamp convention: days since 1904-01-01.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        print(f"WARNING: file not found: {filepath}")
        return pd.DataFrame()

    print(f"Loading: {filepath}")

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row[:len(NI_CHANNEL_NAMES)])
        if len(vals) < len(NI_CHANNEL_NAMES):
            continue

        # Skip header / nonnumeric first cell.
        try:
            ts_serial = float(vals[0])
        except (TypeError, ValueError):
            continue

        if np.isnan(ts_serial):
            continue

        numeric_vals = [ts_serial]
        for v in vals[1:]:
            try:
                numeric_vals.append(float(v) if v is not None else np.nan)
            except (TypeError, ValueError):
                numeric_vals.append(np.nan)

        if len(numeric_vals) == len(NI_CHANNEL_NAMES):
            rows.append(numeric_vals)

    if not rows:
        raise ValueError(f"No numeric NI rows found in {filepath}")

    df = pd.DataFrame(rows, columns=NI_CHANNEL_NAMES)

    labview_epoch = pd.Timestamp("1904-01-01")
    df["datetime"] = labview_epoch + pd.to_timedelta(df["timestamp_serial"], unit="D")
    df = df.drop(columns=["timestamp_serial"]).set_index("datetime").sort_index()

    # Make sure all channels are numeric.
    for col in SENSOR_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def load_multiple_ni(files):
    dfs = []
    for fp in files:
        df_i = load_ni_excel(fp)
        if not df_i.empty:
            dfs.append(df_i)

    if not dfs:
        raise FileNotFoundError("No NI files were loaded. Check FILES paths.")

    df = pd.concat(dfs).sort_index()
    df = df[~df.index.duplicated(keep="last")]

    if START_TIME is not None:
        df = df[df.index >= pd.Timestamp(START_TIME)]
    if END_TIME is not None:
        df = df[df.index <= pd.Timestamp(END_TIME)]

    if MASK_UNPHYSICAL:
        for col in SENSOR_COLS:
            bad = (df[col] < PHYSICAL_MIN_C) | (df[col] > PHYSICAL_MAX_C)
            df.loc[bad, col] = np.nan

    if AUTO_CONVERT_F_TO_C:
        df = auto_convert_possible_fahrenheit(df)

    print(f"Combined period: {df.index.min()} → {df.index.max()}")
    print(f"Records: {len(df)}")
    return df


# =============================================================================
# OPTIONAL UNIT CHECK
# =============================================================================

def auto_convert_possible_fahrenheit(df):
    """Heuristic conversion for channels that look like Fahrenheit.

    Use carefully. For raw audit, keep AUTO_CONVERT_F_TO_C = False.
    """
    out = df.copy()

    print("\nUnit check:")
    for col in SENSOR_COLS:
        s = out[col].dropna()
        if s.empty:
            continue

        p05 = s.quantile(0.05)
        med = s.median()
        p95 = s.quantile(0.95)

        looks_f = (50 < p05 < 100) and (60 < med < 110) and (70 < p95 < 130)

        if looks_f:
            converted = (out[col] - 32.0) * 5.0 / 9.0
            c_med = converted.dropna().median()

            if 15 <= c_med <= 50:
                out[col] = converted
                print(f"  {col}: converted F → C | raw median={med:.1f}, converted median={c_med:.1f}")
            else:
                print(f"  {col}: looks high but not converted | median={med:.1f}")
        else:
            print(f"  {col}: kept raw | p05={p05:.1f}, median={med:.1f}, p95={p95:.1f}")

    return out


def sensor_summary(df, save_path=None):
    rows = []
    for col in SENSOR_COLS:
        s = df[col].dropna()
        if s.empty:
            rows.append({
                "sensor": col,
                "n": 0,
                "min": np.nan,
                "mean": np.nan,
                "median": np.nan,
                "max": np.nan,
                "amplitude": np.nan,
            })
            continue

        rows.append({
            "sensor": col,
            "n": int(s.count()),
            "min": float(s.min()),
            "mean": float(s.mean()),
            "median": float(s.median()),
            "max": float(s.max()),
            "amplitude": float(s.max() - s.min()),
        })

    summary = pd.DataFrame(rows)

    if save_path is not None:
        summary.to_csv(save_path, index=False)
        print(f"Saved summary: {save_path}")

    return summary


# =============================================================================
# PLOTTING WITH END-OF-LINE LABELS
# =============================================================================

def _spread_label_positions(y_values, min_gap):
    """Spread label y-positions to reduce overlap at the right edge."""
    if len(y_values) == 0:
        return {}

    items = sorted(y_values.items(), key=lambda kv: kv[1])
    adjusted = {}
    last_y = -np.inf

    for name, y in items:
        y_new = max(y, last_y + min_gap)
        adjusted[name] = y_new
        last_y = y_new

    # Shift down slightly if label stack drifts too far upward.
    original_max = max(y_values.values())
    adjusted_max = max(adjusted.values())
    overshoot = adjusted_max - original_max

    if overshoot > 0:
        for name in adjusted:
            adjusted[name] -= overshoot * 0.5

    return adjusted


def plot_inline_labels(
    df,
    cols,
    title,
    save_path,
    ylabel="Temperature",
    resample_rule="1min",
    figsize=(16, 8),
):
    """Plot selected columns and write each label at the right end of its line."""
    data = df[cols].copy()

    if resample_rule is not None:
        data = data.resample(resample_rule).mean()

    fig, ax = plt.subplots(figsize=figsize)

    lines = {}
    for col in cols:
        s = data[col].dropna()
        if s.empty:
            continue
        line, = ax.plot(s.index, s.values, linewidth=1.4, alpha=0.95)
        lines[col] = line

    ax.set_title(title)
    ax.set_xlabel("Datetime")
    ax.set_ylabel(ylabel)
    ax.grid(True, alpha=0.25)

    if not lines:
        ax.text(0.5, 0.5, "No data", transform=ax.transAxes, ha="center", va="center")
        plt.tight_layout()
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        plt.close(fig)
        return

    # Extend right side to make room for inline labels.
    x_min = data.index.min()
    x_max = data.index.max()
    duration = x_max - x_min
    pad = duration * 0.09 if duration > pd.Timedelta(0) else pd.Timedelta(hours=1)
    x_label = x_max + pad * 0.15
    ax.set_xlim(x_min, x_max + pad)

    # Compute final y for every plotted line.
    final_y = {}
    final_x = {}
    for col, line in lines.items():
        s = data[col].dropna()
        final_x[col] = s.index[-1]
        final_y[col] = float(s.iloc[-1])

    y_min, y_max = ax.get_ylim()
    y_span = y_max - y_min
    min_gap = y_span * 0.025
    label_y = _spread_label_positions(final_y, min_gap=min_gap)

    for col, line in lines.items():
        color = line.get_color()
        y0 = final_y[col]
        y1 = label_y[col]

        # small connector from line end to text position
        ax.plot([final_x[col], x_label], [y0, y1], color=color, linewidth=0.8, alpha=0.55)
        ax.text(
            x_label,
            y1,
            f"{col}",
            color=color,
            fontsize=9,
            va="center",
            ha="left",
            fontweight="bold",
        )

    # No separate legend. Labels are inline at each line end.
    plt.tight_layout()
    fig.savefig(save_path, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved plot: {save_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    df = load_multiple_ni(FILES)

    summary = sensor_summary(df, OUTPUT_DIR / "sensor_summary.csv")
    print("\nSensor summary:")
    print(summary.to_string(index=False))

    # Full all-sensor plot
    plot_inline_labels(
        df,
        cols=SENSOR_COLS,
        title="All NI Sensor Channels — Inline End Labels",
        save_path=OUTPUT_DIR / "all_NI_sensors_inline_labels.png",
        ylabel="Temperature",
        resample_rule="1min",
        figsize=(18, 9),
    )

    # Optional grouped plots, easier to read than one crowded plot.
    groups = {
        "CAM_box": ["T1Kd", "T1Kb", "T1Ke", "T1Ka", "T1Kc", "T2A", "T2A2", "T1Tb"],
        "C3_box": ["T3Kb", "T3Kd", "T3Ke", "T3Kc", "T3Ka", "T1A"],
        "RR_box": ["T2Ka", "T2Kd", "T2Kc", "2Ke", "T1Ta", "T1Ta2"],
        "roof_candidates": ["T2A2", "T1Tb", "T1A", "T3Ka", "T1Ta2", "T2Ka"],
        "room_air_candidates": ["T1Ka", "T3Kd", "T1Ta"],
    }

    for name, cols in groups.items():
        cols = [c for c in cols if c in df.columns]
        if cols:
            plot_inline_labels(
                df,
                cols=cols,
                title=f"NI Sensor Check — {name}",
                save_path=OUTPUT_DIR / f"{name}_inline_labels.png",
                ylabel="Temperature",
                resample_rule="1min",
                figsize=(16, 7),
            )

    print("\nDone.")
    print(f"Output folder: {OUTPUT_DIR.resolve()}")


if __name__ == "__main__":
    main()
