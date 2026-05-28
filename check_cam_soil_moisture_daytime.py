"""
CAM soil moisture daytime-drop diagnostic.

Purpose
-------
Checks whether CAM-side soil moisture actually decreases during daytime.
This is useful for verifying whether daytime CAM cooling can reasonably be
explained by substrate evaporation, or whether the CAM substrate is mostly flat
and cooling should be attributed more to shading / thermal inertia.

Default expected data files in the same folder:
    datasoilmoisture.zip
    weatherfile mar-april.xlsx

Expected soil-moisture ZIP members:
    sensor 1 COM5_CAM.csv   -> shallow sensor, assumed around 2 cm
    sensor 2 COM6_CAM.csv   -> deeper sensor, assumed around 7 cm

Typical usage
-------------
    python check_cam_soil_moisture_daytime.py

With explicit base folder:
    python check_cam_soil_moisture_daytime.py --base-dir "E:/Pagi/SKRRRRRRRipsi/data"

If the weather Excel is missing, the script still runs using local daytime hours
06:00-18:00 instead of solar radiation.
"""

from __future__ import annotations

import argparse
import io
import json
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd


# ==============================================================================
# 01 — CONFIGURATION
# ==============================================================================

DEFAULT_CAM_WINDOW = ("2026-03-31 11:58:00", "2026-04-02 21:42:00")
DEFAULT_ZIP = "datasoilmoisture.zip"
DEFAULT_WEATHER = "weatherfile mar-april.xlsx"
DEFAULT_WEATHER_SHEET = "3-24april"

CAM_MEMBERS = {
    "shallow": "sensor 1 COM5_CAM.csv",
    "deep": "sensor 2 COM6_CAM.csv",
}


@dataclass
class CheckConfig:
    base_dir: str = "."
    soil_zip: str = DEFAULT_ZIP
    weather_file: str = DEFAULT_WEATHER
    weather_sheet: str = DEFAULT_WEATHER_SHEET
    output_dir: str = "outputs_soil_check"
    start: str = DEFAULT_CAM_WINDOW[0]
    end: str = DEFAULT_CAM_WINDOW[1]
    solar_threshold: float = 50.0      # W/m2; daytime = G_sol > threshold
    fallback_day_start: int = 6        # used if weather file is not available
    fallback_day_end: int = 18
    min_points_per_segment: int = 30
    slope_flat_threshold: float = 0.001  # theta units per hour; +/- below this = flat
    substrate_depth_m: float = 0.10       # only for rough uniform-water equivalent
    timezone: str = "Asia/Jakarta"


# ==============================================================================
# 02 — LOADERS
# ==============================================================================

def _read_zip_text(z: zipfile.ZipFile, member_name: str) -> str:
    """Read text inside ZIP with fallback encodings."""
    raw = z.read(member_name)
    if raw[:200].count(b"\x00") > 40:
        return raw.decode("utf-16le", errors="ignore").replace("\x00", "")
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("latin1", errors="ignore")


def load_cam_soil_moisture(zip_path: Path, timezone: str = "Asia/Jakarta") -> pd.DataFrame:
    """Load CAM RIKA soil moisture from datasoilmoisture.zip."""
    if not zip_path.exists():
        raise FileNotFoundError(f"Soil moisture ZIP not found: {zip_path}")

    frames = []
    with zipfile.ZipFile(zip_path, "r") as z:
        available = set(z.namelist())
        missing = [m for m in CAM_MEMBERS.values() if m not in available]
        if missing:
            raise FileNotFoundError(
                "Missing CAM member(s) inside soil ZIP:\n"
                + "\n".join(f"  - {m}" for m in missing)
                + "\n\nAvailable members:\n"
                + "\n".join(f"  - {m}" for m in sorted(available))
            )

        for label, member in CAM_MEMBERS.items():
            text = _read_zip_text(z, member)
            raw = pd.read_csv(io.StringIO(text), sep=";")

            if "Timestamp" not in raw.columns:
                raise ValueError(f"{member} has no 'Timestamp' column")

            ts_utc = pd.to_datetime(raw["Timestamp"], utc=True, errors="coerce")
            ts_local = ts_utc.dt.tz_convert(timezone).dt.tz_localize(None)

            moisture_col = "Moisture (%)"
            temp_col = "Temperature"
            if moisture_col not in raw.columns:
                raise ValueError(f"{member} has no '{moisture_col}' column")

            df = pd.DataFrame(index=ts_local)
            df.index.name = "timestamp"
            df[f"moisture_{label}_pct"] = pd.to_numeric(raw[moisture_col], errors="coerce")
            df[f"theta_{label}"] = df[f"moisture_{label}_pct"] / 100.0
            if temp_col in raw.columns:
                df[f"soil_temp_{label}_C"] = pd.to_numeric(raw[temp_col], errors="coerce")

            df = df.dropna(how="all").sort_index()
            frames.append(df)

    out = pd.concat(frames, axis=1).sort_index()

    # Filter obvious nonphysical VWC glitches, but do not overclean.
    for col in ["theta_shallow", "theta_deep"]:
        if col in out:
            bad = (out[col] < 0.03) | (out[col] > 0.95)
            if int(bad.sum()) > 0:
                print(f"Flagging {int(bad.sum())} nonphysical points in {col} as NaN")
                out.loc[bad, col] = np.nan

    # Interpolate only short gaps so slope calculations are not dominated by small logger drops.
    out = out.resample("1min").mean()
    for col in out.columns:
        out[col] = out[col].interpolate("time", limit=15, limit_direction="both")

    return out


def _to_float(value) -> float:
    if value is None:
        return np.nan
    if isinstance(value, str):
        value = value.strip()
        if value in {"", "---", "--"}:
            return np.nan
    return pd.to_numeric(value, errors="coerce")


def load_weather_excel(weather_path: Path, sheet_name: str = DEFAULT_WEATHER_SHEET) -> Optional[pd.DataFrame]:
    """Load Davis weather Excel if available. Returns None if file is missing."""
    if not weather_path.exists():
        print(f"Weather file not found: {weather_path}")
        print("Continuing with fallback daytime definition based on local clock hours.")
        return None

    from openpyxl import load_workbook

    wb = load_workbook(weather_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Weather sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    records = []

    # Matches the loader style used in green_roof_model_clean.py: data starts at row 4.
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        ts = pd.to_datetime(f"{row[0]} {row[1]}", dayfirst=True, errors="coerce")
        if pd.isna(ts):
            continue
        records.append(
            {
                "timestamp": ts,
                "T_a_C": _to_float(row[2]),
                "RH_pct": _to_float(row[5]),
                "wind": _to_float(row[7]),
                "rain_mm": _to_float(row[17]),
                "G_sol_Wm2": _to_float(row[19]),
            }
        )

    if not records:
        raise ValueError(f"No weather records parsed from {weather_path}")

    df = pd.DataFrame(records).set_index("timestamp").sort_index()
    df = df[~df.index.duplicated(keep="first")]
    df = df.resample("1min").mean()

    for col in ["T_a_C", "RH_pct", "wind", "G_sol_Wm2"]:
        if col in df:
            df[col] = df[col].interpolate("time", limit=30, limit_direction="both")
    if "rain_mm" in df:
        df["rain_mm"] = df["rain_mm"].fillna(0).clip(lower=0)
    if "G_sol_Wm2" in df:
        df["G_sol_Wm2"] = df["G_sol_Wm2"].fillna(0).clip(lower=0)

    return df


# ==============================================================================
# 03 — ANALYSIS HELPERS
# ==============================================================================

def restrict_window(df: pd.DataFrame, start: str, end: str) -> pd.DataFrame:
    start_ts = pd.Timestamp(start)
    end_ts = pd.Timestamp(end)
    return df[(df.index >= start_ts) & (df.index <= end_ts)].copy()


def add_daytime_flag(df: pd.DataFrame, cfg: CheckConfig) -> pd.DataFrame:
    """Add daytime flag. Uses solar radiation when available, otherwise local clock time."""
    out = df.copy()
    if "G_sol_Wm2" in out.columns and out["G_sol_Wm2"].notna().sum() > 0:
        out["is_daytime"] = out["G_sol_Wm2"] > cfg.solar_threshold
        out["daytime_method"] = f"solar > {cfg.solar_threshold:g} W/m2"
    else:
        hours = out.index.hour + out.index.minute / 60.0
        out["is_daytime"] = (hours >= cfg.fallback_day_start) & (hours < cfg.fallback_day_end)
        out["daytime_method"] = f"clock {cfg.fallback_day_start:02d}:00-{cfg.fallback_day_end:02d}:00"
    return out


def linear_slope_per_hour(series: pd.Series) -> float:
    """OLS slope in theta units per hour."""
    s = series.dropna()
    if len(s) < 2:
        return np.nan
    x = (s.index - s.index[0]).total_seconds().to_numpy() / 3600.0
    y = s.to_numpy(dtype=float)
    if np.allclose(x.max(), x.min()):
        return np.nan
    slope, _intercept = np.polyfit(x, y, 1)
    return float(slope)


def classify_slope(slope_per_h: float, flat_threshold: float) -> str:
    if pd.isna(slope_per_h):
        return "insufficient"
    if slope_per_h < -flat_threshold:
        return "descending"
    if slope_per_h > flat_threshold:
        return "ascending"
    return "flat"


def summarize_period(
    df: pd.DataFrame,
    mask: pd.Series,
    label: str,
    cfg: CheckConfig,
) -> List[Dict[str, object]]:
    """Summarize one mask per date for shallow and deep theta."""
    rows: List[Dict[str, object]] = []
    work = df.loc[mask].copy()
    if work.empty:
        return rows

    for date, g in work.groupby(work.index.date):
        if len(g) < cfg.min_points_per_segment:
            continue

        for theta_col, sensor_label in [
            ("theta_shallow", "shallow_approx_2cm"),
            ("theta_deep", "deep_approx_7cm"),
        ]:
            if theta_col not in g:
                continue
            s = g[theta_col].dropna()
            if len(s) < cfg.min_points_per_segment:
                continue

            slope = linear_slope_per_hour(s)
            delta = float(s.iloc[-1] - s.iloc[0])
            duration_h = float((s.index[-1] - s.index[0]).total_seconds() / 3600.0)
            rain_total = float(g.get("rain_mm", pd.Series(0.0, index=g.index)).sum())
            solar_mean = float(g.get("G_sol_Wm2", pd.Series(np.nan, index=g.index)).mean())
            solar_max = float(g.get("G_sol_Wm2", pd.Series(np.nan, index=g.index)).max())

            # Rough uniform-water equivalent only. This does NOT mean the sensor
            # measures the whole 10 cm storage; it is a scale check.
            water_equiv_mm = float(delta * cfg.substrate_depth_m * 1000.0)

            rows.append(
                {
                    "period": label,
                    "date": str(date),
                    "sensor": sensor_label,
                    "start": str(s.index[0]),
                    "end": str(s.index[-1]),
                    "duration_h": duration_h,
                    "n_points": int(len(s)),
                    "theta_start": float(s.iloc[0]),
                    "theta_end": float(s.iloc[-1]),
                    "theta_delta": delta,
                    "theta_min": float(s.min()),
                    "theta_max": float(s.max()),
                    "slope_theta_per_h": slope,
                    "classification": classify_slope(slope, cfg.slope_flat_threshold),
                    "rough_uniform_water_equiv_delta_mm": water_equiv_mm,
                    "rain_total_mm": rain_total,
                    "G_sol_mean_Wm2": solar_mean,
                    "G_sol_max_Wm2": solar_max,
                }
            )
    return rows


def summarize_day_night(df: pd.DataFrame, cfg: CheckConfig) -> pd.DataFrame:
    day_rows = summarize_period(df, df["is_daytime"], "daytime", cfg)
    night_rows = summarize_period(df, ~df["is_daytime"], "night_or_low_solar", cfg)
    return pd.DataFrame(day_rows + night_rows)


def compute_rolling_dtheta(df: pd.DataFrame, window_min: int = 30) -> pd.DataFrame:
    """Add rolling dtheta/dt in theta units per hour."""
    out = df.copy()
    for col in ["theta_shallow", "theta_deep"]:
        if col not in out:
            continue
        # Difference over rolling window. min_periods keeps it robust at edges.
        shifted = out[col].shift(window_min)
        out[f"d{col}_per_h_{window_min}min"] = (out[col] - shifted) / (window_min / 60.0)
    return out


def make_interpretation(summary: pd.DataFrame, cfg: CheckConfig) -> Dict[str, object]:
    """Create compact interpretation from slope summary."""
    result: Dict[str, object] = {
        "daytime_method": None,
        "slope_flat_threshold_theta_per_h": cfg.slope_flat_threshold,
        "interpretation": {},
    }
    if summary.empty:
        result["interpretation"]["overall"] = "No valid day/night segments found."
        return result

    for sensor in summary["sensor"].unique():
        day = summary[(summary["sensor"] == sensor) & (summary["period"] == "daytime")]
        if day.empty:
            result["interpretation"][sensor] = "No valid daytime segment."
            continue

        counts = day["classification"].value_counts().to_dict()
        median_slope = float(day["slope_theta_per_h"].median())
        median_delta = float(day["theta_delta"].median())
        desc_frac = float((day["classification"] == "descending").mean())

        if desc_frac >= 0.67:
            statement = "mostly descending during daytime"
        elif desc_frac <= 0.33 and abs(median_slope) <= cfg.slope_flat_threshold:
            statement = "mostly flat during daytime"
        elif median_slope > cfg.slope_flat_threshold:
            statement = "often increasing during daytime"
        else:
            statement = "mixed / weak daytime decline"

        result["interpretation"][sensor] = {
            "statement": statement,
            "classification_counts": counts,
            "median_daytime_slope_theta_per_h": median_slope,
            "median_daytime_delta_theta": median_delta,
            "descending_fraction": desc_frac,
        }

    return result


# ==============================================================================
# 04 — PLOTS
# ==============================================================================

def shade_daytime(ax, df: pd.DataFrame) -> None:
    """Shade contiguous daytime regions."""
    if "is_daytime" not in df:
        return
    mask = df["is_daytime"].fillna(False).to_numpy()
    idx = df.index
    if len(mask) == 0:
        return

    start = None
    for i, val in enumerate(mask):
        if val and start is None:
            start = idx[i]
        if start is not None and ((not val) or i == len(mask) - 1):
            end = idx[i] if val and i == len(mask) - 1 else idx[i - 1]
            ax.axvspan(start, end, alpha=0.08)
            start = None


def plot_diagnostic(df: pd.DataFrame, summary: pd.DataFrame, cfg: CheckConfig, save_path: Path) -> None:
    df_plot = compute_rolling_dtheta(df)

    fig, axes = plt.subplots(4, 1, figsize=(14, 12), sharex=True)

    ax = axes[0]
    shade_daytime(ax, df_plot)
    if "theta_shallow" in df_plot:
        ax.plot(df_plot.index, df_plot["theta_shallow"], label="theta shallow ~2 cm", linewidth=1.8)
    if "theta_deep" in df_plot:
        ax.plot(df_plot.index, df_plot["theta_deep"], label="theta deep ~7 cm", linewidth=1.8)
    ax.set_ylabel("VWC θ (-)")
    ax.set_title("CAM soil moisture: daytime shaded regions")
    ax.grid(True, alpha=0.3)
    ax.legend()

    ax = axes[1]
    shade_daytime(ax, df_plot)
    if "G_sol_Wm2" in df_plot:
        ax.plot(df_plot.index, df_plot["G_sol_Wm2"], label="G_sol", linewidth=1.5)
        ax.axhline(cfg.solar_threshold, linestyle="--", linewidth=1, label="day threshold")
    else:
        ax.text(0.02, 0.5, "No weather/solar file; daytime uses clock hours", transform=ax.transAxes)
    ax.set_ylabel("Solar (W/m²)")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="upper right")

    ax2 = ax.twinx()
    if "rain_mm" in df_plot:
        ax2.bar(df_plot.index, df_plot["rain_mm"], width=0.002, alpha=0.25, label="rain")
        ax2.set_ylabel("Rain (mm/min)")

    ax = axes[2]
    shade_daytime(ax, df_plot)
    for col, label in [
        ("dtheta_shallow_per_h_30min", "dθ shallow/dt, 30-min"),
        ("dtheta_deep_per_h_30min", "dθ deep/dt, 30-min"),
    ]:
        if col in df_plot:
            ax.plot(df_plot.index, df_plot[col], label=label, linewidth=1.2)
    ax.axhline(0, linestyle="--", linewidth=1)
    ax.axhline(-cfg.slope_flat_threshold, linestyle=":", linewidth=1, label="descending threshold")
    ax.axhline(cfg.slope_flat_threshold, linestyle=":", linewidth=1)
    ax.set_ylabel("dθ/dt per hour")
    ax.set_title("Rolling moisture slope: negative means drying")
    ax.grid(True, alpha=0.3)
    ax.legend()

    ax = axes[3]
    if summary.empty:
        ax.text(0.02, 0.5, "No valid summary segments", transform=ax.transAxes)
    else:
        day = summary[summary["period"] == "daytime"].copy()
        if not day.empty:
            labels = day["date"] + "\n" + day["sensor"].str.replace("_", " ")
            ax.bar(np.arange(len(day)), day["slope_theta_per_h"])
            ax.axhline(0, linestyle="--", linewidth=1)
            ax.axhline(-cfg.slope_flat_threshold, linestyle=":", linewidth=1)
            ax.axhline(cfg.slope_flat_threshold, linestyle=":", linewidth=1)
            ax.set_xticks(np.arange(len(day)))
            ax.set_xticklabels(labels, rotation=45, ha="right")
            ax.set_ylabel("Daytime slope θ/hour")
            ax.set_title("Per-day daytime slope summary")
            ax.grid(True, axis="y", alpha=0.3)

    axes[-1].set_xlabel("Datetime")
    fig.tight_layout()
    fig.savefig(save_path, dpi=200, bbox_inches="tight")
    plt.close(fig)


def plot_solar_vs_drying(df: pd.DataFrame, cfg: CheckConfig, save_path: Path) -> None:
    """Scatter plot: solar radiation vs rolling dtheta/dt."""
    if "G_sol_Wm2" not in df:
        return

    d = compute_rolling_dtheta(df).dropna(subset=["G_sol_Wm2"])
    fig, ax = plt.subplots(figsize=(9, 6))

    plotted = False
    for col, label in [
        ("dtheta_shallow_per_h_30min", "shallow ~2 cm"),
        ("dtheta_deep_per_h_30min", "deep ~7 cm"),
    ]:
        if col in d:
            valid = d[["G_sol_Wm2", col]].dropna()
            if len(valid) > 0:
                ax.scatter(valid["G_sol_Wm2"], valid[col], s=12, alpha=0.45, label=label)
                plotted = True

    if not plotted:
        plt.close(fig)
        return

    ax.axhline(0, linestyle="--", linewidth=1)
    ax.axhline(-cfg.slope_flat_threshold, linestyle=":", linewidth=1)
    ax.axvline(cfg.solar_threshold, linestyle="--", linewidth=1, label="day threshold")
    ax.set_xlabel("Solar radiation G_sol (W/m²)")
    ax.set_ylabel("Rolling dθ/dt, 30-min (θ per hour)")
    ax.set_title("Does high daytime solar coincide with CAM soil drying?")
    ax.grid(True, alpha=0.3)
    ax.legend()
    fig.tight_layout()
    fig.savefig(save_path, dpi=200, bbox_inches="tight")
    plt.close(fig)


# ==============================================================================
# 05 — MAIN
# ==============================================================================

def parse_args() -> CheckConfig:
    parser = argparse.ArgumentParser(description="Check CAM soil moisture daytime descending behavior.")
    parser.add_argument("--base-dir", default=".", help="Folder containing data files.")
    parser.add_argument("--soil-zip", default=DEFAULT_ZIP, help="Soil moisture ZIP filename/path.")
    parser.add_argument("--weather-file", default=DEFAULT_WEATHER, help="Weather Excel filename/path.")
    parser.add_argument("--weather-sheet", default=DEFAULT_WEATHER_SHEET, help="Weather Excel sheet name.")
    parser.add_argument("--output-dir", default="outputs_soil_check", help="Output folder.")
    parser.add_argument("--start", default=DEFAULT_CAM_WINDOW[0], help="Start datetime for CAM check.")
    parser.add_argument("--end", default=DEFAULT_CAM_WINDOW[1], help="End datetime for CAM check.")
    parser.add_argument("--solar-threshold", type=float, default=50.0, help="Solar threshold for daytime, W/m².")
    parser.add_argument("--slope-flat-threshold", type=float, default=0.001, help="Theta/hour threshold for flat vs trend.")
    parser.add_argument("--substrate-depth-m", type=float, default=0.10, help="Rough water equivalent depth scale.")
    ns = parser.parse_args()
    return CheckConfig(**vars(ns))


def resolve_path(base: Path, maybe_path: str) -> Path:
    p = Path(maybe_path)
    return p if p.is_absolute() else base / p


def main() -> None:
    cfg = parse_args()
    base = Path(cfg.base_dir)
    out_dir = resolve_path(base, cfg.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    soil_zip_path = resolve_path(base, cfg.soil_zip)
    weather_path = resolve_path(base, cfg.weather_file)

    print("=" * 78)
    print("CAM SOIL MOISTURE DAYTIME-DROP CHECK")
    print("=" * 78)
    print(f"Base dir       : {base.resolve()}")
    print(f"Soil ZIP       : {soil_zip_path}")
    print(f"Weather file   : {weather_path}")
    print(f"Window         : {cfg.start} -> {cfg.end}")
    print(f"Solar threshold: {cfg.solar_threshold:g} W/m²")
    print(f"Flat threshold : ±{cfg.slope_flat_threshold:g} θ/hour")

    soil = load_cam_soil_moisture(soil_zip_path, timezone=cfg.timezone)
    weather = load_weather_excel(weather_path, sheet_name=cfg.weather_sheet)

    if weather is not None:
        df = soil.join(weather, how="left")
    else:
        df = soil.copy()

    df = restrict_window(df, cfg.start, cfg.end)
    if df.empty:
        raise ValueError("No data inside selected CAM window.")

    # Fill small weather gaps after joining to soil timestamps.
    for col in ["T_a_C", "RH_pct", "wind", "G_sol_Wm2"]:
        if col in df:
            df[col] = df[col].interpolate("time", limit=30, limit_direction="both")
    if "rain_mm" in df:
        df["rain_mm"] = df["rain_mm"].fillna(0)

    df = add_daytime_flag(df, cfg)
    daytime_method = str(df["daytime_method"].dropna().iloc[0])

    summary = summarize_day_night(df, cfg)
    interpretation = make_interpretation(summary, cfg)
    interpretation["daytime_method"] = daytime_method
    interpretation["window"] = {"start": cfg.start, "end": cfg.end}

    clean_csv = out_dir / "cam_soil_moisture_joined_clean.csv"
    summary_csv = out_dir / "cam_soil_day_night_slope_summary.csv"
    summary_json = out_dir / "cam_soil_daytime_check_summary.json"
    plot_path = out_dir / "cam_soil_moisture_daytime_check.png"
    scatter_path = out_dir / "cam_solar_vs_soil_drying_scatter.png"

    df.to_csv(clean_csv, index_label="timestamp")
    summary.to_csv(summary_csv, index=False)
    with summary_json.open("w", encoding="utf-8") as f:
        json.dump({"config": asdict(cfg), "result": interpretation}, f, indent=2, default=str)

    plot_diagnostic(df, summary, cfg, plot_path)
    plot_solar_vs_drying(df, cfg, scatter_path)

    print("\n=== OUTPUT FILES ===")
    print(f"Joined clean data : {clean_csv}")
    print(f"Slope summary     : {summary_csv}")
    print(f"JSON summary      : {summary_json}")
    print(f"Main plot         : {plot_path}")
    if scatter_path.exists():
        print(f"Solar scatter     : {scatter_path}")

    print("\n=== QUICK INTERPRETATION ===")
    print(f"Daytime definition: {daytime_method}")
    if summary.empty:
        print("No valid day/night segments found.")
    else:
        day = summary[summary["period"] == "daytime"]
        print(day[[
            "date", "sensor", "theta_delta", "slope_theta_per_h",
            "classification", "rain_total_mm", "G_sol_mean_Wm2", "G_sol_max_Wm2",
        ]].to_string(index=False))

        print("\nInterpretation by sensor:")
        for sensor, item in interpretation["interpretation"].items():
            if isinstance(item, dict):
                print(
                    f"  {sensor}: {item['statement']} | "
                    f"median slope={item['median_daytime_slope_theta_per_h']:.5f} θ/h | "
                    f"median Δθ={item['median_daytime_delta_theta']:.5f}"
                )
            else:
                print(f"  {sensor}: {item}")

    print("\nNote:")
    print("  A flat 2 cm / 7 cm sensor does not prove zero surface evaporation.")
    print("  It means daytime evaporation is not strong enough, deep enough, or visible enough")
    print("  in those sensors to appear as a clear VWC decline.")


if __name__ == "__main__":
    main()
