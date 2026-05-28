"""
================================================================================
GREEN ROOF SIMULATION PATCH / RUNNER — MAY 2026
================================================================================
This file is meant to be placed in the same folder as `new_baru.py`.
It imports the existing model, applies the fixes discussed in review, and adds
loaders for the actual weather/NI/RIKA files.

Main purpose now:
- temporary scientific-guess simulation that runs and can be compared to data,
- split validation windows for CAM and C3,
- safer physics/data handling before final parameter values are available.

Key changes vs new_baru.py:
1) Weather Excel loader for `weatherfile mar-april.xlsx`.
2) NI loader that supports both raw LabVIEW serial XLSX and cleaned XLSX.
3) RIKA soil-moisture ZIP loader.
4) Dew-point-based sky temperature.
5) Substrate thermal conductivity separated from hydraulic conductivity.
6) Positive hydraulic diffusivity when psi_sat is stored as suction magnitude.
7) G_sol passed into moisture solver, so CAM day/night stomata logic works.
8) Rainfall is treated as conserved flux, not linearly interpolated state.
9) Results store datetimes, so RMSE/MAE are computed at common timestamps.
10) CAM and C3 simulation windows are split.

Notes:
- Sensor target for validation is inner roof surface temperature:
  CAM = T1Tb, C3 = T3Ka.
- T1Ka/T3Kd can still be used as the interior-air boundary condition for Eq.14.
- All guessed parameters are collected in apply_scientific_guess_parameters().
================================================================================
"""

from __future__ import annotations

import io
import os
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, Iterable, Optional, Sequence, Tuple, Union

import numpy as np
import pandas as pd

# Import the existing code. Keep this file in the same folder as new_baru.py.
import new_baru as gr


# ==============================================================================
# VALIDATION WINDOWS
# ==============================================================================

WINDOWS: Dict[str, Tuple[pd.Timestamp, pd.Timestamp]] = {
    # Main CAM window: chosen because Apr 4 onward has RIKA anomaly 0 -> 90s.
    "CAM": (pd.Timestamp("2026-03-31 11:58:00"),
            pd.Timestamp("2026-04-02 21:42:00")),

    # Main C3 window: best clean overlap after the big RIKA gap.
    # Contains one night and one daytime heating period; shorter than CAM but cleaner.
    "C3":  (pd.Timestamp("2026-04-09 11:05:00"),
            pd.Timestamp("2026-04-10 14:08:00")),
}

# Alternative C3 window if you want to inspect another case.
WINDOWS_ALT: Dict[str, Tuple[pd.Timestamp, pd.Timestamp]] = {
    "C3_short_1": (pd.Timestamp("2026-04-06 17:38:00"),
                   pd.Timestamp("2026-04-07 10:06:00")),
}


# ==============================================================================
# SCIENTIFIC-GUESS PARAMETERS
# ==============================================================================

def apply_scientific_guess_parameters(
    *,
    rho_g: float = 400.0,          # kg/m3, measured/available from user
    theta_sat: float = 0.55,       # m3/m3, temporary literature/RIKA-based guess
    k_theta_sat: float = 5e-6,     # m/s, temporary effective hydraulic conductivity
    H_g: float = 0.06,             # m, temporary effective substrate depth
    H_slab: float = 0.10,          # m, reference-paper slab default; replace later
    lambda_dry: float = 0.08,      # W/mK, lightweight soil+rice-husk baseline
    lambda_sat: Optional[float] = None,
    h_in: float = 8.0,             # W/m2K, fixed for now
) -> None:
    """Fill missing parameters for temporary runs.

    lambda_sat default:
        lambda_sat = lambda_dry + theta_sat * lambda_water
    This is a simple pore-water contribution estimate. Replace/sensitivity-test later.
    """
    gr.substrat.rho_g = float(rho_g)
    gr.substrat.theta_sat = float(theta_sat)
    gr.substrat.k_theta_sat = float(k_theta_sat)
    gr.substrat.lambda_dry = float(lambda_dry)

    if lambda_sat is None:
        lambda_sat = gr.substrat.lambda_dry + gr.substrat.theta_sat * gr.substrat.lambda_water
    # new_baru.SubstrateParameters has no lambda_sat field, so attach dynamically.
    gr.substrat.lambda_sat = float(lambda_sat)

    gr.geom.H_g = float(H_g)
    gr.geom.h_in = float(h_in)
    gr.slab.H_slab = float(H_slab)
    gr.geom.H_slab = float(H_slab)


# ==============================================================================
# LOADERS
# ==============================================================================

def _to_float(x):
    if x is None:
        return np.nan
    if isinstance(x, str):
        x = x.strip()
        if x in {"", "---", "--"}:
            return np.nan
    return pd.to_numeric(x, errors="coerce")


def load_weather_excel(
    filepath: Union[str, Path],
    sheet_name: str = "3-24april",
    date_start: Optional[Union[str, pd.Timestamp]] = None,
    date_end: Optional[Union[str, pd.Timestamp]] = None,
) -> pd.DataFrame:
    """Load Davis weather Excel file.

    Expected columns from WeatherLink export:
    Date(0), Time(1), T_a(2), RH(5), wind speed(7), rain(17), solar rad(19).

    Output columns:
    T_a [degC], RH [%], u [m/s or logger unit as provided], rain [mm/min],
    G_sol [W/m2], rain_flux [kg/m2/s].
    """
    from openpyxl import load_workbook

    filepath = str(filepath)
    start = pd.Timestamp(date_start) if date_start is not None else None
    end = pd.Timestamp(date_end) if date_end is not None else None

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    records = []
    # First three rows are headers.
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        ts = pd.to_datetime(f"{row[0]} {row[1]}", dayfirst=True, errors="coerce")
        if pd.isna(ts):
            continue
        if start is not None and ts < start:
            continue
        if end is not None and ts > end:
            continue
        records.append({
            "timestamp": ts,
            "T_a": _to_float(row[2]),
            "RH": _to_float(row[5]),
            "u": _to_float(row[7]),
            "rain": _to_float(row[17]),
            "G_sol": _to_float(row[19]),
        })

    if not records:
        raise ValueError(f"No weather data found for requested period in {filepath}")

    df = pd.DataFrame.from_records(records).set_index("timestamp").sort_index()

    # The uploaded weather file contains duplicated rows; keep the first occurrence.
    df = df[~df.index.duplicated(keep="first")]

    # Basic physical cleaning.
    df["T_a"] = pd.to_numeric(df["T_a"], errors="coerce")
    df["RH"] = pd.to_numeric(df["RH"], errors="coerce").clip(0, 100)
    df["u"] = pd.to_numeric(df["u"], errors="coerce").clip(lower=0.1)
    df["rain"] = pd.to_numeric(df["rain"], errors="coerce").fillna(0).clip(lower=0)
    df["G_sol"] = pd.to_numeric(df["G_sol"], errors="coerce").fillna(0).clip(lower=0)

    # Interpolate state variables only.
    for col in ["T_a", "RH", "u", "G_sol"]:
        df[col] = df[col].interpolate(method="time", limit=30, limit_direction="both")

    # If rain is mm per minute, 1 mm = 1 kg/m2; flux is kg/m2/s.
    # Do not linearly interpolate rain as a state variable.
    df["rain_flux"] = df["rain"] / 60.0

    return df.dropna(subset=["T_a", "RH", "u", "G_sol"])



def create_weather_cache(
    filepath: Union[str, Path],
    cache_path: Union[str, Path],
    sheet_name: str = "3-24april",
) -> pd.DataFrame:
    """Create a CSV cache from the heavy weather XLSX.

    The weather XLSX is large. Parsing once and saving CSV makes later runs much faster.
    """
    df = load_weather_excel(filepath, sheet_name=sheet_name)
    cache_path = Path(cache_path)
    df.to_csv(cache_path, index_label="timestamp")
    return df


def load_weather_cache_or_excel(
    filepath: Union[str, Path],
    cache_path: Union[str, Path],
    sheet_name: str = "3-24april",
    date_start: Optional[Union[str, pd.Timestamp]] = None,
    date_end: Optional[Union[str, pd.Timestamp]] = None,
) -> pd.DataFrame:
    """Load weather from CSV cache if available; otherwise parse XLSX.

    For first run, you may call create_weather_cache() once. If no cache exists,
    this function falls back to the Excel loader.
    """
    cache_path = Path(cache_path)
    if cache_path.exists():
        df = pd.read_csv(cache_path, parse_dates=["timestamp"]).set_index("timestamp").sort_index()
        if date_start is not None:
            df = df[df.index >= pd.Timestamp(date_start)]
        if date_end is not None:
            df = df[df.index <= pd.Timestamp(date_end)]
        return df
    return load_weather_excel(filepath, sheet_name=sheet_name, date_start=date_start, date_end=date_end)


def _excel_col_to_index(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha())
    idx = 0
    for ch in letters:
        idx = idx * 26 + ord(ch.upper()) - 64
    return idx - 1


def _load_NI_raw_labview_xlsx(filepath: Union[str, Path]) -> pd.DataFrame:
    """Load raw LabVIEW xlsx by XML parsing.

    Used for `Pengukuran 30_1/30_2 Maret 2026.xlsx`, because openpyxl can fail
    on styles in those files.
    """
    filepath = str(filepath)
    with zipfile.ZipFile(filepath, "r") as z:
        xml_content = z.read("xl/worksheets/sheet1.xml")

    tree = ET.fromstring(xml_content)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rows_xml = tree.findall(f".//{ns}row")

    names = [
        "timestamp_serial",
        "T1Kd", "T1Kb", "T1Ke", "T1Ka", "T1Kc",
        "T3Kb", "T3Kd", "T3Ke", "T3Kc", "T3Ka",
        "T2Ka", "T2Kd", "T2Kc", "2Ke", "T1A",
        "T2A", "T2A2", "T1Tb", "T1Ta", "T1Ta2",
    ]

    data = []
    for row in rows_xml[1:]:
        vals = []
        for cell in row.findall(f"{ns}c"):
            idx = _excel_col_to_index(cell.attrib.get("r", ""))
            while len(vals) <= idx:
                vals.append(np.nan)
            v = cell.find(f"{ns}v")
            vals[idx] = float(v.text) if v is not None else np.nan
        if len(vals) >= 21:
            data.append(vals[:21])

    df = pd.DataFrame(data, columns=names)
    df["timestamp"] = pd.Timestamp("1904-01-01") + pd.to_timedelta(
        df["timestamp_serial"], unit="D"
    )
    return df.drop(columns=["timestamp_serial"]).set_index("timestamp").sort_index()


def _load_NI_clean_xlsx(filepath: Union[str, Path]) -> pd.DataFrame:
    """Load cleaned NI workbook with sheet `NI Sensor Data`."""
    df = pd.read_excel(
        filepath,
        sheet_name="NI Sensor Data",
        header=1,
        skiprows=[2],
        engine="openpyxl",
    )
    df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
    df = df.dropna(subset=["Timestamp"]).set_index("Timestamp").sort_index()
    return df


def _add_NI_alias_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # General anomaly filter for temperature-like channels.
    temp_cols = [c for c in df.columns if c.startswith("T") or c == "2Ke"]
    for col in temp_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        bad = (df[col] < -10) | (df[col] > 80)
        df.loc[bad, col] = np.nan
        df[col] = df[col].interpolate(method="time", limit=30, limit_direction="both")

    # Alias mapping used by the model.
    if "T1Ke" in df: df["T_g_top_CAM"] = df["T1Ke"]
    if "T2A" in df: df["T_g_bot_CAM"] = df["T2A"]
    if "T1Tb" in df: df["T_s_in_CAM"] = df["T1Tb"]
    if "T1Ta" in df: df["T_s_ext_CAM"] = df["T1Ta"]
    if "T3Ka" in df: df["T_s_in_C3"] = df["T3Ka"]
    if "T1Ka" in df: df["T_in_CAM"] = df["T1Ka"]
    if "T3Kd" in df: df["T_in_C3"] = df["T3Kd"]
    if "T2Ka" in df: df["T_r_in_RR"] = df["T2Ka"]
    return df


def load_NI_sensor_data_auto(filepath: Union[str, Path]) -> pd.DataFrame:
    """Load either cleaned NI workbook or raw LabVIEW serial workbook."""
    filepath = str(filepath)
    try:
        # Try the cleaned file layout first.
        df = _load_NI_clean_xlsx(filepath)
    except Exception:
        df = _load_NI_raw_labview_xlsx(filepath)
    return _add_NI_alias_columns(df)


def load_multiple_NI_sensor_data(filepaths: Sequence[Union[str, Path]]) -> pd.DataFrame:
    dfs = [load_NI_sensor_data_auto(p) for p in filepaths]
    df = pd.concat(dfs).sort_index()
    df = df[~df.index.duplicated(keep="last")]
    return _add_NI_alias_columns(df)


def _read_zip_text(z: zipfile.ZipFile, name: str) -> str:
    data = z.read(name)
    # Some RIKA CSVs are UTF-16LE-like with null bytes.
    if data[:200].count(b"\x00") > 40:
        return data.decode("utf-16le", errors="ignore").replace("\x00", "")
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("latin1", errors="ignore")


def load_rika_soil_file_from_zip(
    zip_path: Union[str, Path],
    member_name: str,
    theta_min_valid: float = 0.03,
    theta_max_valid: float = 0.88,
) -> pd.DataFrame:
    """Load one RIKA CSV from ZIP.

    Timestamp is recorded in GMT. Output index is converted to WIB/Asia-Jakarta
    and timezone-naive.
    """
    with zipfile.ZipFile(zip_path, "r") as z:
        txt = _read_zip_text(z, member_name)

    df = pd.read_csv(io.StringIO(txt), sep=";")
    df["timestamp_utc"] = pd.to_datetime(df["Timestamp"], utc=True, errors="coerce")
    df["timestamp"] = df["timestamp_utc"].dt.tz_convert("Asia/Jakarta").dt.tz_localize(None)
    df["soil_temp"] = pd.to_numeric(df.get("Temperature"), errors="coerce")
    df["moisture_pct"] = pd.to_numeric(df.get("Moisture (%)"), errors="coerce")
    df["theta"] = df["moisture_pct"] / 100.0
    df = df.dropna(subset=["timestamp"]).set_index("timestamp").sort_index()

    # Flag obvious impossible / known glitch values.
    bad = (df["theta"] < theta_min_valid) | (df["theta"] > theta_max_valid)
    df.loc[bad, ["moisture_pct", "theta"]] = np.nan
    return df[["soil_temp", "moisture_pct", "theta"]]


def load_rika_pair(
    zip_path: Union[str, Path],
    plant_type: str,
) -> pd.DataFrame:
    """Load paired RIKA sensors for CAM or C3.

    Assumption for now:
    - sensor 1 = shallow sensor, approx 2 cm
    - sensor 2 = deeper sensor, approx 7 cm
    """
    plant_type = plant_type.upper()
    if plant_type == "CAM":
        members = ("sensor 1 COM5_CAM.csv", "sensor 2 COM6_CAM.csv")
    elif plant_type == "C3":
        members = ("sensor 1 COM5_C3.csv", "sensor 2 COM6_C3.csv")
    else:
        raise ValueError("plant_type must be 'CAM' or 'C3'")

    s1 = load_rika_soil_file_from_zip(zip_path, members[0]).rename(
        columns={"soil_temp": "soil_temp_2cm", "moisture_pct": "moisture_2cm_pct", "theta": "theta_2cm"}
    )
    s2 = load_rika_soil_file_from_zip(zip_path, members[1]).rename(
        columns={"soil_temp": "soil_temp_7cm", "moisture_pct": "moisture_7cm_pct", "theta": "theta_7cm"}
    )
    return pd.concat([s1, s2], axis=1).sort_index()


# ==============================================================================
# CORRECTED THERMODYNAMIC HELPERS
# ==============================================================================

def dew_point_C(T_C: float, RH: float) -> float:
    """Magnus dew point [degC] from air temperature and relative humidity."""
    RH = float(np.clip(RH, 1.0, 100.0))
    a, b = 17.27, 237.7
    gamma = (a * T_C) / (b + T_C) + np.log(RH / 100.0)
    return (b * gamma) / (a - gamma)


def sky_temperature(T_a_K: float, RH: float) -> float:
    """Effective sky temperature [K].

    Uses dew-point-based clear-sky emissivity correlation, not RH directly.
    """
    T_C = T_a_K - 273.15
    Tdp = dew_point_C(T_C, RH)
    eps_sky = 0.711 + 0.56 * (Tdp / 100.0) + 0.73 * (Tdp / 100.0) ** 2
    eps_sky = float(np.clip(eps_sky, 0.60, 1.00))
    return eps_sky ** 0.25 * T_a_K


# ==============================================================================
# CORRECTED SUBSTRATE PROPERTIES / MOISTURE SOLVER
# ==============================================================================

def compute_substrate_properties(
    theta: np.ndarray,
    substrate: gr.SubstrateParameters,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """Moisture-dependent substrate properties.

    Important correction:
    - k_theta remains hydraulic conductivity [m/s].
    - lambda_g is thermal conductivity [W/mK] and is calculated from saturation.
    """
    theta = np.asarray(theta, dtype=float)
    b = substrate.b
    if substrate.theta_sat is None or substrate.k_theta_sat is None or substrate.rho_g is None:
        raise ValueError("substrate.theta_sat, k_theta_sat, and rho_g must be set before running simulation")

    theta_safe = np.maximum(theta, substrate.theta_min + 1e-6)
    S = np.clip(theta_safe / substrate.theta_sat, 0.0, 1.0)

    # Eq. B.8 style volumetric heat capacity used in the original code.
    rho_cp_g = substrate.cp_g * (0.2 + theta_safe) * substrate.rho_g

    # Hydraulic conductivity for Richards/moisture transport only.
    k_theta = substrate.k_theta_sat * S ** (2.0 * b + 3.0)

    # Thermal conductivity for heat conduction only.
    lambda_sat = getattr(substrate, "lambda_sat", None)
    if lambda_sat is None:
        lambda_sat = substrate.lambda_dry + substrate.theta_sat * substrate.lambda_water
    lambda_g = substrate.lambda_dry + S * (lambda_sat - substrate.lambda_dry)

    # psi_sat is stored as suction magnitude (+), so D_theta should be positive.
    psi_abs = abs(substrate.psi_sat)
    D_theta = b * substrate.k_theta_sat * psi_abs / theta_safe * S ** (b + 3.0)

    return rho_cp_g, lambda_g, k_theta, D_theta


def substrate_vapor_resistance(theta_surface: float, substrate: gr.SubstrateParameters) -> float:
    """Simple effective vapor resistance [s/m] increasing as substrate dries.

    This replaces the dimensionally strange placeholder r_vap = lambda_dry*50.
    It is intentionally simple for temporary simulation; calibrate/sensitivity-test later.
    """
    S = float(np.clip(theta_surface / substrate.theta_sat, 0.05, 1.0))
    return 100.0 * S ** (-2.0)


def solve_substrate_heat(
    T_g: np.ndarray,
    T_f: float,
    T_slab_top: float,
    G_sol: float,
    T_a_K: float,
    RH: float,
    u: float,
    theta: np.ndarray,
    plant: gr.PlantParameters,
    substrate: gr.SubstrateParameters,
    H_g: float,
    dt: float,
) -> np.ndarray:
    """Corrected wrapper: same as original, but uses corrected properties and r_vap."""
    sigma = 5.67e-8
    gamma = gr.psychrometric_constant()
    rho_air = 1.2
    cp_air = 1005

    Nz = len(T_g)
    dz = H_g / (Nz - 1)

    rho_cp_g, lambda_g, _k_theta, _ = compute_substrate_properties(theta, substrate)

    T_sky = sky_temperature(T_a_K, RH)
    P_a = gr.ambient_vapor_pressure(T_a_K, RH)
    P_g_sat = gr.saturation_pressure(T_g[0])
    epsilon_fg = 1 / (1 / plant.epsilon_f + 1 / substrate.epsilon_g - 1)

    R_g_net = ((1 - substrate.rho_g_rad) * plant.tau_f * G_sol
               + epsilon_fg * sigma * (T_f ** 4 - T_g[0] ** 4))

    d0 = 0.701 * plant.H_f ** 0.975
    Z_M = 0.131 * plant.H_f ** 0.997
    Z_u = plant.H_f + 2.0

    log_r = np.log(max((plant.H_f - d0) / Z_M, 1.01)) / np.log(max((Z_u - d0) / Z_M, 1.01))
    u_f = u * max(log_r, 0.01)
    a_drag = (0.28 * plant.LAI * plant.H_f * plant.d_f) ** 0.5
    u_g = max(u_f * np.exp(-a_drag * (1 - 0.05 / max(plant.H_f, 0.01))), 0.01)

    r_c = 1 / (0.004 + 0.012 * u_g)
    r_vap = substrate_vapor_resistance(float(theta[0]), substrate)
    r_a = gr.compute_aerodynamic_resistance(u, plant)

    h_conv_g = plant.LAI * (rho_air * cp_air) / (r_a + r_c)
    h_eva_g = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_vap)

    q_top = R_g_net - h_conv_g * (T_g[0] - T_a_K) - h_eva_g * max(P_g_sat - P_a, 0.0)

    aw = np.zeros(Nz); ap = np.zeros(Nz); ae = np.zeros(Nz); bv = np.zeros(Nz)

    for i in range(1, Nz - 1):
        lw = 0.5 * (lambda_g[i - 1] + lambda_g[i])
        le = 0.5 * (lambda_g[i] + lambda_g[i + 1])
        aw[i] = lw / dz ** 2
        ae[i] = le / dz ** 2
        ap[i] = rho_cp_g[i] / dt + aw[i] + ae[i]
        bv[i] = rho_cp_g[i] / dt * T_g[i]

    ae[0] = lambda_g[0] / dz ** 2
    ap[0] = rho_cp_g[0] / dt + ae[0]
    bv[0] = rho_cp_g[0] / dt * T_g[0] + q_top / dz

    lam_e_bc = lambda_g[-1] / dz ** 2
    aw[-1] = lambda_g[-1] / dz ** 2
    ap[-1] = rho_cp_g[-1] / dt + aw[-1] + lam_e_bc
    bv[-1] = rho_cp_g[-1] / dt * T_g[-1] + lam_e_bc * T_slab_top

    return gr.tdma_solver(-aw, ap, -ae, bv)


def solve_substrate_moisture(
    theta: np.ndarray,
    T_f: float,
    T_g_surface: float,
    T_a_K: float,
    RH: float,
    u: float,
    G_sol: float,
    j_irrigation: float,
    plant: gr.PlantParameters,
    substrate: gr.SubstrateParameters,
    H_g: float,
    dt: float,
) -> Tuple[np.ndarray, float]:
    """Corrected moisture solver.

    G_sol is passed into stomatal resistance, so CAM daytime closure is active.
    j_irrigation is kg/m2/s water input. Rain flux should already be kg/m2/s.
    """
    gamma = gr.psychrometric_constant()
    rho_air = 1.2
    cp_air = 1005
    rho_water = 1000.0

    Nz = len(theta)
    dz = H_g / (Nz - 1)

    _, _, k_theta, D_theta = compute_substrate_properties(theta, substrate)

    P_a = gr.ambient_vapor_pressure(T_a_K, RH)
    P_f_sat = gr.saturation_pressure(T_f)
    P_g_sat = gr.saturation_pressure(T_g_surface)
    r_a = gr.compute_aerodynamic_resistance(u, plant)
    r_stoma = gr.compute_stomatal_resistance(
        G_sol, T_f, P_f_sat, P_a, float(np.mean(theta)), plant, substrate
    )
    r_vap = substrate_vapor_resistance(float(theta[0]), substrate)

    h_eva_f = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_stoma)
    h_eva_g = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_vap)

    j_eva_f = max(0.0, h_eva_f * (P_f_sat - P_a) / substrate.l_fg)
    j_eva_g = max(0.0, h_eva_g * (P_g_sat - P_a) / substrate.l_fg)
    j_eva = j_eva_f + j_eva_g

    j_eva_ms = j_eva / rho_water
    j_input_ms = j_irrigation / rho_water

    aw = np.zeros(Nz); ap = np.zeros(Nz); ae = np.zeros(Nz); bv = np.zeros(Nz)

    for i in range(1, Nz - 1):
        Dw = 0.5 * (D_theta[i - 1] + D_theta[i])
        De = 0.5 * (D_theta[i] + D_theta[i + 1])
        kw = 0.5 * (k_theta[i - 1] + k_theta[i])
        ke = 0.5 * (k_theta[i] + k_theta[i + 1])
        aw[i] = Dw / dz ** 2
        ae[i] = De / dz ** 2
        ap[i] = 1 / dt + aw[i] + ae[i]
        bv[i] = theta[i] / dt + (ke - kw) / dz

    j_net = j_input_ms - j_eva_ms
    ae[0] = D_theta[0] / dz ** 2
    ap[0] = 1 / dt + ae[0]
    bv[0] = theta[0] / dt + j_net / dz + k_theta[0] / dz

    if theta[-1] >= substrate.theta_sat * 0.99:
        aw[-1] = D_theta[-1] / dz ** 2
        ap[-1] = 1 / dt + aw[-1]
        bv[-1] = theta[-1] / dt - k_theta[-1] / dz
    else:
        aw[-1] = D_theta[-1] / dz ** 2
        ap[-1] = 1 / dt + aw[-1]
        bv[-1] = theta[-1] / dt

    theta_new = gr.tdma_solver(-aw, ap, -ae, bv)
    theta_new = np.clip(theta_new, substrate.theta_min, substrate.theta_sat)
    return theta_new, float(j_eva)


# ==============================================================================
# RESAMPLING / INITIAL CONDITIONS / SIMULATION
# ==============================================================================

def prepare_weather_1s(weather_df: pd.DataFrame, dt: float = 1.0) -> pd.DataFrame:
    rule = f"{int(dt)}s"
    states = weather_df[["T_a", "RH", "u", "G_sol"]].resample(rule).interpolate("time")

    if "rain_flux" in weather_df.columns:
        rain_flux = weather_df["rain_flux"].fillna(0).resample(rule).ffill().fillna(0)
    else:
        rain_flux = (weather_df["rain"].fillna(0) / 60.0).resample(rule).ffill().fillna(0)

    states["rain_flux"] = rain_flux
    return states


def _theta_initial_array(
    theta_initial: Optional[Union[float, dict, Sequence[float]]],
    H_g: float,
    Nz: int,
    substrate: gr.SubstrateParameters,
) -> np.ndarray:
    z = np.linspace(0, H_g, Nz)
    if theta_initial is None:
        return np.full(Nz, substrate.theta_sat * 0.80)
    if isinstance(theta_initial, dict):
        depths = np.asarray(theta_initial.get("depths", [0.02, 0.07]), dtype=float)
        values = np.asarray(theta_initial.get("values"), dtype=float)
        if values.size != depths.size:
            raise ValueError("theta_initial dict must contain same-length depths and values")
        return np.clip(np.interp(z, depths, values, left=values[0], right=values[-1]), substrate.theta_min, substrate.theta_sat)
    if isinstance(theta_initial, (list, tuple, np.ndarray)) and len(theta_initial) == 2:
        return np.clip(np.interp(z, [0.02, 0.07], list(theta_initial), left=theta_initial[0], right=theta_initial[1]), substrate.theta_min, substrate.theta_sat)
    return np.full(Nz, float(theta_initial))


def run_simulation(
    weather_df: pd.DataFrame,
    plant: gr.PlantParameters,
    substrate: gr.SubstrateParameters,
    slab: gr.SlabParameters,
    geom: gr.GeometryParameters,
    num: gr.NumericalParameters,
    theta_initial: Optional[Union[float, dict, Sequence[float]]] = None,
    j_irrigation: float = 0.0,
    T_in_series: Optional[pd.Series] = None,
    T_g_top_initial_C: Optional[float] = None,
    T_s_in_initial_C: Optional[float] = None,
    save_every_s: int = 60,
) -> dict:
    """Corrected simulation loop with datetime output and rain_flux handling."""
    Nz_g = num.Nz_substrate
    Nz_s = num.Nz_slab
    dt = num.dt
    if geom.H_g is None or slab.H_slab is None:
        raise ValueError("geom.H_g and slab.H_slab must be set")

    weather_1s = prepare_weather_1s(weather_df, dt=dt)
    N_steps = len(weather_1s)

    T_in_1s = None
    if T_in_series is not None:
        T_in_1s = T_in_series.sort_index().resample(f"{int(dt)}s").interpolate("time").reindex(weather_1s.index).interpolate("time")

    T_a_init = float(weather_1s["T_a"].iloc[0]) + 273.15
    T_g_top_init = (T_g_top_initial_C + 273.15) if T_g_top_initial_C is not None else T_a_init
    T_s_in_init = (T_s_in_initial_C + 273.15) if T_s_in_initial_C is not None else T_a_init

    T_g = np.linspace(T_g_top_init, T_a_init, Nz_g)
    T_s = np.linspace(T_a_init, T_s_in_init, Nz_s)
    T_f = T_a_init
    theta = _theta_initial_array(theta_initial, geom.H_g, Nz_g, substrate)

    results = {
        "datetime": [], "time": [], "T_f": [],
        "T_g_top": [], "T_g_mid": [], "T_g_bot": [], "T_s_in": [],
        "theta_top": [], "theta_mid": [], "theta_bot": [],
        "q_s_in": [], "j_eva": [], "T_a": [], "G_sol": [], "T_in_used": [],
    }

    for step, (ts, row) in enumerate(weather_1s.iterrows()):
        T_a_K = float(row["T_a"]) + 273.15
        G_sol = max(float(row["G_sol"]), 0.0)
        RH = float(np.clip(row["RH"], 1, 99))
        u = max(float(row["u"]), 0.1)
        j_pr = float(j_irrigation) + max(float(row.get("rain_flux", 0.0)), 0.0)

        if T_in_1s is not None and ts in T_in_1s.index and not pd.isna(T_in_1s.loc[ts]):
            T_in_current = float(T_in_1s.loc[ts]) + 273.15
        else:
            T_in_current = geom.T_in_default

        theta_avg = float(np.mean(theta))

        T_f = gr.solve_foliage_temperature(
            T_f_prev=T_f, T_a_K=T_a_K, T_g_surface=T_g[0],
            G_sol=G_sol, RH=RH, u=u, theta_avg=theta_avg,
            plant=plant, substrate=substrate, dt=dt,
        )

        T_g = solve_substrate_heat(
            T_g=T_g, T_f=T_f, T_slab_top=T_s[0],
            G_sol=G_sol, T_a_K=T_a_K, RH=RH, u=u, theta=theta,
            plant=plant, substrate=substrate, H_g=geom.H_g, dt=dt,
        )

        theta, j_eva = solve_substrate_moisture(
            theta=theta, T_f=T_f, T_g_surface=T_g[0], T_a_K=T_a_K,
            RH=RH, u=u, G_sol=G_sol, j_irrigation=j_pr,
            plant=plant, substrate=substrate, H_g=geom.H_g, dt=dt,
        )

        _, lambda_g, _, _ = compute_substrate_properties(theta, substrate)
        T_s, q_s_in = gr.solve_slab_heat(
            T_s=T_s, T_g_bottom=T_g[-1], lambda_g_bottom=lambda_g[-1],
            slab=slab, geom=geom, dt=dt, T_in_K=T_in_current,
        )

        if step % int(save_every_s / dt) == 0:
            results["datetime"].append(ts)
            results["time"].append(step * dt)
            results["T_f"].append(T_f - 273.15)
            results["T_g_top"].append(T_g[0] - 273.15)
            results["T_g_mid"].append(T_g[Nz_g // 2] - 273.15)
            results["T_g_bot"].append(T_g[-1] - 273.15)
            results["T_s_in"].append(T_s[-1] - 273.15)
            results["theta_top"].append(float(theta[0]))
            results["theta_mid"].append(float(theta[Nz_g // 2]))
            results["theta_bot"].append(float(theta[-1]))
            results["q_s_in"].append(float(q_s_in))
            results["j_eva"].append(float(j_eva))
            results["T_a"].append(T_a_K - 273.15)
            results["G_sol"].append(G_sol)
            results["T_in_used"].append(T_in_current - 273.15)

    results["Q_gain"] = float(np.trapezoid(results["q_s_in"], dx=save_every_s))
    return results


def series_from_results(results: dict, key: str) -> pd.Series:
    return pd.Series(results[key], index=pd.to_datetime(results["datetime"]), name=key)


def validation_metrics(sim: pd.Series, obs: pd.Series) -> Dict[str, float]:
    sim = sim.sort_index().resample("1min").mean()
    obs = obs.sort_index().resample("1min").mean()
    common = sim.index.intersection(obs.index)
    diff = sim.loc[common] - obs.loc[common]
    return {
        "n": int(diff.count()),
        "bias_C": float(diff.mean()),
        "mae_C": float(diff.abs().mean()),
        "rmse_C": float(np.sqrt((diff ** 2).mean())),
    }


# ==============================================================================
# CASE PREPARATION / RUNNERS
# ==============================================================================

def subset_by_window(df: pd.DataFrame, plant_type: str) -> pd.DataFrame:
    start, end = WINDOWS[plant_type.upper()]
    return df[(df.index >= start) & (df.index <= end)].copy()


def get_theta_initial_from_rika(rika_df: pd.DataFrame, start: pd.Timestamp) -> dict:
    """Get shallow/deep theta nearest to start time; returns profile dict."""
    # Resample to 1-min and interpolate small gaps only.
    r = rika_df.resample("1min").mean().interpolate("time", limit=10, limit_direction="both")
    # nearest not always available, use first valid after start if needed.
    local = r[(r.index >= start - pd.Timedelta("10min")) & (r.index <= start + pd.Timedelta("10min"))]
    if local[["theta_2cm", "theta_7cm"]].dropna().empty:
        local = r[r.index >= start].head(30)
    vals = local[["theta_2cm", "theta_7cm"]].dropna().iloc[0]
    return {"depths": [0.02, 0.07], "values": [float(vals["theta_2cm"]), float(vals["theta_7cm"])]}


def prepare_case_data(
    plant_type: str,
    base_dir: Union[str, Path] = "/mnt/data",
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load weather, NI, and RIKA data for the selected plant window."""
    plant_type = plant_type.upper()
    base_dir = Path(base_dir)
    start, end = WINDOWS[plant_type]

    weather = load_weather_cache_or_excel(
        base_dir / "weatherfile mar-april.xlsx",
        base_dir / "weather_clean_cache.csv",
        date_start=start,
        date_end=end,
    )

    ni_files = [
        base_dir / "Pengukuran 30_1 Maret 2026.xlsx",
        base_dir / "Pengukuran 30_2 Maret 2026.xlsx",
        base_dir / "NI_sensor_data_clean.xlsx",
    ]
    ni = load_multiple_NI_sensor_data(ni_files)
    ni = ni[(ni.index >= start) & (ni.index <= end)].copy()

    rika = load_rika_pair(base_dir / "datasoilmoisture.zip", plant_type)
    rika = rika[(rika.index >= start) & (rika.index <= end)].copy()

    return weather, ni, rika


def run_validation_case(
    plant_type: str,
    base_dir: Union[str, Path] = "/mnt/data",
    calibrate_lai: bool = False,
    lai_bounds: Tuple[float, float] = (0.3, 3.5),
) -> Tuple[dict, Dict[str, float]]:
    """Run one plant validation case and return results + T_s,in metrics.

    If calibrate_lai=False, uses the current LAI value from new_baru.py.
    """
    apply_scientific_guess_parameters()

    plant_type = plant_type.upper()
    plant = gr.bromelia if plant_type == "CAM" else gr.wedelia
    weather, ni, rika = prepare_case_data(plant_type, base_dir)
    start, _end = WINDOWS[plant_type]

    theta_initial = get_theta_initial_from_rika(rika, start)

    if plant_type == "CAM":
        T_in_series = ni["T_in_CAM"]
        target_col = "T_s_in_CAM"
        T_g_top_initial = float(ni["T_g_top_CAM"].dropna().iloc[0]) if "T_g_top_CAM" in ni else None
    else:
        T_in_series = ni["T_in_C3"]
        target_col = "T_s_in_C3"
        T_g_top_initial = None

    T_s_in_initial = float(ni[target_col].dropna().iloc[0])

    if calibrate_lai:
        from scipy.optimize import minimize_scalar

        def obj(lai: float) -> float:
            plant.LAI = float(lai)
            res = run_simulation(
                weather_df=weather,
                plant=plant,
                substrate=gr.substrat,
                slab=gr.slab,
                geom=gr.geom,
                num=gr.num,
                theta_initial=theta_initial,
                T_in_series=T_in_series,
                T_g_top_initial_C=T_g_top_initial,
                T_s_in_initial_C=T_s_in_initial,
            )
            m = validation_metrics(series_from_results(res, "T_s_in"), ni[target_col])
            return m["rmse_C"]

        opt = minimize_scalar(obj, bounds=lai_bounds, method="bounded", options={"xatol": 0.03})
        plant.LAI = float(opt.x)

    results = run_simulation(
        weather_df=weather,
        plant=plant,
        substrate=gr.substrat,
        slab=gr.slab,
        geom=gr.geom,
        num=gr.num,
        theta_initial=theta_initial,
        T_in_series=T_in_series,
        T_g_top_initial_C=T_g_top_initial,
        T_s_in_initial_C=T_s_in_initial,
    )
    metrics = validation_metrics(series_from_results(results, "T_s_in"), ni[target_col])
    metrics["LAI_used"] = float(plant.LAI)
    metrics["window_start"] = str(WINDOWS[plant_type][0])
    metrics["window_end"] = str(WINDOWS[plant_type][1])
    return results, metrics


# ==============================================================================
# PATCH new_baru NAMESPACE FOR INTERACTIVE USE
# ==============================================================================

def apply_patches_to_new_baru_namespace() -> None:
    gr.sky_temperature = sky_temperature
    gr.compute_substrate_properties = compute_substrate_properties
    gr.solve_substrate_heat = solve_substrate_heat
    gr.solve_substrate_moisture = solve_substrate_moisture
    gr.run_simulation = run_simulation


apply_patches_to_new_baru_namespace()


if __name__ == "__main__":
    print("Patched green-roof model loaded.")
    print("Main windows:")
    for k, (a, b) in WINDOWS.items():
        print(f"  {k}: {a} -> {b} ({(b-a).total_seconds()/3600:.1f} h)")
    print("\nExample:")
    print("  import green_roof_fixed_runner as fx")
    print("  res_cam, m_cam = fx.run_validation_case('CAM', calibrate_lai=False)")
    print("  res_c3,  m_c3  = fx.run_validation_case('C3',  calibrate_lai=False)")
