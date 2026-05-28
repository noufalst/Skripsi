"""
================================================================================
CONTROL CABIN / REFERENCE ROOF — SLAB-ONLY VALIDATION
================================================================================
Tujuan:
    Membandingkan model reference roof / kabin kontrol slab beton 10 cm
    terhadap data aktual NI sensor atap indoor RR (T2Ka).

Model:
    1D transient conduction pada slab beton.
    Top boundary     : solar absorption + convection to ambient + longwave to sky
    Bottom boundary  : convection to cabin/interior air proxy

Acuan struktur:
    Dibuat dengan format section yang mirip new_baru.py / new_baru_revised_same_structure_v2.py
    supaya gampang dibandingkan dan dipahami.

Run langsung:
    py control_cabin_slab_only.py

Output:
    - metrics_control_slab.csv
    - validation_CONTROL_SLAB_CAM_WINDOW.png
    - validation_CONTROL_SLAB_C3_WINDOW.png
================================================================================
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Tuple, Optional, Sequence
from openpyxl import load_workbook
import zipfile
import warnings

warnings.filterwarnings("ignore")

# ==============================================================================
# SECTION 1: PARAMETERS
# ==============================================================================

@dataclass
class SlabOnlyParameters:
    """Parameter slab beton reference roof / kabin kontrol."""
    lambda_s: float = 1.74       # W/mK, cast concrete dari paper acuan
    rho_s: float = 2300.0        # kg/m3, cast concrete
    cp_s: float = 840.0          # J/kgK, cast concrete
    H_slab: float = 0.10         # m, slab 10 cm
    alpha_s: float = 0.40        # absorptivitas shortwave concrete
    epsilon_s: float = 0.82      # emissivitas longwave concrete
    h_in: float = 8.0            # W/m2K, convection underside to cabin air
    A_roof: float = 1.0          # m2, 100 cm x 100 cm. Tidak memengaruhi T, hanya Q_total.

@dataclass
class NumericalParameters:
    """Parameter numerik."""
    dt: float = 1.0              # s
    Nz_slab: int = 67            # mengikuti paper RR/slab
    save_every_s: int = 60       # simpan per menit

slab = SlabOnlyParameters()
num = NumericalParameters()

# Window default sesuai yang sedang dipakai untuk validasi GR.
VALIDATION_WINDOWS = {
    "CAM_WINDOW": (pd.Timestamp("2026-03-31 11:58:00"), pd.Timestamp("2026-04-02 21:42:00")),
    "C3_WINDOW":  (pd.Timestamp("2026-04-09 11:05:00"), pd.Timestamp("2026-04-10 14:08:00")),
}

# ==============================================================================
# SECTION 2: DATA LOADERS
# ==============================================================================

def _to_float(x):
    if x is None:
        return np.nan
    if isinstance(x, str):
        x = x.strip()
        if x in {"", "---", "--"}:
            return np.nan
    return pd.to_numeric(x, errors="coerce")


def load_weather_data(filepath: str,
                      date_start: Optional[str] = None,
                      date_end: Optional[str] = None,
                      sheet_name: str = "3-24april") -> pd.DataFrame:
    """
    Load weatherfile mar-april.xlsx.

    Output:
        T_a   [°C]
        RH    [%]
        u     [m/s atau sesuai logger; dipakai sebagai wind proxy]
        rain  [mm per interval]
        G_sol [W/m2]
    """
    filepath = str(filepath)
    start = pd.Timestamp(date_start) if date_start else None
    end = pd.Timestamp(date_end) if date_end else None

    print(f"Loading weather: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    records = []
    # Data mulai baris 4 pada file Davis export.
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        ts = pd.to_datetime(f"{row[0]} {row[1]}", dayfirst=True, errors="coerce")
        if pd.isna(ts):
            continue
        if start is not None and ts < start:
            continue
        if end is not None and ts > end:
            continue
        records.append({
            "timestamp": ts,
            "T_a": _to_float(row[2]),
            "RH": _to_float(row[5]),
            "u": _to_float(row[7]),
            "rain": _to_float(row[17]),
            "G_sol": _to_float(row[19]),
        })

    if not records:
        raise ValueError("Tidak ada weather data pada window yang diminta.")

    df = pd.DataFrame.from_records(records).set_index("timestamp").sort_index()
    before = len(df)
    df = df[~df.index.duplicated(keep="first")]
    if before - len(df) > 0:
        print(f"  Duplicate timestamp removed: {before-len(df)}")

    df["T_a"] = pd.to_numeric(df["T_a"], errors="coerce")
    df["RH"] = pd.to_numeric(df["RH"], errors="coerce").clip(0, 100)
    df["u"] = pd.to_numeric(df["u"], errors="coerce").clip(lower=0.1)
    df["rain"] = pd.to_numeric(df["rain"], errors="coerce").fillna(0).clip(lower=0)
    df["G_sol"] = pd.to_numeric(df["G_sol"], errors="coerce").fillna(0).clip(lower=0)

    for col in ["T_a", "RH", "u", "G_sol"]:
        df[col] = df[col].interpolate(method="time", limit=30, limit_direction="both")

    df = df.dropna(subset=["T_a", "RH", "u", "G_sol"])
    print(f"  Weather period: {df.index[0]} → {df.index[-1]} | {len(df)} rows")
    return df


def load_weather_cache_or_excel(filepath: str,
                                cache_path: str = "weather_clean_cache.csv",
                                date_start: Optional[str] = None,
                                date_end: Optional[str] = None) -> pd.DataFrame:
    """Load weather dari cache bila ada; kalau belum ada, parse Excel lalu cache full file."""
    cache = Path(cache_path)
    if cache.exists():
        df = pd.read_csv(cache, parse_dates=["timestamp"]).set_index("timestamp").sort_index()
        if date_start:
            df = df[df.index >= pd.Timestamp(date_start)]
        if date_end:
            df = df[df.index <= pd.Timestamp(date_end)]
        return df

    full = load_weather_data(filepath)
    full.to_csv(cache, index_label="timestamp")
    if date_start:
        full = full[full.index >= pd.Timestamp(date_start)]
    if date_end:
        full = full[full.index <= pd.Timestamp(date_end)]
    return full


def load_NI_sensor_data(filepath: str) -> pd.DataFrame:
    """
    Load raw NI LabVIEW Excel.

    Timestamp adalah LabVIEW serial days since 1904-01-01.
    Target RR/control cabin:
        T2Ka = Atap Indoor RR / Reference Roof inner surface.

    Interior proxy untuk boundary bawah:
        default di script ini memakai mean(T2Kd, T2Kc), yaitu dinding timur/barat RR,
        karena tidak ada channel khusus ruangan RR yang terkonfirmasi.
    """
    filepath = str(filepath)
    print(f"Loading NI raw: {filepath}")

    with zipfile.ZipFile(filepath, "r") as z:
        with z.open("xl/worksheets/sheet1.xml") as f:
            xml_content = f.read()

    tree = ET.fromstring(xml_content)
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rows_xml = tree.findall(f".//{ns}row")

    ni_channel_names = [
        "timestamp_serial",
        "T1Kd", "T1Kb", "T1Ke", "T1Ka", "T1Kc",
        "T3Kb", "T3Kd", "T3Ke", "T3Kc", "T3Ka",
        "T2Ka", "T2Kd", "T2Kc", "2Ke",  "T1A",
        "T2A",  "T2A2", "T1Tb", "T1Ta", "T1Ta2",
    ]

    data = []
    for row in rows_xml[1:]:
        vals = []
        for cell in row.findall(f"{ns}c"):
            v = cell.find(f"{ns}v")
            vals.append(float(v.text) if v is not None else np.nan)
        if len(vals) == 21:
            data.append(vals)

    if not data:
        raise ValueError(f"Tidak ada row numerik terbaca dari {filepath}")

    df = pd.DataFrame(data, columns=ni_channel_names)
    labview_epoch = pd.Timestamp("1904-01-01")
    df["timestamp"] = labview_epoch + pd.to_timedelta(df["timestamp_serial"], unit="D")
    df = df.set_index("timestamp").drop(columns=["timestamp_serial"]).sort_index()

    # Basic anomaly filter semua channel suhu: nilai ekstrem tidak fisik -> NaN, interpolasi gap pendek.
    temp_cols = [c for c in df.columns if c.startswith("T") or c in ["2Ke"]]
    for col in temp_cols:
        bad = (df[col] < -10) | (df[col] > 90)
        if bad.any():
            print(f"  {col}: mask anomaly {int(bad.sum())} rows")
            df.loc[bad, col] = np.nan
            df[col] = df[col].interpolate(method="time", limit=30, limit_direction="both")

    # Alias control cabin / reference roof.
    df["T_r_in_RR"] = df["T2Ka"]
    df["T_in_RR_proxy"] = df[["T2Kd", "T2Kc"]].mean(axis=1)
    return df


def load_multiple_NI_sensor_data(filepaths: Sequence[str]) -> pd.DataFrame:
    frames = []
    for fp in filepaths:
        p = Path(fp)
        if p.exists():
            frames.append(load_NI_sensor_data(str(p)))
        else:
            print(f"Skipping missing NI file: {p}")
    if not frames:
        raise FileNotFoundError("Tidak ada file NI raw yang ditemukan.")
    df = pd.concat(frames).sort_index()
    df = df[~df.index.duplicated(keep="first")]
    print(f"NI combined: {df.index[0]} → {df.index[-1]} | {len(df)} rows")
    return df

# ==============================================================================
# SECTION 3: THERMODYNAMIC HELPERS
# ==============================================================================

def dew_point_C(T_C: float, RH: float) -> float:
    RH = np.clip(RH, 1.0, 100.0)
    a, b = 17.27, 237.7
    gamma = (a*T_C)/(b+T_C) + np.log(RH/100.0)
    return (b*gamma)/(a - gamma)


def sky_temperature(T_a_K: float, RH: float) -> float:
    """Effective sky temperature pakai dew-point-based emissivity."""
    T_C = T_a_K - 273.15
    Tdp = dew_point_C(T_C, RH)
    eps_sky = 0.711 + 0.56*(Tdp/100.0) + 0.73*(Tdp/100.0)**2
    eps_sky = float(np.clip(eps_sky, 0.60, 1.00))
    return eps_sky**0.25 * T_a_K


def exterior_convection_coefficient(u: float) -> float:
    """
    Simple exterior convection coefficient for roof surface.
    Correlation umum building surface: h = 5.7 + 3.8u.
    Dipakai sebagai baseline cepat untuk RR slab-only.
    """
    u = max(float(u), 0.1)
    return 5.7 + 3.8*u

# ==============================================================================
# SECTION 4: NUMERICAL SOLVER
# ==============================================================================

def tdma_solver(a: np.ndarray, b: np.ndarray, c: np.ndarray, d: np.ndarray) -> np.ndarray:
    """Thomas algorithm untuk tridiagonal system."""
    n = len(d)
    cp = np.zeros(n)
    dp = np.zeros(n)
    x = np.zeros(n)

    denom = b[0]
    if abs(denom) < 1e-30:
        denom = 1e-30
    cp[0] = c[0] / denom
    dp[0] = d[0] / denom

    for i in range(1, n):
        denom = b[i] - a[i]*cp[i-1]
        if abs(denom) < 1e-30:
            denom = 1e-30
        cp[i] = c[i] / denom if i < n-1 else 0.0
        dp[i] = (d[i] - a[i]*dp[i-1]) / denom

    x[-1] = dp[-1]
    for i in range(n-2, -1, -1):
        x[i] = dp[i] - cp[i]*x[i+1]
    return x

# ==============================================================================
# SECTION 5: SLAB-ONLY REFERENCE ROOF MODEL
# ==============================================================================

def solve_reference_slab_step(T: np.ndarray,
                              T_a_K: float,
                              RH: float,
                              G_sol: float,
                              u: float,
                              T_in_K: float,
                              pars: SlabOnlyParameters,
                              dt: float) -> Tuple[np.ndarray, float, float, float]:
    """
    One implicit step untuk slab beton.

    Coordinate:
        index 0  = exterior/top roof surface
        index -1 = interior/inner roof surface

    Returns:
        T_new [K]
        q_in [W/m2] positive if heat goes into cabin/interior
        q_top [W/m2] positive into slab
        h_ext [W/m2K]
    """
    N = len(T)
    H = pars.H_slab
    dz = H / (N - 1)
    lam = pars.lambda_s
    rho_cp = pars.rho_s * pars.cp_s
    sigma = 5.670374419e-8

    T_sky = sky_temperature(T_a_K, RH)
    h_ext = exterior_convection_coefficient(u)

    # Top surface energy input, explicit berdasarkan temperatur top lama.
    T_top_old = T[0]
    q_solar = pars.alpha_s * G_sol
    q_conv = h_ext * (T_top_old - T_a_K)              # loss if surface warmer
    q_lw = pars.epsilon_s * sigma * (T_top_old**4 - T_sky**4)  # loss if surface warmer than sky
    q_top = q_solar - q_conv - q_lw

    a = np.zeros(N)  # lower diag
    b = np.zeros(N)  # main diag
    c = np.zeros(N)  # upper diag
    d = np.zeros(N)  # rhs

    # Top half-cell with prescribed net flux q_top into slab.
    b[0] = rho_cp/dt + 2*lam/dz**2
    c[0] = -2*lam/dz**2
    d[0] = rho_cp/dt*T[0] + 2*q_top/dz

    # Interior nodes.
    for i in range(1, N-1):
        a[i] = -lam/dz**2
        b[i] = rho_cp/dt + 2*lam/dz**2
        c[i] = -lam/dz**2
        d[i] = rho_cp/dt*T[i]

    # Bottom half-cell with convection to cabin/interior air.
    # q_in = h_in*(T_bottom - T_in), positive leaving slab into cabin.
    a[-1] = -2*lam/dz**2
    b[-1] = rho_cp/dt + 2*lam/dz**2 + 2*pars.h_in/dz
    d[-1] = rho_cp/dt*T[-1] + 2*pars.h_in*T_in_K/dz

    T_new = tdma_solver(a, b, c, d)
    q_in = pars.h_in * (T_new[-1] - T_in_K)
    return T_new, float(q_in), float(q_top), float(h_ext)


def run_reference_slab(weather_df: pd.DataFrame,
                       ni_df: pd.DataFrame,
                       pars: SlabOnlyParameters = slab,
                       numpars: NumericalParameters = num,
                       T_in_strategy: str = "rr_wall_mean",
                       T_in_fixed_C: float = 29.5,
                       T_top_initial_C: Optional[float] = None,
                       T_bottom_initial_C: Optional[float] = None) -> Dict:
    """
    Run model reference roof slab-only.

    T_in_strategy:
        "rr_wall_mean" : memakai mean(T2Kd,T2Kc) sebagai proxy suhu kabin RR.
        "fixed"        : memakai T_in_fixed_C konstan.

    Catatan:
        Untuk validasi T2Ka, T_in tidak boleh T2Ka karena itu target output.
    """
    dt = float(numpars.dt)
    save_every_s = int(numpars.save_every_s)
    N = int(numpars.Nz_slab)

    # Resample weather ke 1 detik. State variables boleh interpolasi.
    weather_1s = weather_df[["T_a", "RH", "u", "G_sol"]].resample("1s").interpolate("time")

    # Interior temperature boundary.
    if T_in_strategy == "rr_wall_mean" and "T_in_RR_proxy" in ni_df.columns:
        T_in_series_C = ni_df["T_in_RR_proxy"].resample("1s").interpolate("time")
        T_in_series_C = T_in_series_C.reindex(weather_1s.index).interpolate("time").ffill().bfill()
    elif T_in_strategy == "fixed":
        T_in_series_C = pd.Series(T_in_fixed_C, index=weather_1s.index)
    else:
        raise ValueError("T_in_strategy harus 'rr_wall_mean' atau 'fixed'.")

    # Initial condition.
    if T_top_initial_C is None:
        T_top_initial_C = float(weather_df["T_a"].iloc[0])
    if T_bottom_initial_C is None:
        if "T_r_in_RR" in ni_df and not ni_df["T_r_in_RR"].dropna().empty:
            T_bottom_initial_C = float(ni_df["T_r_in_RR"].dropna().iloc[0])
        else:
            T_bottom_initial_C = float(weather_df["T_a"].iloc[0])

    T = np.linspace(T_top_initial_C + 273.15, T_bottom_initial_C + 273.15, N)

    results = {
        "datetime": [],
        "time": [],
        "T_r_ext": [],
        "T_r_in": [],
        "T_a": [],
        "G_sol": [],
        "T_in_used": [],
        "q_r_in": [],
        "q_top": [],
        "h_ext": [],
    }

    print(f"\nRunning RR slab-only: {weather_1s.index[0]} → {weather_1s.index[-1]}")
    print(f"  H_slab={pars.H_slab:.3f} m | lambda={pars.lambda_s:.2f} W/mK | Nz={N} | dt={dt}s")
    print(f"  T_in_strategy={T_in_strategy}")

    for step, (ts, row) in enumerate(weather_1s.iterrows()):
        T_a_K = float(row["T_a"] + 273.15)
        RH = float(row["RH"])
        u = float(row["u"])
        G_sol = float(row["G_sol"])
        T_in_K = float(T_in_series_C.loc[ts] + 273.15)

        T, q_in, q_top, h_ext = solve_reference_slab_step(
            T, T_a_K, RH, G_sol, u, T_in_K, pars, dt
        )

        if step % int(save_every_s/dt) == 0:
            results["datetime"].append(ts)
            results["time"].append(step*dt)
            results["T_r_ext"].append(T[0] - 273.15)
            results["T_r_in"].append(T[-1] - 273.15)
            results["T_a"].append(row["T_a"])
            results["G_sol"].append(G_sol)
            results["T_in_used"].append(T_in_K - 273.15)
            results["q_r_in"].append(q_in)
            results["q_top"].append(q_top)
            results["h_ext"].append(h_ext)

        if step % int(6*3600/dt) == 0:
            print(f"  t={step/3600:5.1f}h | T_ext={T[0]-273.15:5.1f}°C | "
                  f"T_in_surf={T[-1]-273.15:5.1f}°C | q_in={q_in:7.1f} W/m2")

    results["Q_gain"] = float(np.trapezoid(results["q_r_in"], dx=save_every_s))
    results["Q_gain_total"] = results["Q_gain"] * pars.A_roof
    print(f"  Q_gain={results['Q_gain']:.1f} J/m2 | Q_total={results['Q_gain_total']:.1f} J")
    return results

# ==============================================================================
# SECTION 6: VALIDATION METRICS + PLOT
# ==============================================================================

def series_from_results(results: Dict, key: str) -> pd.Series:
    return pd.Series(results[key], index=pd.to_datetime(results["datetime"]), name=key)


def validation_metrics(sim: pd.Series, obs: pd.Series) -> Dict[str, float]:
    sim = sim.sort_index().resample("1min").mean()
    obs = obs.sort_index().resample("1min").mean()
    common = sim.index.intersection(obs.index)
    diff = sim.loc[common] - obs.loc[common]
    return {
        "n": int(diff.count()),
        "bias_C": float(diff.mean()),
        "mae_C": float(diff.abs().mean()),
        "rmse_C": float(np.sqrt((diff**2).mean())),
        "amp_measured_C": float(obs.loc[common].max() - obs.loc[common].min()),
        "amp_model_C": float(sim.loc[common].max() - sim.loc[common].min()),
        "amp_error_C": float((sim.loc[common].max() - sim.loc[common].min()) - (obs.loc[common].max() - obs.loc[common].min())),
        "peak_error_C": float(sim.loc[common].max() - obs.loc[common].max()),
        "min_error_C": float(sim.loc[common].min() - obs.loc[common].min()),
    }


def plot_reference_validation(results: Dict,
                              ni_df: pd.DataFrame,
                              label: str,
                              save_path: Optional[str] = None):
    sim = series_from_results(results, "T_r_in").resample("1min").mean()
    obs = ni_df["T_r_in_RR"].sort_index().resample("1min").mean()
    common = sim.index.intersection(obs.index)
    sim_c = sim.loc[common]
    obs_c = obs.loc[common]
    err = sim_c - obs_c
    m = validation_metrics(sim, obs)

    fig, axes = plt.subplots(3, 1, figsize=(14, 9), sharex=True)

    axes[0].plot(obs_c.index, obs_c.values, label="Measured NI: T2Ka (RR inner roof)", linewidth=2)
    axes[0].plot(sim_c.index, sim_c.values, label="Model: slab-only RR", linestyle="--", linewidth=2)
    axes[0].set_ylabel("T_r,in (°C)")
    axes[0].set_title(
        f"Control Cabin / Reference Roof Slab-only Validation — {label}\n"
        f"Bias={m['bias_C']:.2f}°C | MAE={m['mae_C']:.2f}°C | RMSE={m['rmse_C']:.2f}°C | "
        f"AmpErr={m['amp_error_C']:.2f}°C | n={m['n']}"
    )
    axes[0].legend()
    axes[0].grid(True, alpha=0.3)

    axes[1].plot(err.index, err.values, linewidth=1.5)
    axes[1].axhline(0, linestyle="--", linewidth=1)
    axes[1].set_ylabel("Error (°C)")
    axes[1].set_title("Residual: Model - Measured")
    axes[1].grid(True, alpha=0.3)

    time = pd.to_datetime(results["datetime"])
    axes[2].plot(time, results["T_a"], label="T_a", linewidth=1.5)
    axes[2].plot(time, results["T_in_used"], label="T_in proxy", linewidth=1.2)
    axes[2].set_ylabel("Temperature (°C)")
    axes[2].set_xlabel("Datetime")
    axes[2].grid(True, alpha=0.3)

    ax2 = axes[2].twinx()
    ax2.plot(time, results["G_sol"], label="G_sol", linestyle="--", linewidth=1.2)
    ax2.set_ylabel("G_sol (W/m²)")

    lines1, labels1 = axes[2].get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    axes[2].legend(lines1 + lines2, labels1 + labels2, loc="upper right")

    plt.tight_layout()
    if save_path:
        plt.savefig(save_path, dpi=200, bbox_inches="tight")
        print(f"Saved plot: {save_path}")
    plt.show()
    return m

# ==============================================================================
# SECTION 7: CASE PREPARATION + MAIN
# ==============================================================================

def prepare_control_case(case_key: str,
                         base_dir: str = ".") -> Tuple[pd.DataFrame, pd.DataFrame, Tuple[pd.Timestamp, pd.Timestamp]]:
    case_key = case_key.upper()
    start, end = VALIDATION_WINDOWS[case_key]
    base = Path(base_dir)

    weather = load_weather_cache_or_excel(
        str(base / "weatherfile mar-april.xlsx"),
        cache_path=str(base / "weather_clean_cache.csv"),
        date_start=str(start),
        date_end=str(end),
    )

    # Pakai raw NI agar proses cleaning reproducible.
    ni_files = [
        str(base / "Pengukuran 30_1 Maret 2026.xlsx"),
        str(base / "Pengukuran 30_2 Maret 2026.xlsx"),
        str(base / "Pengukuran 3 April 2026.xlsx"),
    ]
    ni = load_multiple_NI_sensor_data(ni_files)
    ni = ni[(ni.index >= start) & (ni.index <= end)].copy()

    if ni.empty:
        raise ValueError(f"NI kosong untuk {case_key}: {start} → {end}")
    if ni["T_r_in_RR"].dropna().empty:
        raise ValueError(f"T2Ka/T_r_in_RR kosong untuk {case_key}")

    return weather, ni, (start, end)


def run_control_validation(case_key: str,
                           base_dir: str = ".",
                           T_in_strategy: str = "rr_wall_mean") -> Tuple[Dict, Dict[str, float]]:
    weather, ni, (start, end) = prepare_control_case(case_key, base_dir)
    results = run_reference_slab(
        weather_df=weather,
        ni_df=ni,
        pars=slab,
        numpars=num,
        T_in_strategy=T_in_strategy,
        T_bottom_initial_C=float(ni["T_r_in_RR"].dropna().iloc[0]),
    )
    metrics = validation_metrics(series_from_results(results, "T_r_in"), ni["T_r_in_RR"])
    metrics.update({
        "case": case_key,
        "window_start": str(start),
        "window_end": str(end),
        "H_slab_m": slab.H_slab,
        "lambda_s_W_mK": slab.lambda_s,
        "T_in_strategy": T_in_strategy,
    })
    return results, metrics


if __name__ == "__main__":
    print("="*70)
    print("CONTROL CABIN / REFERENCE ROOF — SLAB ONLY VALIDATION")
    print("="*70)

    base_dir = "."
    all_metrics = []

    for case in ["CAM_WINDOW", "C3_WINDOW"]:
        print(f"\n=== RUNNING {case} ===")
        res, metrics = run_control_validation(case, base_dir=base_dir, T_in_strategy="rr_wall_mean")
        all_metrics.append(metrics)

        # reload ni window for plot
        _, ni_plot, _ = prepare_control_case(case, base_dir=base_dir)
        plot_reference_validation(
            res, ni_plot, label=case,
            save_path=f"validation_CONTROL_SLAB_{case}.png"
        )

        print("Metrics:")
        for k, v in metrics.items():
            print(f"  {k}: {v}")

    metrics_df = pd.DataFrame(all_metrics)
    metrics_df.to_csv("metrics_control_slab.csv", index=False)
    print("\nSaved metrics: metrics_control_slab.csv")
    print("Done.")
