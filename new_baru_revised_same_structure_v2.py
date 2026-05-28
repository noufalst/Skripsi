"""
================================================================================
GREEN ROOF THERMAL SIMULATION — CORRECTED VERSION
================================================================================
Referensi utama:
    Chagolla-Aranda et al. (2025), Journal of Building Engineering 103, 112053

Tanaman: Bromelia (CAM) vs Wedelia (C3)
Substrat: Tanah + Sekam Padi
Lokasi: Universitas Indonesia, Jakarta

DAFTAR KOREKSI dari versi sebelumnya:
─────────────────────────────────────────────────────────────────────────────
[C1] T_in TIDAK konstan 25°C
     → Asumsi paper tidak berlaku untuk boks outdoor UI
     → T_in diambil dinamis dari NI sensor: T1Ka (CAM) atau T3Kd (C3)
     → Data NI menunjukkan T_in = 24.8–42.8°C (mean 29.5°C)

[C2] tau_f DIPERBARUI berdasarkan morfologi daun
     → Wedelia C3 (daun tipis): tau_f = 0.20 (sebelumnya 0.10)
     → Bromelia CAM (daun tebal+lilin): tau_f = 0.07 (sebelumnya 0.07 sama)
     → Referensi: PROSPECT-D model + CAM vs C3 leaf thickness scaling

[C3] alpha_f DIHITUNG ULANG dari rho_f terukur
     → Wedelia:  alpha_f = 1 - 0.438 - 0.10 = 0.462
     → Bromelia: alpha_f = 1 - 0.390 - 0.07 = 0.540

[C4] r_stoma_min C3 dari data Licor 6800 NYATA
     → r_stoma_min C3 = 167.2 s/m (percentile 5%, data April 2026)
     → Sebelumnya: nilai dummy 150 s/m

[C5] NI data loader menggunakan XML parser
     → extract-text truncate di 1000 baris → data hilang
     → XML parser langsung dari file .xlsx → semua 9032 baris terbaca

[C6] T2A (Tanah bawah CAM) anomali difilter
     → 66 data dengan nilai negatif ekstrem (min -67 miliar °C)
     → Penyebab: sensor disconnect sementara
     → Solusi: replace dengan NaN lalu interpolasi linear

[C7] Sinkronisasi timestamp Davis ↔ NI sensor
     → Davis: interval 1 menit (xx:00:00)
     → NI: interval 1 menit tapi dari LabVIEW (ada detik yang tidak tepat)
     → Solusi: resample keduanya ke grid waktu yang sama

[C8] r_stoma_min CAM masih PLACEHOLDER
     → File Licor CAM belum diupload
     → Nilai sementara: 400.0 s/m (perlu dikonfirmasi dari Licor CAM)
================================================================================
"""

import numpy as np
import pandas as pd
import xml.etree.ElementTree as ET
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Tuple, Optional, Dict, Sequence, Union
import zipfile
import io
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

warnings.filterwarnings('ignore')


# ==============================================================================
# SECTION 1: DATA CLASSES
# ==============================================================================

@dataclass
class PlantParameters:
    """
    Parameter spesifik tanaman.
    CAM (Bromelia): stomata terbuka malam, tertutup siang
    C3  (Wedelia) : stomata terbuka siang, tertutup malam
    """
    name: str
    plant_type: str              # "CAM" atau "C3"

    # Geometri kanopi
    H_f: float = None            # m — tinggi kanopi [REAL: dari pengukuran]
    d_f: float = None            # m — lebar daun [TODO: ukur dari foto]
    LAI: float = None            # m²/m² [REAL: dari data setahun lalu, dikalibrasi]

    # Properti optik daun
    # [C2] tau_f DIPERBARUI: C3 lebih tinggi dari CAM karena daun lebih tipis
    # [C3] alpha_f DIHITUNG dari rho_f terukur (bukan nilai paper)
    alpha_f: float = None        # absorptivitas (-)
    epsilon_f: float = 0.95      # emissivitas — standar semua daun hijau, tidak berubah
    tau_f: float = None          # transmisivitas (-)
    rho_f: float = None          # reflektifitas — dari pengukuran langsung

    # Stomatal resistance
    # [C4] r_stoma_min C3 dari Licor NYATA = 167.2 s/m
    # [C8] r_stoma_min CAM masih PLACEHOLDER = 400.0 s/m
    r_stoma_min: float = None    # s/m
    cover_fraction: float = 1.0  # projected canopy cover fraction, 0-1

@dataclass
class SubstrateParameters:
    """Parameter substrat: tanah + sekam padi."""
    # Properti termal
    lambda_dry: float = None     # W/mK  [LIT/EST: tanah+sekam padi kering]
    lambda_sat: float = None     # W/mK  [EST] substrat kondisi jenuh; dihitung/diuji sensitivitas
    rho_g: float = None          # kg/m³ [LAB: specific gravity test]
    cp_g: float = None           # J/kgK [LIT: 1300]

    # Properti hidraulik
    theta_sat: float = None      # m³/m³ [LAB: water content test]
    k_theta_sat: float = None    # m/s   [LAB: permeability test] -> [Coba cari di lit siapa tau ada]
    psi_sat: float = None        # m     [LIT: 0.35]
    b: float = None              # (-)   [LIT: 5.5]
    theta_min: float = 0.05      # m³/m³ — wilting point, default OK

    # Properti optik permukaan substrat
    rho_g_rad: float = 0.15      # reflektifitas — standar
    epsilon_g: float = 0.95      # emissivitas — standar

    # Properti air — konstanta fisika
    lambda_water: float = 0.60   # W/mK
    cp_water: float = 4180       # J/kgK
    l_fg: float = 2.45e6         # J/kg — latent heat


@dataclass
class SlabParameters:
    """Parameter slab beton — semua nilai standar kecuali H_slab."""
    lambda_s: float = 1.74       # W/mK  — standar beton
    rho_s: float = 2300.0        # kg/m³ — standar beton
    cp_s: float = 840.0          # J/kgK — standar beton
    alpha_s: float = 0.40        # absorptivitas (untuk RR model)
    epsilon_s: float = 0.82      # emissivitas (untuk RR model)
    H_slab: float = None         # m — UKUR LANGSUNG di boks


@dataclass
class GeometryParameters:
    """
    Parameter geometri dan kondisi batas.
    [C1] T_in sekarang DINAMIS dari data NI sensor, bukan konstan 25°C.
    """
    H_g: float = None            # m — tebal substrat [UKUR LANGSUNG]
    H_slab: float = None         # m — tebal slab/atap arah vertikal [UKUR LANGSUNG]
    A_roof: float = None         # m² — luas plan atap; hanya untuk Q_total, bukan q_s_in per m²

    # [C1] T_in_default hanya dipakai sebagai fallback kalau NI data tidak ada
    # Nilai aktual: T1Ka (CAM) = 24.8-42.8°C, T3Kd (C3) = 26.4-31.0°C
    T_in_default: float = 29.5 + 273.15  # K — rata-rata dari NI data (bukan 25°C!)

    h_in: float = 8.0            # W/m²K — koef. konveksi interior, standar
    dynamic_h_in: bool = False

@dataclass
class NumericalParameters:
    """Parameter numerik dari Paper Section 3.1 — tidak perlu diubah."""
    dt: float = 1.0              # s — time step (paper: dt=1s → stable)
    Nz_substrate: int = 107      # node substrat (paper: Nz=107 → mesh independent)
    Nz_slab: int = 67            # node slab (paper: Nz=67)
    convergence: float = 1e-5   # kriteria Gauss-Seidel


# ==============================================================================
# INISIALISASI PARAMETER — NILAI YANG SUDAH DIKETAHUI
# ==============================================================================

# --- Bromelia (CAM) ---
# [REAL] H_f, LAI, rho_f: dari pengukuran setahun lalu
# [C2][C3] tau_f=0.07, alpha_f dihitung dari rho_f
# [C8] r_stoma_min: PLACEHOLDER — perlu Licor CAM
bromelia = PlantParameters(
    name        = "Bromelia (CAM)",
    plant_type  = "CAM",
    H_f         = 0.098,    # [REAL] 9.8 cm dari data setahun lalu
    d_f         = 0.045,    # [ESTIMASI] dari foto — ukur lebih akurat dengan ImageJ
    LAI         = 1.95,     # [REAL] dari data setahun lalu — akan dikalibrasi
    rho_f       = 0.390,    # [REAL] dari pengukuran langsung
    tau_f       = 0.07,     # [C2] daun tebal+lilin → tau rendah (range 0.05-0.10)
    alpha_f     = 0.540,    # [C3] = 1 - 0.390 - 0.07
    epsilon_f   = 0.95,     # standar semua daun
    r_stoma_min = 400.0,    # [C8] PLACEHOLDER — upload Licor CAM untuk nilai nyata
    cover_fraction = 0.95   # projected canopy cover fraction, 0-1
)

# --- Wedelia (C3) ---
# [REAL] H_f, LAI, rho_f: dari pengukuran setahun lalu
# [C2] tau_f=0.20 (bukan 0.10!) — daun C3 tipis meneruskan lebih banyak cahaya
# [C3] alpha_f dihitung dari rho_f
# [C4] r_stoma_min = 167.2 s/m dari Licor NYATA
wedelia = PlantParameters(
    name        = "Wedelia (C3)",
    plant_type  = "C3",
    H_f         = 0.276,    # [REAL] 27.6 cm dari data setahun lalu
    d_f         = 0.060,    # [ESTIMASI] dari foto — ukur lebih akurat dengan ImageJ
    LAI         = 1.3,     # [REAL] dari data setahun lalu — akan dikalibrasi
    rho_f       = 0.438,    # [REAL] dari pengukuran langsung
    tau_f       = 0.20,     # [C2] daun C3 tipis → tau tinggi (range 0.15-0.30)
    alpha_f     = 0.462,    # [C3] = 1 - 0.438 - 0.10 (pakai tau=0.10 awal)
    epsilon_f   = 0.95,     # standar semua daun
    r_stoma_min = 167.2,    # [C4] REAL — percentile 5% dari Licor C3 April 2026
    cover_fraction = 0.5   # projected canopy cover fraction, 0-1
)
# NOTE: alpha_f wedelia = 1 - rho_f - tau_f = 1 - 0.438 - 0.20 = 0.362
#       dengan tau_f yang sudah diperbarui ke 0.20
#       perlu update: alpha_f = 1 - 0.438 - 0.20 = 0.362
wedelia.alpha_f = 1.0 - wedelia.rho_f - wedelia.tau_f  # = 0.362

# --- Substrat: tanah + sekam padi ---
substrat = SubstrateParameters(
    # Nilai sementara untuk running awal. Update kalau data final sudah ada.
    lambda_dry  = 0.2,    # [LIT/EST] W/mK — tanah+sekam padi ringan, dry
    lambda_sat  = None,    # [CALC] default dihitung = lambda_dry + theta_sat*lambda_water
    rho_g       = 400.0,   # [DATA] kg/m³ — bulk density substrat dari user
    cp_g        = 1300.0,  # [LIT] J/kgK — tanah+sekam padi (range 1200-1500)
    theta_sat   = 0.87,    # [EST] m³/m³ — scientific guess awal, update dari RIKA/literatur
    k_theta_sat = 5e-6,    # [EST/CAL] m/s — effective saturated hydraulic conductivity
    psi_sat     = 0.35,    # [LIT] m — suction magnitude untuk loam-like substrate
    b           = 5.5,     # [LIT] estimasi dari grain size (loam: 5-6)
)

# --- Slab beton ---
slab = SlabParameters(
    H_slab = 0.10,         # [EST] m — scientific guess awal, update dari ukuran boks
)

# --- Geometri ---
# [C1] T_in_default diset ke 29.5°C (rata-rata dari NI data), BUKAN 25°C
geom = GeometryParameters(
    H_g            = 0.06,             # [EST] m — tebal substrat efektif sementara
    H_slab         = 0.10,             # [EST] m — samakan dengan slab.H_slab
    T_in_default   = 29.5 + 273.15,   # [C1] dari rata-rata NI data T1Ka
)

# --- Numerical ---
num = NumericalParameters()

# Hitung lambda_sat default jika belum diisi: pendekatan sederhana kontribusi air pori
if substrat.lambda_sat is None:
    substrat.lambda_sat = substrat.lambda_dry + substrat.theta_sat * substrat.lambda_water


# ==============================================================================
# WINDOW VALIDASI SEMENTARA — DATA CAM DAN C3 TIDAK HARUS DI HARI YANG SAMA
# ==============================================================================

VALIDATION_WINDOWS: Dict[str, Tuple[pd.Timestamp, pd.Timestamp]] = {
    # CAM dipilih sebelum anomali RIKA 4 April (drop 0 lalu lompat 90-an).
    "CAM": (pd.Timestamp("2026-03-31 11:58:00"),
            pd.Timestamp("2026-04-02 21:42:00")),

    # C3 window utama: overlap bersih setelah gap besar RIKA.
    "C3":  (pd.Timestamp("2026-04-09 11:05:00"),
            pd.Timestamp("2026-04-10 14:08:00")),
}

# Alternatif C3 untuk dicek manual. Jangan dianggap satu simulasi kontinu dengan window utama.
VALIDATION_WINDOWS_ALT: Dict[str, Tuple[pd.Timestamp, pd.Timestamp]] = {
    "C3_early": (pd.Timestamp("2026-04-06 08:25:00"),
                 pd.Timestamp("2026-04-07 10:06:00")),
    "C3_early_after_anomaly": (pd.Timestamp("2026-04-06 17:38:00"),
                               pd.Timestamp("2026-04-07 10:06:00")),
}

VALIDATION_TARGETS = {
    "CAM": "T2A2",   # default awal
    "C3": "T3Ka",
}

def apply_scientific_guess_parameters(
    rho_g: float = 400.0,
    theta_sat: float = 0.90,
    k_theta_sat: float = 5e-6,
    H_g: float = 0.10,
    H_slab: float = 0.10,
    A_roof: Optional[float] = 1.0,
    lambda_dry: float = 0.12,
    lambda_sat: Optional[float] = None,
    h_in: float = 8.0,
    LAI_CAM: Optional[float] = None,
    LAI_C3: Optional[float] = None,
):
    """Update parameter sementara di satu tempat.

    Pakai ini untuk running awal dengan scientific guess. Kalau data final sudah ada,
    ubah argumennya di sini, bukan tersebar di banyak baris.
    """
    substrat.rho_g = float(rho_g)
    substrat.theta_sat = float(theta_sat)
    substrat.k_theta_sat = float(k_theta_sat)
    substrat.lambda_dry = float(lambda_dry)
    if lambda_sat is None:
        lambda_sat = substrat.lambda_dry + substrat.theta_sat * substrat.lambda_water
    substrat.lambda_sat = float(lambda_sat)

    geom.H_g = float(H_g)
    geom.h_in = float(h_in)
    geom.H_slab = float(H_slab)
    geom.A_roof = None if A_roof is None else float(A_roof)
    slab.H_slab = float(H_slab)

    # LAI tidak wajib dikalibrasi. Kalau belum ada hasil ImageJ, pakai nilai awal
    # sebagai scientific guess agar simulasi tetap running. Nanti update LAI_CAM/LAI_C3
    # dengan hasil ImageJ tanpa perlu mengubah fungsi lain.
    if LAI_CAM is not None:
        bromelia.LAI = float(LAI_CAM)
    elif bromelia.LAI is None:
        bromelia.LAI = 1.5   # fallback konservatif jika semua data LAI kosong

    if LAI_C3 is not None:
        wedelia.LAI = float(LAI_C3)
    elif wedelia.LAI is None:
        wedelia.LAI = 1.2    # fallback konservatif jika semua data LAI kosong
    # Projected canopy cover fraction.
    # Ini bukan LAI. Ini fraksi area horizontal yang tertutup kanopi.
    # CAM diasumsikan hampir full cover, C3 dari foto sekitar 0.40–0.50.
    bromelia.cover_fraction = 0.95
    wedelia.cover_fraction = 0.45

def set_LAI_from_imagej(LAI_CAM: Optional[float] = None,
                        LAI_C3: Optional[float] = None):
    """Update LAI dari hasil ImageJ tanpa menjalankan inverse calibration.

    Pakai fungsi ini setelah kamu punya LAI hasil analisis citra.
    Contoh:
        set_LAI_from_imagej(LAI_CAM=1.35, LAI_C3=1.10)
    """
    if LAI_CAM is not None:
        bromelia.LAI = float(LAI_CAM)
    if LAI_C3 is not None:
        wedelia.LAI = float(LAI_C3)
    print(f"LAI sekarang: CAM={bromelia.LAI:.3f}, C3={wedelia.LAI:.3f}")



# ==============================================================================
# SECTION 2: DATA LOADERS
# ==============================================================================


def _to_float(x):
    """Robust numeric conversion for logger cells."""
    if x is None:
        return np.nan
    if isinstance(x, str):
        x = x.strip()
        if x in {"", "---", "--"}:
            return np.nan
    return pd.to_numeric(x, errors="coerce")


def load_weather_data(filepath: str,
                      date_start: str = None,
                      date_end: str = None,
                      sheet_name: str = "3-24april") -> pd.DataFrame:
    """
    Load data cuaca Davis Vantage Pro 2.

    Sekarang bisa membaca:
    1) WeatherLink TXT/tab-separated lama, atau
    2) Excel baru: weatherfile mar-april.xlsx.

    Output kolom utama:
    - T_a      : suhu udara luar [°C]
    - RH       : relative humidity [%]
    - u        : wind speed dari logger
    - rain     : rain per interval [mm/menit jika data 1 menit]
    - G_sol    : solar radiation [W/m²]
    - rain_flux: rain sebagai flux air [kg/m²s], untuk Eq. (9)

    Catatan penting:
    Rain tidak diperlakukan sebagai state variable. Untuk simulasi 1 detik,
    rain_flux akan di-forward-fill, bukan diinterpolasi linear.
    """
    filepath = str(filepath)
    print(f"Loading weather data: {filepath}")

    start = pd.Timestamp(date_start) if date_start is not None else None
    end = pd.Timestamp(date_end) if date_end is not None else None

    if filepath.lower().endswith((".xlsx", ".xlsm", ".xls")):
        # Excel Davis export: kolom sama dengan TXT lama.
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name]
        records = []
        # File upload punya 3 baris header; data mulai baris 4.
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            ts = pd.to_datetime(f"{row[0]} {row[1]}", dayfirst=True, errors="coerce")
            if pd.isna(ts):
                continue
            if start is not None and ts < start:
                continue
            if end is not None and ts > end:
                continue
            records.append({
                'timestamp': ts,
                'T_a': _to_float(row[2]),
                'RH': _to_float(row[5]),
                'u': _to_float(row[7]),
                'rain': _to_float(row[17]),
                'G_sol': _to_float(row[19]),
            })
        if not records:
            raise ValueError("Tidak ada data weather pada periode yang diminta.")
        df = pd.DataFrame.from_records(records).set_index('timestamp').sort_index()
    else:
        # TXT WeatherLink lama.
        df_raw = pd.read_csv(
            filepath, sep='\t', skiprows=2, header=None,
            na_values=['---', '  ---', ' ---', '---  '], low_memory=False
        )
        df = pd.DataFrame({
            'date'  : df_raw.iloc[:, 0].astype(str).str.strip(),
            'time'  : df_raw.iloc[:, 1].astype(str).str.strip(),
            'T_a'   : pd.to_numeric(df_raw.iloc[:, 2],  errors='coerce'),
            'RH'    : pd.to_numeric(df_raw.iloc[:, 5],  errors='coerce'),
            'u'     : pd.to_numeric(df_raw.iloc[:, 7],  errors='coerce'),
            'rain'  : pd.to_numeric(df_raw.iloc[:, 17], errors='coerce'),
            'G_sol' : pd.to_numeric(df_raw.iloc[:, 19], errors='coerce'),
        })
        # Gunakan dayfirst=True agar aman untuk file Indonesia.
        df['timestamp'] = pd.to_datetime(df['date'] + ' ' + df['time'],
                                         dayfirst=True, errors='coerce')
        df = df.dropna(subset=['timestamp'])
        df = df.set_index('timestamp').drop(columns=['date', 'time']).sort_index()
        if start is not None:
            df = df[df.index >= start]
        if end is not None:
            df = df[df.index <= end]

    # File weather upload mengandung duplikasi timestamp. Ambil satu baris saja.
    before = len(df)
    df = df[~df.index.duplicated(keep='first')]
    n_dup = before - len(df)
    if n_dup > 0:
        print(f"  Duplicate timestamp dibuang: {n_dup}")

    # Koreksi fisik dasar.
    df['G_sol'] = pd.to_numeric(df['G_sol'], errors='coerce').fillna(0).clip(lower=0)
    df['u']     = pd.to_numeric(df['u'], errors='coerce').clip(lower=0.1)
    df['RH']    = pd.to_numeric(df['RH'], errors='coerce').clip(0, 100)
    df['T_a']   = pd.to_numeric(df['T_a'], errors='coerce')
    df['rain']  = pd.to_numeric(df['rain'], errors='coerce').fillna(0).clip(lower=0)

    # Interpolasi state variables saja.
    for col in ['T_a', 'G_sol', 'RH', 'u']:
        df[col] = df[col].interpolate(method='time', limit=30, limit_direction='both')

    # 1 mm rain = 1 kg/m². Jika data per menit, flux = mm/min / 60.
    df['rain_flux'] = df['rain'] / 60.0

    df = df.dropna(subset=['T_a', 'G_sol', 'RH', 'u'])

    print(f"  Periode : {df.index[0]} → {df.index[-1]}")
    print(f"  Records : {len(df)} ({len(df)/60:.1f} jam)")
    print(f"  T_a     : {df['T_a'].min():.1f}–{df['T_a'].max():.1f} °C")
    print(f"  G_sol   : {df['G_sol'].min():.0f}–{df['G_sol'].max():.0f} W/m²")
    print(f"  RH      : {df['RH'].min():.0f}–{df['RH'].max():.0f} %")
    print(f"  Rain    : {df['rain'].sum():.1f} mm (jumlah kolom rain)")
    return df


def create_weather_cache(filepath: str,
                         cache_path: str = "weather_clean_cache.csv",
                         sheet_name: str = "3-24april") -> pd.DataFrame:
    """Parse Excel weather sekali lalu simpan CSV agar run berikutnya cepat."""
    df = load_weather_data(filepath, sheet_name=sheet_name)
    df.to_csv(cache_path, index_label='timestamp')
    print(f"Weather cache saved: {cache_path}")
    return df


def load_weather_cache_or_excel(filepath: str,
                                cache_path: str = "weather_clean_cache.csv",
                                date_start: str = None,
                                date_end: str = None,
                                sheet_name: str = "3-24april") -> pd.DataFrame:
    """Load weather dari cache kalau ada; kalau belum ada, parse Excel/TXT."""
    if Path(cache_path).exists():
        df = pd.read_csv(cache_path, parse_dates=['timestamp']).set_index('timestamp').sort_index()
        if date_start:
            df = df[df.index >= pd.Timestamp(date_start)]
        if date_end:
            df = df[df.index <= pd.Timestamp(date_end)]
        return df
    return load_weather_data(filepath, date_start=date_start, date_end=date_end, sheet_name=sheet_name)


def summarize_weather_data(weather_df: pd.DataFrame,
                           label: str = "weather",
                           print_summary: bool = True) -> pd.DataFrame:
    """Ringkas data weather agar file panjang mudah dibaca.

    Input harus output dari load_weather_data()/load_weather_cache_or_excel().
    Hasil satu baris berisi statistik utama untuk window yang sedang dipakai.
    """
    if weather_df is None or len(weather_df) == 0:
        raise ValueError("weather_df kosong")

    df = weather_df.copy().sort_index()
    diffs = df.index.to_series().diff().dropna().dt.total_seconds()
    dt_med = float(diffs.median()) if len(diffs) else 60.0
    gap_threshold = max(2 * dt_med, 120.0)
    gaps = diffs[diffs > gap_threshold]

    duration_h = (df.index[-1] - df.index[0]).total_seconds() / 3600.0
    solar_kwh_m2 = float((df['G_sol'].fillna(0) * dt_med).sum() / 3.6e6)
    rain_total = float(df['rain'].fillna(0).sum()) if 'rain' in df else np.nan

    daylight = df[df['G_sol'] > 20]
    daytime_G_mean = float(daylight['G_sol'].mean()) if len(daylight) else 0.0
    daylight_hours = len(daylight) * dt_med / 3600.0

    # Klasifikasi sederhana untuk narasi hasil.
    if rain_total >= 10 and df['RH'].mean() >= 80:
        weather_type = "humid-rainy tropical"
    elif solar_kwh_m2 >= 4 and rain_total < 2:
        weather_type = "sunny/drier short-period"
    elif df['RH'].mean() >= 80:
        weather_type = "humid tropical"
    else:
        weather_type = "mixed short-period"

    out = pd.DataFrame([{
        'label': label,
        'start': df.index[0],
        'end': df.index[-1],
        'duration_h': duration_h,
        'records': len(df),
        'median_dt_s': dt_med,
        'n_gaps': int(len(gaps)),
        'max_gap_min': float(gaps.max()/60.0) if len(gaps) else 0.0,
        'T_a_min_C': float(df['T_a'].min()),
        'T_a_mean_C': float(df['T_a'].mean()),
        'T_a_max_C': float(df['T_a'].max()),
        'RH_min_pct': float(df['RH'].min()),
        'RH_mean_pct': float(df['RH'].mean()),
        'RH_max_pct': float(df['RH'].max()),
        'G_sol_max_Wm2': float(df['G_sol'].max()),
        'G_sol_mean_Wm2': float(df['G_sol'].mean()),
        'G_sol_daytime_mean_Wm2': daytime_G_mean,
        'solar_energy_kWh_m2': solar_kwh_m2,
        'daylight_hours': daylight_hours,
        'rain_total_mm': rain_total,
        'wind_mean': float(df['u'].mean()),
        'wind_max': float(df['u'].max()),
        'weather_type': weather_type,
    }])

    if print_summary:
        r = out.iloc[0]
        print("\n" + "="*70)
        print(f"WEATHER SUMMARY — {label}")
        print("="*70)
        print(f"Periode   : {r['start']} → {r['end']} ({r['duration_h']:.1f} jam)")
        print(f"Records   : {int(r['records'])}, median dt = {r['median_dt_s']:.0f} s")
        print(f"Gaps      : {int(r['n_gaps'])} gap, max = {r['max_gap_min']:.1f} menit")
        print(f"T_a       : {r['T_a_min_C']:.1f}–{r['T_a_max_C']:.1f} °C, mean {r['T_a_mean_C']:.1f} °C")
        print(f"RH        : {r['RH_min_pct']:.0f}–{r['RH_max_pct']:.0f} %, mean {r['RH_mean_pct']:.1f} %")
        print(f"G_sol     : max {r['G_sol_max_Wm2']:.0f} W/m², energy {r['solar_energy_kWh_m2']:.2f} kWh/m²")
        print(f"Daylight  : {r['daylight_hours']:.1f} jam, daytime mean G = {r['G_sol_daytime_mean_Wm2']:.0f} W/m²")
        print(f"Rain      : {r['rain_total_mm']:.1f} mm")
        print(f"Wind      : mean {r['wind_mean']:.2f}, max {r['wind_max']:.2f}")
        print(f"Represent : {r['weather_type']}")
    return out


def summarize_weather_windows(base_dir: str = ".",
                              weather_file: str = "weatherfile mar-april.xlsx",
                              cache_path: str = "weather_clean_cache.csv") -> pd.DataFrame:
    """Ringkas weather untuk semua window validasi CAM/C3.

    Contoh:
        summary = summarize_weather_windows()
        summary.to_excel('weather_summary.xlsx', index=False)
    """
    base = Path(base_dir)
    rows = []
    for name, (start, end) in VALIDATION_WINDOWS.items():
        w = load_weather_cache_or_excel(str(base / weather_file),
                                        str(base / cache_path),
                                        date_start=str(start),
                                        date_end=str(end))
        rows.append(summarize_weather_data(w, label=name, print_summary=True))

    # Tambahkan window alternatif jika ada.
    for name, (start, end) in VALIDATION_WINDOWS_ALT.items():
        w = load_weather_cache_or_excel(str(base / weather_file),
                                        str(base / cache_path),
                                        date_start=str(start),
                                        date_end=str(end))
        rows.append(summarize_weather_data(w, label=name, print_summary=True))

    return pd.concat(rows, ignore_index=True)

def load_NI_sensor_data(filepath: str) -> pd.DataFrame:
    """
    Load data suhu dari NI DAQ (LabVIEW → Excel).

    [C5] Menggunakan XML parser langsung — bukan extract-text yang truncate di 1000 baris.
    [C6] T2A (Tanah bawah CAM) difilter dari 66 anomali ekstrem (< -10°C atau > 80°C).
    [C7] Timestamp dikonversi dari LabVIEW serial (hari sejak 1 Jan 1904).

    Channel mapping (dari nomenclature image yang sudah dikonfirmasi):
    Col  1 → T1Kd  : Lantai CAM
    Col  2 → T1Kb  : Dinding Timur CAM
    Col  3 → T1Ke  : Tanah Atas CAM      → T_g_top (kalibrasi bromelia)
    Col  4 → T1Ka  : Ruangan CAM         → T_in dinamis untuk CAM [C1]
    Col  5 → T1Kc  : Dinding Barat CAM
    Col  6 → T3Kb  : Dinding Barat C3
    Col  7 → T3Kd  : Ruangan C3          → T_in dinamis untuk C3 [C1]
    Col  8 → T3Ke  : Lantai C3
    Col  9 → T3Kc  : Dinding Timur C3
    Col 10 → T3Ka  : Atap Indoor C3      → T_s_in (kalibrasi wedelia)
    Col 11 → T2Ka  : Atap Indoor RR      → referensi conventional roof
    Col 12 → T2Kd  : Dinding Timur RR
    Col 13 → T2Kc  : Dinding Barat RR
    Col 14 → 2Ke   : TBD
    Col 15 → T1A   : TBD
    Col 16 → T2A   : Tanah Bawah CAM     → T_g_bot [C6: ada 66 anomali!]
    Col 17 → T2A2  : TBD
    Col 18 → T1Tb  : Atap Indoor CAM     → T_s_in (kalibrasi bromelia)
    Col 19 → T1Ta  : Atap Outdoor CAM
    Col 20 → T1Ta2 : TBD
    """
    print(f"Loading NI sensor data: {filepath}")

    # [C5] Parse XML langsung — tidak pakai extract-text
    import zipfile
    import io

    # Excel (.xlsx) sebenarnya adalah ZIP berisi XML
    with zipfile.ZipFile(filepath, 'r') as z:
        with z.open('xl/worksheets/sheet1.xml') as f:
            xml_content = f.read()

    tree = ET.fromstring(xml_content)
    ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'

    rows_xml = tree.findall(f'.//{ns}row')
    print(f"  Total rows di XML: {len(rows_xml)} (termasuk 1 header)")

    ni_channel_names = [
        'timestamp_serial',
        'T1Kd', 'T1Kb', 'T1Ke', 'T1Ka', 'T1Kc',
        'T3Kb', 'T3Kd', 'T3Ke', 'T3Kc', 'T3Ka',
        'T2Ka', 'T2Kd', 'T2Kc', '2Ke',  'T1A',
        'T2A',  'T2A2', 'T1Tb', 'T1Ta', 'T1Ta2'
    ]

    data = []
    for row in rows_xml[1:]:   # skip header row
        vals = []
        for cell in row.findall(f'{ns}c'):
            v = cell.find(f'{ns}v')
            vals.append(float(v.text) if v is not None else np.nan)
        if len(vals) == 21:
            data.append(vals)

    df = pd.DataFrame(data, columns=ni_channel_names)

    # [C7] Konversi LabVIEW timestamp (hari sejak 1 Jan 1904) ke datetime
    labview_epoch = pd.Timestamp('1904-01-01')
    df['timestamp'] = labview_epoch + pd.to_timedelta(
        df['timestamp_serial'], unit='D'
    )
    df = df.set_index('timestamp').drop(columns=['timestamp_serial'])

    # [C6] Filter anomali T2A — 66 data dengan nilai ekstrem
    # Nilai < -10°C atau > 80°C adalah tidak fisik → sensor disconnect
    print(f"\n  [C6] Filtering anomali T2A...")
    mask_anomaly = (df['T2A'] < -10) | (df['T2A'] > 80)
    n_anomaly = mask_anomaly.sum()
    print(f"  Anomali T2A: {n_anomaly} data")

    if n_anomaly > 0:
        # Cek max cluster
        clusters = []
        count = 0
        for val in mask_anomaly:
            if val:
                count += 1
            else:
                if count > 0:
                    clusters.append(count)
                    count = 0
        if count > 0:
            clusters.append(count)
        max_cluster = max(clusters) if clusters else 0
        print(f"  Max berturutan: {max_cluster} data")

        # Replace dengan NaN lalu interpolasi
        df.loc[mask_anomaly, 'T2A'] = np.nan
        df['T2A'] = df['T2A'].interpolate(
            method='linear',
            limit=30,           # max 30 menit gap
            limit_direction='both'
        )
        print(f"  Status: replaced dengan NaN → interpolasi linear ✓")

    # ============================================================
    # General NI temperature anomaly cleaning
    # Applies to all raw temperature channels
    # ============================================================

    raw_temp_cols = [
        c for c in df.columns
        if c.startswith("T") and pd.api.types.is_numeric_dtype(df[c])
    ]

    for col in raw_temp_cols:
        # batas fisik suhu eksperimen outdoor/indoor
        mask_bad = (df[col] < -10) | (df[col] > 80)

        n_bad = int(mask_bad.sum())
        if n_bad > 0:
            print(f"  Cleaning anomaly {col}: {n_bad} points")
            df.loc[mask_bad, col] = np.nan

            # Interpolasi hanya gap pendek
            df[col] = df[col].interpolate(
                method="time",
                limit=30,
                limit_direction="both"
            )
    # Buat kolom deskriptif untuk simulasi
    df['T_g_top_CAM'] = df['T1Ke']   # Tanah atas CAM → kalibrasi bromelia (dual)
    df['T_g_bot_CAM'] = df['T2A']    # Tanah bawah CAM → validasi profil
    df['T_s_in_CAM']  = df['T1Tb']   # Atap indoor CAM → kalibrasi bromelia (dual)
    df['T_s_ext_CAM'] = df['T1Ta']   # Atap outdoor CAM
    df['T_s_in_C3']   = df['T3Ka']   # Atap indoor C3 → kalibrasi wedelia
    # [C1] T_in dinamis — BUKAN 25°C konstan
    df['T_in_CAM']    = df['T1Ka']   # Ruangan CAM → dipakai sebagai T_in dinamis
    df['T_in_C3']     = df['T3Kd']   # Ruangan C3 → dipakai sebagai T_in dinamis
    df['T_r_in_RR']   = df['T2Ka']   # Referensi conventional roof

    interval = (df.index[1] - df.index[0]).total_seconds()

    print(f"\n  Periode : {df.index[0]} → {df.index[-1]}")
    print(f"  Records : {len(df)} ({len(df)/60:.1f} jam)")
    print(f"  Interval: {interval:.0f} detik")
    print(f"\n  Key variables (setelah cleaning):")
    for col, label in [
        ('T_g_top_CAM', 'T_g_top CAM (T1Ke)'),
        ('T_g_bot_CAM', 'T_g_bot CAM (T2A) '),
        ('T_s_in_CAM',  'T_s_in CAM (T1Tb) '),
        ('T_s_in_C3',   'T_s_in C3  (T3Ka) '),
        ('T_in_CAM',    '[C1] T_in CAM (T1Ka)'),
        ('T_in_C3',     '[C1] T_in C3  (T3Kd)'),
    ]:
        s = df[col].dropna()
        print(f"    {label}: {s.min():.1f}–{s.max():.1f}°C, mean={s.mean():.1f}")

    return df



def load_NI_sensor_data_auto(filepath: str) -> pd.DataFrame:
    """Wrapper agar nama function jelas untuk file NI raw/clean.

    Untuk saat ini langsung memanggil load_NI_sensor_data() karena file NI yang kamu
    upload masih kompatibel dengan XML parser di atas.
    """
    return load_NI_sensor_data(filepath)


def load_multiple_NI_sensor_data(filepaths: Sequence[str]) -> pd.DataFrame:
    """Gabungkan beberapa file NI akibat logger/laptop mati.

    File yang bisa digabung misalnya:
    - Pengukuran 30_1 Maret 2026.xlsx
    - Pengukuran 30_2 Maret 2026.xlsx
    - NI_sensor_data_clean.xlsx

    Gap besar tidak diinterpolasi di sini; data hanya digabung, sort, dan duplicate
    timestamp dibuang. Interpolasi kecil tetap dilakukan per-file di loader.
    """
    dfs = []
    for fp in filepaths:
        if Path(fp).exists():
            dfs.append(load_NI_sensor_data(fp))
        else:
            print(f"WARNING: file NI tidak ditemukan: {fp}")
    if not dfs:
        raise ValueError("Tidak ada file NI yang berhasil dibaca.")
    df = pd.concat(dfs).sort_index()
    df = df[~df.index.duplicated(keep='last')]
    print(f"\nNI gabungan: {df.index[0]} → {df.index[-1]} | {len(df)} records")
    return df


def load_licor_data(filepath: str) -> pd.DataFrame:
    """
    Load data gas exchange dari Licor 6800.
    Header kompleks: 14 baris sebelum data, baris 15 adalah satuan (skip).
    """
    print(f"Loading Licor data: {filepath}")

    df = pd.read_excel(filepath, header=14, skiprows=[15])

    for col in ['A', 'E', 'gsw', 'TleafEB', 'VPDleaf', 'Qin']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    print(f"  Records: {len(df)}")
    if 'gsw' in df.columns:
        gsw_valid = df[(df['gsw'] > 0.001) & (df.get('A', pd.Series([1]*len(df))) > 0)]['gsw']
        print(f"  Valid gsw (>0.001, A>0): {len(gsw_valid)}")
        if len(gsw_valid) > 0:
            r = 1 / (gsw_valid * 0.0224)
            print(f"  r_stoma: {r.min():.1f}–{r.max():.1f} s/m")

    return df


def extract_r_stoma_min(licor_df: pd.DataFrame,
                         plant_type: str = "C3",
                         percentile: float = 5.0) -> float:
    """
    Ekstrak r_stoma_min dari data Licor.
    r_stoma (s/m) = 1 / (gsw_mol × 0.0224)
    Pakai percentile ke-5 (bukan minimum) untuk stabilitas.
    [C4] Untuk C3: sudah dikonfirmasi = 167.2 s/m dari data April 2026
    """
    mask = (licor_df['gsw'] > 0.001)
    if 'A' in licor_df.columns:
        mask = mask & (licor_df['A'] > 0)

    gsw_valid = licor_df[mask]['gsw']

    if len(gsw_valid) == 0:
        print(f"WARNING: Tidak ada gsw valid! Pakai nilai default.")
        return 167.2 if plant_type == "C3" else 400.0

    r_stoma = 1 / (gsw_valid * 0.0224)
    r_stoma_min = float(np.percentile(r_stoma, percentile))

    print(f"\nr_stoma_min {plant_type}:")
    print(f"  Min absolut     : {r_stoma.min():.1f} s/m")
    print(f"  Percentile {percentile:.0f}%  : {r_stoma_min:.1f} s/m  ← dipakai")
    print(f"  Mean            : {r_stoma.mean():.1f} s/m")

    return r_stoma_min



def _read_zip_text(z: zipfile.ZipFile, name: str) -> str:
    """Read text file inside ZIP with fallback encoding."""
    data = z.read(name)
    # Beberapa export logger punya null bytes/UTF-16-like.
    if data[:200].count(b"\x00") > 40:
        return data.decode('utf-16le', errors='ignore').replace('\x00', '')
    try:
        return data.decode('utf-8')
    except UnicodeDecodeError:
        return data.decode('latin1', errors='ignore')


def load_rika_soil_file_from_zip(zip_path: str,
                                 member_name: str,
                                 theta_min_valid: float = 0.03,
                                 theta_max_valid: float = 0.88) -> pd.DataFrame:
    """
    Load satu file CSV RIKA dari ZIP.

    Timestamp di logger terbaca sebagai GMT/UTC, lalu dikonversi ke WIB
    dan dibuat timezone-naive supaya match dengan weather/NI.
    """
    with zipfile.ZipFile(zip_path, 'r') as z:
        txt = _read_zip_text(z, member_name)

    df = pd.read_csv(io.StringIO(txt), sep=';')
    df['timestamp_utc'] = pd.to_datetime(df['Timestamp'], utc=True, errors='coerce')
    df['timestamp'] = df['timestamp_utc'].dt.tz_convert('Asia/Jakarta').dt.tz_localize(None)
    df['soil_temp'] = pd.to_numeric(df.get('Temperature'), errors='coerce')
    df['moisture_pct'] = pd.to_numeric(df.get('Moisture (%)'), errors='coerce')
    df['theta'] = df['moisture_pct'] / 100.0
    df = df.dropna(subset=['timestamp']).set_index('timestamp').sort_index()

    # Filter glitch yang jelas tidak fisik untuk VWC green roof.
    bad = (df['theta'] < theta_min_valid) | (df['theta'] > theta_max_valid)
    if bad.sum() > 0:
        print(f"  {member_name}: {bad.sum()} data moisture di-flag NaN")
        df.loc[bad, ['moisture_pct', 'theta']] = np.nan

    return df[['soil_temp', 'moisture_pct', 'theta']]


def load_soil_moisture_data(filepath: str,
                            plant_type: str = "CAM") -> pd.DataFrame:
    """
    Load data soil moisture RIKA dari datasoilmoisture.zip.

    Output kolom:
    - theta_2cm, theta_7cm
    - moisture_2cm_pct, moisture_7cm_pct
    - soil_temp_2cm, soil_temp_7cm

    Asumsi sementara:
    sensor 1 = kedalaman dangkal ±2 cm
    sensor 2 = kedalaman dalam ±7 cm
    """
    plant_type = plant_type.upper()
    if plant_type == "CAM":
        members = ("sensor 1 COM5_CAM.csv", "sensor 2 COM6_CAM.csv")
    elif plant_type == "C3":
        members = ("sensor 1 COM5_C3.csv", "sensor 2 COM6_C3.csv")
    else:
        raise ValueError("plant_type harus 'CAM' atau 'C3'")

    s1 = load_rika_soil_file_from_zip(filepath, members[0]).rename(columns={
        'soil_temp': 'soil_temp_2cm',
        'moisture_pct': 'moisture_2cm_pct',
        'theta': 'theta_2cm',
    })
    s2 = load_rika_soil_file_from_zip(filepath, members[1]).rename(columns={
        'soil_temp': 'soil_temp_7cm',
        'moisture_pct': 'moisture_7cm_pct',
        'theta': 'theta_7cm',
    })
    df = pd.concat([s1, s2], axis=1).sort_index()

    # ============================================================
    # MANUAL CLEANING FOR KNOWN RIKA GLITCHES
    # Timestamp RIKA sudah dikonversi dari GMT/UTC ke WIB di loader sebelumnya.
    # Jadi waktu di bawah ini adalah WIB.
    # ============================================================

    if plant_type == "CAM":
        # CAM sensor 1 glitch:
        # sekitar 1 Apr 2026 16:40 WIB, data drop/lompat tidak fisik.
        glitch_start = pd.Timestamp("2026-04-01 16:39:00")
        glitch_end   = pd.Timestamp("2026-04-01 16:43:00")

        cols = ["theta_2cm", "moisture_2cm_pct"]
        existing_cols = [c for c in cols if c in df.columns]

        if existing_cols:
            mask = (df.index >= glitch_start) & (df.index <= glitch_end)
            n_bad = int(mask.sum())
            if n_bad > 0:
                print(f"  RIKA CAM manual glitch mask: {n_bad} rows at {glitch_start} → {glitch_end}")
                df.loc[mask, existing_cols] = np.nan

    elif plant_type == "C3":
        # C3 early-window sensor 2 glitch:
        # ini hanya penting kalau kamu pakai window C3 awal 6 Apr.
        glitch_start = pd.Timestamp("2026-04-06 17:33:00")
        glitch_end   = pd.Timestamp("2026-04-06 17:37:00")

        cols = ["theta_7cm", "moisture_7cm_pct"]
        existing_cols = [c for c in cols if c in df.columns]

        if existing_cols:
            mask = (df.index >= glitch_start) & (df.index <= glitch_end)
            n_bad = int(mask.sum())
            if n_bad > 0:
                print(f"  RIKA C3 manual glitch mask: {n_bad} rows at {glitch_start} → {glitch_end}")
                df.loc[mask, existing_cols] = np.nan

    # Interpolasi hanya gap pendek setelah masking.
    # limit=15 artinya maksimal sekitar 15 record; aman untuk glitch beberapa menit.
    for col in ["theta_2cm", "theta_7cm", "moisture_2cm_pct", "moisture_7cm_pct"]:
        if col in df.columns:
            df[col] = df[col].interpolate(
                method="time",
                limit=15,
                limit_direction="both"
            )

    print(f"RIKA {plant_type}: {df.index[0]} → {df.index[-1]} | {len(df)} records")
    return df

def synchronize_datasets(weather_df: pd.DataFrame,
                          ni_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    [C7] Sinkronisasi timestamp Davis dan NI sensor.

    Masalah: Davis interval 1 menit tepat (xx:00),
    NI dari LabVIEW mungkin ada detik tidak tepat (xx:58:16).
    Solusi: resample keduanya ke grid 1 menit yang sama.
    """
    print("\nSinkronisasi dataset...")

    # Temukan periode overlap
    start = max(weather_df.index[0], ni_df.index[0])
    end   = min(weather_df.index[-1], ni_df.index[-1])

    print(f"  Overlap: {start} → {end}")
    duration_h = (end - start).total_seconds() / 3600
    print(f"  Durasi : {duration_h:.1f} jam = {duration_h/24:.1f} hari")

    # Filter ke periode overlap
    w = weather_df[(weather_df.index >= start) & (weather_df.index <= end)].copy()
    n = ni_df[(ni_df.index >= start) & (ni_df.index <= end)].copy()

    # Resample NI ke grid 1 menit aligned dengan Davis
    # (menangani slight timestamp drift dari LabVIEW)
    n_resampled = n.resample('1min').mean()

    print(f"  Weather records : {len(w)}")
    print(f"  NI records      : {len(n)} → {len(n_resampled)} (setelah resample)")

    return w, n_resampled


def resample_for_simulation(df: pd.DataFrame,
                             dt: float = 1.0) -> pd.DataFrame:
    """Resample data ke interval dt detik untuk simulasi."""
    rule = f'{int(dt)}s'
    df_resampled = df.resample(rule).interpolate(method='time')
    print(f"  Resample: {len(df)} → {len(df_resampled)} records ({dt:.0f}s interval)")
    return df_resampled




def prepare_weather_1s(weather_df: pd.DataFrame, dt: float = 1.0) -> pd.DataFrame:
    """
    Resample weather ke interval dt detik dengan perlakuan berbeda:
    - T_a, RH, u, G_sol: interpolasi time
    - rain_flux: forward-fill supaya water balance tidak rusak
    """
    rule = f'{int(dt)}s'
    states = weather_df[['T_a', 'RH', 'u', 'G_sol']].resample(rule).interpolate('time')

    if 'rain_flux' in weather_df.columns:
        rain_flux = weather_df['rain_flux'].fillna(0).resample(rule).ffill().fillna(0)
    else:
        rain_flux = (weather_df['rain'].fillna(0)/60.0).resample(rule).ffill().fillna(0)
    states['rain_flux'] = rain_flux
    return states


def make_theta_initial_profile(theta_initial,
                               H_g: float,
                               Nz: int,
                               substrate: SubstrateParameters) -> np.ndarray:
    """
    Buat profil awal theta.

    Bisa menerima:
    - None              → 0.8*theta_sat seragam
    - float             → seragam
    - [theta_2cm, theta_7cm]
    - dict {'depths':[...], 'values':[...]}.
    """
    z = np.linspace(0, H_g, Nz)
    if theta_initial is None:
        return np.full(Nz, substrate.theta_sat * 0.80)
    if isinstance(theta_initial, dict):
        depths = np.asarray(theta_initial.get('depths', [0.02, 0.07]), dtype=float)
        values = np.asarray(theta_initial.get('values'), dtype=float)
        return np.clip(np.interp(z, depths, values, left=values[0], right=values[-1]),
                       substrate.theta_min, substrate.theta_sat)
    if isinstance(theta_initial, (list, tuple, np.ndarray)) and len(theta_initial) == 2:
        return np.clip(np.interp(z, [0.02, 0.07], theta_initial,
                                 left=theta_initial[0], right=theta_initial[-1]),
                       substrate.theta_min, substrate.theta_sat)
    return np.full(Nz, float(theta_initial))


def get_theta_initial_from_rika(rika_df: pd.DataFrame,
                                start_time: pd.Timestamp) -> dict:
    """Ambil theta awal dari RIKA dekat start_time."""
    start_time = pd.Timestamp(start_time)
    r = rika_df.resample('1min').mean().interpolate('time', limit=10, limit_direction='both')
    local = r[(r.index >= start_time - pd.Timedelta('10min')) &
              (r.index <= start_time + pd.Timedelta('10min'))]
    if local[['theta_2cm', 'theta_7cm']].dropna().empty:
        local = r[r.index >= start_time].head(30)
    vals = local[['theta_2cm', 'theta_7cm']].dropna().iloc[0]
    return {'depths': [0.02, 0.07],
            'values': [float(vals['theta_2cm']), float(vals['theta_7cm'])]}


# ==============================================================================
# SECTION 3: HELPER TERMODINAMIKA
# ==============================================================================

def saturation_pressure(T_K: float) -> float:
    """Tekanan uap saturasi [Pa] via persamaan Magnus."""
    T_C = T_K - 273.15
    return 610.78 * np.exp(17.269 * T_C / (T_C + 237.29))



def dew_point_C(T_C: float, RH: float) -> float:
    """Dew point [°C] via Magnus approximation."""
    RH = float(np.clip(RH, 1.0, 100.0))
    a, b = 17.27, 237.7
    gamma = (a*T_C)/(b+T_C) + np.log(RH/100.0)
    return (b*gamma)/(a-gamma)


def sky_temperature(T_a_K: float, RH: float) -> float:
    """
    Suhu langit efektif [K].

    Koreksi: korelasi emissivity langit menggunakan dew point, bukan RH langsung.
    Jika RH langsung dimasukkan, epsilon_sky terlalu sering ter-clip ke 1.
    """
    T_C = T_a_K - 273.15
    Tdp = dew_point_C(T_C, RH)
    epsilon_sky = 0.711 + 0.56*(Tdp/100.0) + 0.73*(Tdp/100.0)**2
    epsilon_sky = float(np.clip(epsilon_sky, 0.60, 1.00))
    return epsilon_sky**0.25 * T_a_K

def ambient_vapor_pressure(T_a_K: float, RH: float) -> float:
    """Tekanan uap aktual udara [Pa]. VPD = P_sat - P_a."""
    return (RH/100) * saturation_pressure(T_a_K)


def psychrometric_constant() -> float:
    """Konstanta psikrometrik γ ≈ 66.8 Pa/K."""
    return 1005 * 101325 / (0.622 * 2.45e6)

def interior_ceiling_convective_h(T_surface_K: float, T_air_K: float) -> float:
    """
    Natural convection coefficient for interior ceiling/underside roof surface.

    For a hot downward-facing ceiling surface, buoyancy is stable,
    so convection is weak. Walton/ASHRAE-style correlation:
        h = a * |ΔT|^(1/3)

    a = 0.76 for stable horizontal surface
    a = 1.52 for unstable horizontal surface

    Unit: W/m2K
    """
    dT = float(T_surface_K - T_air_K)
    dT_abs = max(abs(dT), 0.1)

    # Underside ceiling:
    # hot surface above cooler air => stable horizontal convection
    if dT > 0:
        a = 0.76
    else:
        # cooler ceiling with warmer air below => unstable
        a = 1.52

    h = a * (dT_abs ** (1.0 / 3.0))

    # keep within physically reasonable natural convection range
    return float(np.clip(h, 0.5, 4.0))
# ==============================================================================
# SECTION 4: BLOK 1 — FOLIAGE MODEL
# ==============================================================================

def compute_aerodynamic_resistance(u: float,
                                   plant: PlantParameters) -> float:
    """
    Resistansi aerodinamik r_a [s/m]. Persamaan A.2, A.3, A.4.
    Berbeda antara CAM dan C3 karena H_f berbeda.
    """
    k = 0.41
    d0 = 0.701 * plant.H_f**0.975    # A.3
    Z0 = 0.131 * plant.H_f**0.997    # A.4
    z_ref = plant.H_f + 2.0
    ratio = max((z_ref - d0) / Z0, 1.01)
    r_a = 1 / (k**2 * max(u, 0.1)) * (np.log(ratio))**2    # A.2
    return max(r_a, 5.0)


def compute_stomatal_resistance(G_sol: float,
                                T_f_K: float,
                                P_f_sat: float,
                                P_a: float,
                                theta: float,
                                plant: PlantParameters,
                                substrate: SubstrateParameters) -> float:
    """
    Stomatal resistance aktual r_stoma [s/m]. Persamaan A.6.
    Perbedaan CAM vs C3: modifier pada G_sol > 50 W/m².
    [C8] CAM factor = 12.0 masih placeholder → dari Licor CAM nanti.
    """
    # f1–f4: fungsi modifikasi A.7–A.10
    f1 = 1 + np.exp(-0.034 * (G_sol - 3.5))
    T_C = T_f_K - 273.15
    f2 = (np.exp(0.3*T_C) + 258) / (np.exp(0.3*T_C) + 27)
    VPD = max(P_f_sat - P_a, 0)
    f3 = 4e-3 + np.exp(-0.73 * 0.622e3/101325 * VPD)
    theta_safe = max(theta, substrate.theta_min + 0.001)
    f4 = substrate.theta_sat / theta_safe

    r_stoma = (plant.r_stoma_min / plant.LAI) * f1 * f2 * f3 * f4

    # Modifikasi CAM vs C3
    if plant.plant_type == "CAM":
        if G_sol > 50:    # siang hari
            # [C8] 12.0 = PLACEHOLDER — akan diupdate dari Licor CAM
            r_stoma *= 12.0
        else:             # malam hari
            r_stoma *= 0.5

    return max(r_stoma, 5.0)


def solve_foliage_temperature(T_f_prev: float,
                              T_a_K: float,
                              T_g_surface: float,
                              G_sol: float,
                              RH: float,
                              u: float,
                              theta_avg: float,
                              plant: PlantParameters,
                              substrate: SubstrateParameters,
                              dt: float) -> float:
    """
    Solve T_f baru dengan implicit scheme. Persamaan (1)(2)(3).
    """
    sigma   = 5.67e-8
    rho_air = 1.2
    cp_air  = 1005
    gamma   = psychrometric_constant()

    T_sky   = sky_temperature(T_a_K, RH)
    P_a     = ambient_vapor_pressure(T_a_K, RH)
    P_f_sat = saturation_pressure(T_f_prev)

    epsilon_fg = 1 / (1/plant.epsilon_f + 1/substrate.epsilon_g - 1)    # Eq.3

    R_f_net = (plant.alpha_f * G_sol
               - plant.epsilon_f * sigma * (T_f_prev**4 - T_sky**4)
               - epsilon_fg * sigma * (T_f_prev**4 - T_g_surface**4))   # Eq.2

    r_a      = compute_aerodynamic_resistance(u, plant)
    r_stoma  = compute_stomatal_resistance(G_sol, T_f_prev, P_f_sat, P_a,
                                           theta_avg, plant, substrate)
    h_conv_f = plant.LAI * (rho_air * cp_air) / r_a                     # A.1
    h_eva_f  = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_stoma) # A.5

    rho_cp_f = rho_air * cp_air
    d_eff    = 0.001

    dPdT = saturation_pressure(T_f_prev+0.5) - saturation_pressure(T_f_prev-0.5)

    A = rho_cp_f*d_eff/dt + h_conv_f + h_eva_f*dPdT

    # Derivation of B from Eq.1 (implicit):
    # (rho_cp*d/dt + h_conv + h_eva*dPdT)*T_f_new =
    #   (rho_cp*d/dt)*T_f_old + R_net + h_conv*T_a
    #   + h_eva*(P_a - P_f_sat + dPdT*T_f_old)
    # Note: sign of h_eva term is NEGATIVE of (P_f_sat - P_a - dPdT*T_f_old)
    B = (rho_cp_f*d_eff/dt * T_f_prev + R_f_net
         + h_conv_f*T_a_K + h_eva_f*(P_a - P_f_sat + dPdT*T_f_prev))

    return B / A


# ==============================================================================
# SECTION 5: BLOK 2 — SUBSTRATE MODEL
# ==============================================================================


def compute_substrate_properties(theta: np.ndarray,
                                  substrate: SubstrateParameters
                                  ) -> Tuple[np.ndarray, np.ndarray,
                                             np.ndarray, np.ndarray]:
    """
    Properti substrat sebagai fungsi VWC.

    Koreksi penting:
    - k_theta tetap hydraulic conductivity [m/s] untuk moisture transport.
    - lambda_g adalah thermal conductivity [W/mK] dan dihitung dari derajat
      kejenuhan air, bukan dari theta*k_theta.
    """
    theta = np.asarray(theta, dtype=float)
    b = substrate.b

    if substrate.theta_sat is None or substrate.k_theta_sat is None or substrate.rho_g is None:
        raise ValueError("theta_sat, k_theta_sat, dan rho_g harus diisi sebelum simulasi")

    theta_safe = np.maximum(theta, substrate.theta_min + 1e-6)
    S = np.clip(theta_safe/substrate.theta_sat, 0.0, 1.0)

    # B.8 style dari kode awal: kapasitas panas volumetrik naik dengan VWC.
    rho_cp_g = substrate.cp_g * (0.2 + theta_safe) * substrate.rho_g

    # B.10: hydraulic conductivity untuk Richards/moisture transport.
    k_theta = substrate.k_theta_sat * S**(2*b + 3)

    # Thermal conductivity: bounded oleh dry dan saturated condition.
    lambda_sat = substrate.lambda_sat
    if lambda_sat is None:
        lambda_sat = substrate.lambda_dry + substrate.theta_sat * substrate.lambda_water
    lambda_g = substrate.lambda_dry + S * (lambda_sat - substrate.lambda_dry)

    # B.11: psi_sat di sini disimpan sebagai magnitude positif, jadi D_theta dibuat positif.
    psi_abs = abs(substrate.psi_sat)
    D_theta = b * substrate.k_theta_sat * psi_abs / theta_safe * S**(b + 3)

    return rho_cp_g, lambda_g, k_theta, D_theta


def substrate_vapor_resistance(theta_surface: float,
                               substrate: SubstrateParameters) -> float:
    """
    Resistansi uap permukaan substrat [s/m].

    Temporary scientific guess: saat substrat makin kering, resistansi uap naik.
    Ini menggantikan placeholder lama r_vap = lambda_dry*50 yang mencampur
    properti termal dengan resistansi massa.
    """
    S = float(np.clip(theta_surface/substrate.theta_sat, 0.05, 1.0))
    return 100.0 * S**(-2.0)

def tdma_solver(a: np.ndarray, b: np.ndarray,
                c: np.ndarray, d: np.ndarray) -> np.ndarray:
    """
    Tridiagonal Matrix Algorithm (Thomas Algorithm).
    O(n) complexity — lebih efisien dari O(n³) matrix inversion.
    """
    n   = len(d)
    c_  = np.zeros(n)
    d_  = np.zeros(n)
    x   = np.zeros(n)

    c_[0] = c[0] / b[0]
    d_[0] = d[0] / b[0]

    for i in range(1, n):
        denom = b[i] - a[i] * c_[i-1]
        denom = denom if abs(denom) > 1e-30 else 1e-30
        c_[i] = c[i] / denom
        d_[i] = (d[i] - a[i] * d_[i-1]) / denom

    x[-1] = d_[-1]
    for i in range(n-2, -1, -1):
        x[i] = d_[i] - c_[i] * x[i+1]

    return x


def solve_substrate_heat(T_g: np.ndarray,
                          T_f: float,
                          T_slab_top: float,
                          G_sol: float,
                          T_a_K: float,
                          RH: float,
                          u: float,
                          theta: np.ndarray,
                          plant: PlantParameters,
                          substrate: SubstrateParameters,
                          H_g: float,
                          dt: float) -> np.ndarray:
    """Solve T_g[z] via FVM + TDMA. Persamaan (4)(5)(6)(7)."""
    sigma   = 5.67e-8
    gamma   = psychrometric_constant()
    rho_air = 1.2
    cp_air  = 1005

    Nz = len(T_g)
    dz = H_g / (Nz - 1)

    rho_cp_g, lambda_g, k_theta, _ = compute_substrate_properties(theta, substrate)

    # Boundary atas
    T_sky   = sky_temperature(T_a_K, RH)
    P_a     = ambient_vapor_pressure(T_a_K, RH)
    P_g_sat = saturation_pressure(T_g[0])
    epsilon_fg = 1 / (1/plant.epsilon_f + 1/substrate.epsilon_g - 1)

    # R_g_net = ((1 - substrate.rho_g_rad) * plant.tau_f * G_sol          # Eq.6
    #            + epsilon_fg * sigma * (T_f**4 - T_g[0]**4))

    # Solar radiation reaching substrate.
    # Original paper assumes uniform canopy: G_sub = tau_f * G_sol.
    # For real planted box, not all horizontal area is covered by leaves.
    # Open area receives direct solar, covered area receives transmitted solar.
    cover = float(np.clip(getattr(plant, "cover_fraction", 1.0), 0.0, 1.0))
    G_to_substrate = ((1.0 - cover) + cover * plant.tau_f) * G_sol

    R_g_net = ((1 - substrate.rho_g_rad) * G_to_substrate
               + epsilon_fg * sigma * (T_f**4 - T_g[0]**4))
    d0  = 0.701 * plant.H_f**0.975
    Z_M = 0.131 * plant.H_f**0.997
    Z_u = plant.H_f + 2.0
    k_vk = 0.41

    log_r = np.log(max((plant.H_f - d0)/Z_M, 1.01)) / np.log(max((Z_u - d0)/Z_M, 1.01))
    u_f = u * max(log_r, 0.01)
    a_drag = (0.28 * plant.LAI * plant.H_f * plant.d_f)**0.5
    u_g = max(u_f * np.exp(-a_drag * (1 - 0.05/max(plant.H_f, 0.01))), 0.01)

    r_c   = 1 / (0.004 + 0.012 * u_g)                                   # B.3
    r_vap = substrate_vapor_resistance(float(theta[0]), substrate)       # moisture-dependent
    r_a   = compute_aerodynamic_resistance(u, plant)

    h_conv_g = plant.LAI * (rho_air * cp_air) / (r_a + r_c)             # B.1
    h_eva_g  = plant.LAI / gamma * (rho_air * cp_air) / (r_a + r_vap)   # B.2

    q_top = R_g_net - h_conv_g*(T_g[0]-T_a_K) - h_eva_g*max(P_g_sat-P_a, 0.0)  # Eq.5

    # FVM tridiagonal
    aw = np.zeros(Nz); ap = np.zeros(Nz)
    ae = np.zeros(Nz); bv = np.zeros(Nz)

    for i in range(1, Nz-1):
        lw = 0.5*(lambda_g[i-1]+lambda_g[i])
        le = 0.5*(lambda_g[i]+lambda_g[i+1])
        aw[i] = lw/dz**2
        ae[i] = le/dz**2
        ap[i] = rho_cp_g[i]/dt + aw[i] + ae[i]
        bv[i] = rho_cp_g[i]/dt * T_g[i]

    ae[0]  = lambda_g[0]/dz**2
    ap[0]  = rho_cp_g[0]/dt + ae[0]
    bv[0]  = rho_cp_g[0]/dt * T_g[0] + q_top/dz

    # BC bawah: T_slab_top = Dirichlet boundary
    # ap must include BOTH left neighbor (aw) AND boundary conductance (a_E)
    lam_e_bc = lambda_g[-1]/dz**2       # conductance to slab boundary
    aw[-1]   = lambda_g[-1]/dz**2       # conductance to left neighbor
    ap[-1]   = rho_cp_g[-1]/dt + aw[-1] + lam_e_bc   # FIX: was missing lam_e_bc
    bv[-1]   = rho_cp_g[-1]/dt*T_g[-1] + lam_e_bc*T_slab_top

    # FIX: off-diagonal must be NEGATIVE for correct heat flow direction
    # aw, ae store positive conductance values → negate before TDMA
    return tdma_solver(-aw, ap, -ae, bv)



def solve_substrate_moisture(theta: np.ndarray,
                              T_f: float,
                              T_g_surface: float,
                              T_a_K: float,
                              RH: float,
                              u: float,
                              G_sol: float,
                              j_irrigation: float,
                              plant: PlantParameters,
                              substrate: SubstrateParameters,
                              H_g: float,
                              dt: float) -> Tuple[np.ndarray, float]:
    """
    Solve θ[z] via Richards equation + TDMA. Persamaan (8)(9)(10)(11).

    Koreksi: G_sol aktual dikirim ke compute_stomatal_resistance(), sehingga
    CAM siang/malam bekerja juga di moisture solver. Sebelumnya G_sol=0 selalu.
    """
    gamma   = psychrometric_constant()
    rho_air = 1.2
    cp_air  = 1005

    Nz = len(theta)
    dz = H_g / (Nz - 1)

    _, _, k_theta, D_theta = compute_substrate_properties(theta, substrate)

    P_a     = ambient_vapor_pressure(T_a_K, RH)
    P_f_sat = saturation_pressure(T_f)
    P_g_sat = saturation_pressure(T_g_surface)
    r_a     = compute_aerodynamic_resistance(u, plant)
    r_stoma = compute_stomatal_resistance(G_sol, T_f, P_f_sat, P_a,
                                          np.mean(theta), plant, substrate)
    r_vap   = substrate_vapor_resistance(float(theta[0]), substrate)

    h_eva_f = plant.LAI / gamma * (rho_air*cp_air) / (r_a + r_stoma)
    h_eva_g = plant.LAI / gamma * (rho_air*cp_air) / (r_a + r_vap)

    # Clip to zero: tidak memodelkan kondensasi, hanya evaporasi/transpirasi.
    j_eva_f = max(0.0, h_eva_f * (P_f_sat - P_a) / substrate.l_fg)     # Eq.10 [kg/m²s]
    j_eva_g = max(0.0, h_eva_g * (P_g_sat - P_a) / substrate.l_fg)
    j_eva   = j_eva_f + j_eva_g                                      # [kg/m²s]

    # Convert ke [m/s] untuk Richards equation.
    rho_water   = 1000.0
    j_eva_ms    = j_eva / rho_water
    j_irrig_ms  = j_irrigation / rho_water

    aw = np.zeros(Nz); ap = np.zeros(Nz)
    ae = np.zeros(Nz); bv = np.zeros(Nz)

    for i in range(1, Nz-1):
        Dw = 0.5*(D_theta[i-1]+D_theta[i])
        De = 0.5*(D_theta[i]+D_theta[i+1])
        kw = 0.5*(k_theta[i-1]+k_theta[i])
        ke = 0.5*(k_theta[i]+k_theta[i+1])
        aw[i] = Dw/dz**2
        ae[i] = De/dz**2
        ap[i] = 1/dt + aw[i] + ae[i]
        bv[i] = theta[i]/dt + (ke-kw)/dz
# APAKAH INI BIANG MASALAHNYA?
    # j_net  = j_irrig_ms - j_eva_ms                                      # Eq.9 [m/s]
    # ae[0]  = D_theta[0]/dz**2
    # ap[0]  = 1/dt + ae[0]
    # bv[0]  = theta[0]/dt + j_net/dz + k_theta[0]/dz

    j_net  = j_irrig_ms - j_eva_ms  # positive = water input at top [m/s]

    ae[0]  = D_theta[0]/dz**2
    ap[0]  = 1/dt + ae[0]

    # Top boundary water balance:
    # infiltration/rain adds water, evaporation removes water,
    # gravity conductivity should not act as an artificial water source.
    bv[0]  = theta[0]/dt + (j_net - k_theta[0]) / dz

    # apakah ini biang masalahnya? Kalau theta sudah hampir jenuh, maka D_theta besar → aw besar → ap besar → bv kecil → theta_new tidak naik sesuai ekspektasi.
    # if theta[-1] >= substrate.theta_sat*0.99:                            # Eq.11
    #     aw[-1] = D_theta[-1]/dz**2
    #     ap[-1] = 1/dt + aw[-1]
    #     bv[-1] = theta[-1]/dt - k_theta[-1]/dz
    # else:
    #     aw[-1] = D_theta[-1]/dz**2
    #     ap[-1] = 1/dt + aw[-1]
    #     bv[-1] = theta[-1]/dt

    # Bottom boundary: free drainage / unit-gradient drainage.
    # Green roof substrate should be able to drain downward, not store water
    # until it reaches full saturation.
    aw[-1] = D_theta[-1]/dz**2
    ap[-1] = 1/dt + aw[-1]
    bv[-1] = theta[-1]/dt - k_theta[-1]/dz

    theta_new = tdma_solver(-aw, ap, -ae, bv)
    theta_new = np.clip(theta_new, substrate.theta_min, substrate.theta_sat)

    return theta_new, float(j_eva)

# ==============================================================================
# SECTION 6: BLOK 3 — SLAB MODEL
# ==============================================================================

def solve_slab_heat(T_s: np.ndarray,
                    T_g_bottom: float,
                    lambda_g_bottom: float,
                    slab: SlabParameters,
                    geom: GeometryParameters,
                    dt: float,
                    T_in_K: Optional[float] = None) -> Tuple[np.ndarray, float]:
    """
    Solve T_s[z] dan q_s_in via FVM + TDMA. Persamaan (12)(13)(14).

    [C1] T_in_K sekarang PARAMETER DINAMIS:
    - Kalau T_in_K diberikan → pakai nilai dari NI sensor saat itu
    - Kalau None → fallback ke geom.T_in_default (rata-rata 29.5°C, bukan 25°C!)
    """
    Nz = len(T_s)
    dz = slab.H_slab / (Nz - 1)
    rho_cp_s = slab.rho_s * slab.cp_s

    # [C1] Gunakan T_in dinamis atau fallback ke default
    T_in = T_in_K if T_in_K is not None else geom.T_in_default

    aw = np.zeros(Nz); ap = np.zeros(Nz)
    ae = np.zeros(Nz); bv = np.zeros(Nz)

    for i in range(1, Nz-1):
        aw[i] = slab.lambda_s/dz**2
        ae[i] = slab.lambda_s/dz**2
        ap[i] = rho_cp_s/dt + aw[i] + ae[i]
        bv[i] = rho_cp_s/dt * T_s[i]

    # BC atas slab: T_g_bottom = Dirichlet boundary dari substrat
    lam_w_bc = lambda_g_bottom/dz**2    # conductance ke boundary substrat
    ae[0]    = slab.lambda_s/dz**2
    ap[0]    = rho_cp_s/dt + ae[0] + lam_w_bc   # FIX: was missing lam_w_bc
    bv[0]    = rho_cp_s/dt*T_s[0] + lam_w_bc*T_g_bottom            # Eq.13

    # h_in   = geom.h_in          -> suruh ganti ma gpt
    if getattr(geom, "dynamic_h_in", False):
        h_in = interior_ceiling_convective_h(T_s[-1], T_in)
    else:
        h_in = geom.h_in
    aw[-1] = slab.lambda_s/dz**2
    ap[-1] = rho_cp_s/dt + aw[-1] + h_in/dz
    bv[-1] = rho_cp_s/dt*T_s[-1] + h_in*T_in/dz                        # Eq.14

    # FIX: negate off-diagonal for correct conduction direction
    T_s_new = tdma_solver(-aw, ap, -ae, bv)

    q_s_in = h_in * (T_s_new[-1] - T_in)   # Eq.14: positif=masuk, negatif=keluar

    return T_s_new, q_s_in


# ==============================================================================
# SECTION 7: MAIN SIMULATION LOOP
# ==============================================================================


def run_simulation(weather_df: pd.DataFrame,
                   plant: PlantParameters,
                   substrate: SubstrateParameters,
                   slab: SlabParameters,
                   geom: GeometryParameters,
                   num: NumericalParameters,
                   theta_initial = None,
                   j_irrigation: float = 0.0,
                   T_in_series: pd.Series = None,
                   T_g_top_initial_C: Optional[float] = None,
                   T_s_in_initial_C: Optional[float] = None,
                   save_every_s: int = 60) -> dict:
    """
    Main simulation loop mengikuti Figure 2 flowchart dari paper.

    Perubahan dari versi awal:
    - Weather 1 detik dibuat dengan prepare_weather_1s(): rain_flux tidak
      diinterpolasi linear.
    - Hasil menyimpan datetime, sehingga validasi bisa align by timestamp.
    - theta_initial bisa berupa profil dari RIKA 2 cm dan 7 cm.
    - G_sol dikirim ke moisture solver agar CAM siang/malam aktif.
    """
    print(f"\n{'='*60}")
    print(f"Simulasi: {plant.name} ({plant.plant_type})")
    if T_in_series is not None:
        print(f"[C1] T_in boundary: dinamis dari NI "
              f"({T_in_series.min():.1f}–{T_in_series.max():.1f}°C)")
    else:
        print(f"[C1] T_in boundary: default {geom.T_in_default-273.15:.1f}°C")
    print(f"{'='*60}")

    if geom.H_g is None or slab.H_slab is None:
        raise ValueError("geom.H_g dan slab.H_slab harus diisi sebelum simulasi")

    Nz_g = num.Nz_substrate
    Nz_s = num.Nz_slab
    dt   = num.dt

    print("Resample weather data ke 1 detik...")
    weather_1s = prepare_weather_1s(weather_df, dt=dt)
    N_steps = len(weather_1s)
    print(f"Total timesteps: {N_steps:,} ({N_steps/3600:.1f} jam)")

    T_in_1s = None
    if T_in_series is not None:
        T_in_1s = (T_in_series.sort_index()
                   .resample(f'{int(dt)}s').interpolate('time')
                   .reindex(weather_1s.index).interpolate('time'))

    # Kondisi awal
    T_a_init = float(weather_1s['T_a'].iloc[0]) + 273.15
    T_g_top_init = T_a_init if T_g_top_initial_C is None else T_g_top_initial_C + 273.15
    T_s_in_init  = T_a_init if T_s_in_initial_C is None else T_s_in_initial_C + 273.15

    if theta_initial is None:
        print(f"theta_initial tidak ada → pakai {substrate.theta_sat*0.80:.2f} (80% sat)")

    # Profil awal lebih realistis untuk validasi pendek.
    T_g   = np.linspace(T_g_top_init, T_a_init, Nz_g)
    T_s   = np.linspace(T_a_init, T_s_in_init, Nz_s)
    T_f   = T_a_init
    theta = make_theta_initial_profile(theta_initial, geom.H_g, Nz_g, substrate)

    results = {
        'datetime'  : [],
        'time'      : [], 'T_f'    : [], 'T_g_top': [],
        'T_g_mid'   : [], 'T_g_bot': [], 'T_s_in' : [],
        'theta_top' : [], 'theta_mid': [], 'theta_bot': [],
        'q_s_in'    : [], 'j_eva'  : [],
        'T_a'       : [], 'G_sol'  : [],
        'T_in_used' : [],
        'j_pr': [],
        'theta_mean': [],
    }

    print("Mulai simulasi...\n")

    for step, (ts, row) in enumerate(weather_1s.iterrows()):
        T_a_K = float(row['T_a']) + 273.15
        G_sol = max(float(row['G_sol']), 0.0)
        RH    = float(np.clip(row['RH'], 1, 99))
        u     = max(float(row['u']), 0.1)
        j_pr  = float(j_irrigation) + max(float(row.get('rain_flux', 0.0)), 0.0)

        if T_in_1s is not None and ts in T_in_1s.index and not pd.isna(T_in_1s.loc[ts]):
            T_in_current = float(T_in_1s.loc[ts]) + 273.15
        else:
            T_in_current = geom.T_in_default

        theta_avg = float(np.mean(theta))

        # BLOK 1: Foliage
        T_f = solve_foliage_temperature(
            T_f_prev=T_f, T_a_K=T_a_K, T_g_surface=T_g[0],
            G_sol=G_sol, RH=RH, u=u, theta_avg=theta_avg,
            plant=plant, substrate=substrate, dt=dt
        )

        # BLOK 2a: Substrat heat
        T_g = solve_substrate_heat(
            T_g=T_g, T_f=T_f, T_slab_top=T_s[0],
            G_sol=G_sol, T_a_K=T_a_K, RH=RH, u=u, theta=theta,
            plant=plant, substrate=substrate, H_g=geom.H_g, dt=dt
        )

        # BLOK 2b: Substrat moisture
        theta, j_eva = solve_substrate_moisture(
            theta=theta, T_f=T_f, T_g_surface=T_g[0],
            T_a_K=T_a_K, RH=RH, u=u, G_sol=G_sol, j_irrigation=j_pr,
            plant=plant, substrate=substrate, H_g=geom.H_g, dt=dt
        )

        # BLOK 3: Slab + q_s_in
        _, lambda_g, _, _ = compute_substrate_properties(theta, substrate)
        T_s, q_s_in = solve_slab_heat(
            T_s=T_s, T_g_bottom=T_g[-1],
            lambda_g_bottom=lambda_g[-1],
            slab=slab, geom=geom, dt=dt,
            T_in_K=T_in_current
        )

        if step % int(save_every_s/dt) == 0:
            results['datetime'].append(ts)
            results['time'].append(step * dt)
            results['T_f'].append(T_f - 273.15)
            results['T_g_top'].append(T_g[0] - 273.15)
            results['T_g_mid'].append(T_g[Nz_g//2] - 273.15)
            results['T_g_bot'].append(T_g[-1] - 273.15)
            results['T_s_in'].append(T_s[-1] - 273.15)
            results['theta_top'].append(float(theta[0]))
            results['theta_mid'].append(float(theta[Nz_g//2]))
            results['theta_bot'].append(float(theta[-1]))
            results['q_s_in'].append(float(q_s_in))
            results['j_eva'].append(float(j_eva))
            results['T_a'].append(T_a_K - 273.15)
            results['G_sol'].append(G_sol)
            results['T_in_used'].append(T_in_current - 273.15)
            results['theta_mean'].append(float(np.mean(theta)))
            results['j_pr'].append(float(j_pr))

        if step % int(6*3600/dt) == 0:
            print(f"  t={step*dt/3600:5.1f}h | "
                  f"T_f={T_f-273.15:5.1f}°C | "
                  f"T_g={T_g[0]-273.15:5.1f}°C | "
                  f"T_s_in={T_s[-1]-273.15:5.1f}°C | "
                  f"θ={theta[0]:.3f} | "
                  f"T_in={T_in_current-273.15:.1f}°C | "
                  f"q={q_s_in:6.1f} W/m²")

    # Q_gain — Persamaan (18)
    results['Q_gain'] = float(np.trapezoid(results['q_s_in'], dx=save_every_s))  # J/m²
    if getattr(geom, 'A_roof', None) is not None:
        results['Q_gain_total'] = results['Q_gain'] * geom.A_roof  # J untuk luas atap aktual

    print(f"\nSelesai. Q_gain = {results['Q_gain']:.1f} J/m²")
    if 'Q_gain_total' in results:
        print(f"        Q_gain_total = {results['Q_gain_total']:.1f} J (A_roof={geom.A_roof:.4f} m²)")
    return results


def series_from_results(results: dict, key: str) -> pd.Series:
    """Konversi output simulasi ke Series bertimestamp."""
    return pd.Series(results[key], index=pd.to_datetime(results['datetime']), name=key)


def validation_metrics(sim: pd.Series, obs: pd.Series) -> Dict[str, float]:
    """Hitung bias/MAE/RMSE hanya pada common timestamp."""
    sim = sim.sort_index().resample('1min').mean()
    obs = obs.sort_index().resample('1min').mean()
    common = sim.index.intersection(obs.index)
    diff = sim.loc[common] - obs.loc[common]
    return {
        'n': int(diff.count()),
        'bias_C': float(diff.mean()),
        'mae_C': float(diff.abs().mean()),
        'rmse_C': float(np.sqrt((diff**2).mean())),
    }

# ==============================================================================
# SECTION 8: KALIBRASI LAI
# ==============================================================================


def calibrate_LAI(weather_df: pd.DataFrame,
                  NI_data: pd.DataFrame,
                  plant: PlantParameters,
                  substrate: SubstrateParameters,
                  slab: SlabParameters,
                  geom: GeometryParameters,
                  num: NumericalParameters,
                  theta_initial = None,
                  LAI_bounds: Tuple[float, float] = (0.3, 3.5),
                  T_in_series: pd.Series = None,
                  target_col: str = None,
                  T_g_top_initial_C: Optional[float] = None,
                  T_s_in_initial_C: Optional[float] = None) -> Tuple[float, dict]:
    """
    Kalibrasi LAI berbasis timestamp.

    Target default:
    - CAM: T_s_in_CAM (T1Tb) sebagai target utama; T_g_top_CAM bisa dipakai nanti
    - C3 : T_s_in_C3  (T3Ka)

    Catatan: untuk update sementara, kalibrasi hanya ke T_s_in agar sederhana dan
    tidak terlalu overfit.
    """
    from scipy.optimize import minimize_scalar

    if T_in_series is None:
        T_in_series = NI_data['T_in_CAM'] if plant.plant_type == 'CAM' else NI_data['T_in_C3']

    if target_col is None:
        target_col = 'T_s_in_CAM' if plant.plant_type == 'CAM' else 'T_s_in_C3'

    if T_s_in_initial_C is None and target_col in NI_data:
        T_s_in_initial_C = float(NI_data[target_col].dropna().iloc[0])

    if T_g_top_initial_C is None and plant.plant_type == 'CAM' and 'T_g_top_CAM' in NI_data:
        T_g_top_initial_C = float(NI_data['T_g_top_CAM'].dropna().iloc[0])

    print(f"\nKalibrasi LAI {plant.name}")
    print(f"Target: model T_s_in vs NI {target_col}")

    iteration = [0]

    def objective(LAI):
        iteration[0] += 1
        plant.LAI = float(LAI)
        results = run_simulation(
            weather_df=weather_df, plant=plant,
            substrate=substrate, slab=slab,
            geom=geom, num=num,
            theta_initial=theta_initial,
            T_in_series=T_in_series,
            T_g_top_initial_C=T_g_top_initial_C,
            T_s_in_initial_C=T_s_in_initial_C,
        )
        sim = series_from_results(results, 'T_s_in')
        obs = NI_data[target_col]
        m = validation_metrics(sim, obs)
        print(f"  Iterasi {iteration[0]:2d} | LAI={LAI:.3f} | RMSE={m['rmse_C']:.3f}°C")
        return m['rmse_C']

    result = minimize_scalar(objective, bounds=LAI_bounds,
                             method='bounded', options={'xatol': 0.03})

    LAI_optimal = float(result.x)
    plant.LAI = LAI_optimal
    print(f"\nLAI optimal ({plant.name}) = {LAI_optimal:.3f}")

    cal_results = run_simulation(
        weather_df=weather_df, plant=plant,
        substrate=substrate, slab=slab,
        geom=geom, num=num,
        theta_initial=theta_initial,
        T_in_series=T_in_series,
        T_g_top_initial_C=T_g_top_initial_C,
        T_s_in_initial_C=T_s_in_initial_C,
    )

    return LAI_optimal, cal_results


def prepare_validation_case(plant_type: str,
                            base_dir: str = "/mnt/data") -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    """Load weather, NI, RIKA sesuai window validasi CAM/C3."""
    plant_type = plant_type.upper()
    start, end = VALIDATION_WINDOWS[plant_type]
    start = pd.Timestamp(start)
    end = pd.Timestamp(end)
    base = Path(base_dir)

    weather = load_weather_cache_or_excel(
        str(base / 'weatherfile mar-april.xlsx'),
        cache_path=str(base / 'weather_clean_cache.csv'),
        date_start=str(start), date_end=str(end),
    )

    ni_files = [
        str(base / 'Pengukuran 30_1 Maret 2026.xlsx'),
        str(base / 'Pengukuran 30_2 Maret 2026.xlsx'),
        str(base / 'Pengukuran 3 April 2026.xlsx'),
    ]
    ni = load_multiple_NI_sensor_data(ni_files)
    ni = ni[(ni.index >= start) & (ni.index <= end)].copy()

    rika = load_soil_moisture_data(str(base / 'datasoilmoisture.zip'), plant_type=plant_type)
    rika = rika[(rika.index >= start) & (rika.index <= end)].copy()

    theta_initial = get_theta_initial_from_rika(rika, start)
    return weather, ni, rika, theta_initial


def run_validation_case(plant_type: str,
                        base_dir: str = "/mnt/data",
                        calibrate_lai: bool = False) -> Tuple[dict, Dict[str, float]]:
    """Shortcut untuk run CAM/C3 dengan window validasi sementara."""
    # apply_scientific_guess_parameters()
    plant_type = plant_type.upper()
    plant = bromelia if plant_type == 'CAM' else wedelia
    weather, ni, rika, theta_initial = prepare_validation_case(plant_type, base_dir)

# hahahahah ganti benrar
    # if plant_type == 'CAM':
    #     T_in_series = ni['T_in_CAM']
    #     target_col = 'T_s_in_CAM'
    #     T_g_top_initial = float(ni['T_g_top_CAM'].dropna().iloc[0]) if 'T_g_top_CAM' in ni else None
    # else:
    #     T_in_series = ni['T_in_C3']
    #     target_col = 'T_s_in_C3'
    #     T_g_top_initial = None
    if plant_type == 'CAM':
        T_in_series = ni['T_in_CAM']

        # Target validasi CAM bisa diganti dari runner.
        # Default: T1Tb
        target_col = VALIDATION_TARGETS.get("CAM", "T1Tb")

        T_g_top_initial = (
            float(ni['T_g_top_CAM'].dropna().iloc[0])
            if 'T_g_top_CAM' in ni and not ni['T_g_top_CAM'].dropna().empty
            else None
        )

    else:
        T_in_series = ni['T_in_C3']

        # Target validasi C3 default: T3Ka
        target_col = VALIDATION_TARGETS.get("C3", "T3Ka")

        T_g_top_initial = None


    if target_col not in ni.columns:
        raise ValueError(
            f"Kolom target validasi '{target_col}' tidak ditemukan di NI data. "
            f"Kolom tersedia: {ni.columns.tolist()}"
        )

    if ni[target_col].dropna().empty:
        raise ValueError(
            f"Kolom target validasi '{target_col}' kosong pada window {plant_type}."
        )
    T_s_in_initial = float(ni[target_col].dropna().iloc[0])

    if calibrate_lai:
        _, results = calibrate_LAI(
            weather_df=weather, NI_data=ni, plant=plant,
            substrate=substrat, slab=slab, geom=geom, num=num,
            theta_initial=theta_initial, T_in_series=T_in_series,
            target_col=target_col,
            T_g_top_initial_C=T_g_top_initial,
            T_s_in_initial_C=T_s_in_initial,
        )
    else:
        results = run_simulation(
            weather_df=weather, plant=plant,
            substrate=substrat, slab=slab, geom=geom, num=num,
            theta_initial=theta_initial,
            T_in_series=T_in_series,
            T_g_top_initial_C=T_g_top_initial,
            T_s_in_initial_C=T_s_in_initial,
        )

    metrics = validation_metrics(series_from_results(results, 'T_s_in'), ni[target_col])
    metrics['LAI_used'] = float(plant.LAI)
    metrics['window_start'] = str(VALIDATION_WINDOWS[plant_type][0])
    metrics['window_end'] = str(VALIDATION_WINDOWS[plant_type][1])
    return results, metrics

# ==============================================================================
# SECTION 9: VISUALISASI
# ==============================================================================

def plot_results(results_cam: dict,
                  results_c3: dict,
                  weather_df: pd.DataFrame,
                  NI_data: pd.DataFrame = None,
                  save_path: str = None):
    """
    6-panel comparison plot: CAM vs C3.
    [C1] Panel tambahan untuk T_in aktual vs 25°C asumsi paper.
    """
    t_cam = np.array(results_cam['time']) / 3600
    t_c3  = np.array(results_c3['time'])  / 3600

    fig = plt.figure(figsize=(16, 16))
    fig.suptitle('Green Roof Thermal Simulation — CAM (Bromelia) vs C3 (Wedelia)\n'
                 'Universitas Indonesia | Koreksi: T_in dinamis, tau_f diperbarui',
                 fontsize=13, fontweight='bold')

    gs = gridspec.GridSpec(3, 2, figure=fig, hspace=0.45, wspace=0.3)

    # (a) Suhu CAM
    ax = fig.add_subplot(gs[0, 0])
    ax.plot(t_cam, results_cam['T_a'],     'k--', lw=1.2, label='T_a', alpha=0.7)
    ax.plot(t_cam, results_cam['T_f'],     'b-',  lw=1.5, label='T_f (foliage)')
    ax.plot(t_cam, results_cam['T_g_top'], 'g-o', lw=1.2, label='T_g_top',
            markevery=60, ms=3)
    ax.plot(t_cam, results_cam['T_s_in'],  'r-^', lw=1.2, label='T_s_in',
            markevery=60, ms=3)
    if NI_data is not None:
        t_ni = np.arange(len(NI_data)) / 60
        ax.plot(t_ni, NI_data['T_g_top_CAM'].values, 'g.', ms=2, alpha=0.4,
                label='T_g_top (NI)')
        ax.plot(t_ni, NI_data['T_s_in_CAM'].values,  'r.', ms=2, alpha=0.4,
                label='T_s_in (NI)')
    ax.set_title('(a) Suhu — Bromelia (CAM)')
    ax.set_ylabel('Suhu (°C)'); ax.set_xlabel('Waktu (jam)')
    ax.legend(fontsize=7); ax.grid(True, alpha=0.3)

    # (b) Suhu C3
    ax = fig.add_subplot(gs[0, 1])
    ax.plot(t_c3,  results_c3['T_a'],     'k--', lw=1.2, label='T_a', alpha=0.7)
    ax.plot(t_c3,  results_c3['T_f'],     'b-',  lw=1.5, label='T_f (foliage)')
    ax.plot(t_c3,  results_c3['T_g_top'], 'g-o', lw=1.2, label='T_g_top',
            markevery=60, ms=3)
    ax.plot(t_c3,  results_c3['T_s_in'],  'r-^', lw=1.2, label='T_s_in',
            markevery=60, ms=3)
    if NI_data is not None:
        ax.plot(t_ni, NI_data['T_s_in_C3'].values, 'r.', ms=2, alpha=0.4,
                label='T_s_in (NI)')
    ax.set_title('(b) Suhu — Wedelia (C3)')
    ax.set_ylabel('Suhu (°C)'); ax.set_xlabel('Waktu (jam)')
    ax.legend(fontsize=7); ax.grid(True, alpha=0.3)

    # (c) Heat flux
    ax = fig.add_subplot(gs[1, 0])
    ax.plot(t_cam, results_cam['q_s_in'], 'b-', lw=1.5, label='CAM (Bromelia)')
    ax.plot(t_c3,  results_c3['q_s_in'],  'r-', lw=1.5, label='C3 (Wedelia)')
    ax.axhline(0, color='gray', ls='--', alpha=0.5)
    ax.fill_between(t_cam, results_cam['q_s_in'], 0,
                    where=np.array(results_cam['q_s_in'])>0,
                    alpha=0.15, color='blue')
    Q_cam = results_cam['Q_gain']
    Q_c3  = results_c3['Q_gain']
    ax.text(0.02, 0.97, f"Q_gain CAM = {Q_cam:.0f} J/m²\nQ_gain C3  = {Q_c3:.0f} J/m²",
            transform=ax.transAxes, va='top', fontsize=8,
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.6))
    ax.set_title('(c) Heat Flux ke Interior')
    ax.set_ylabel('q_s_in (W/m²)'); ax.set_xlabel('Waktu (jam)')
    ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # (d) VWC
    ax = fig.add_subplot(gs[1, 1])
    ax.plot(t_cam, results_cam['theta_top'], 'b-',  lw=1.5, label='CAM θ_top')
    ax.plot(t_cam, results_cam['theta_bot'], 'b--', lw=1.0, label='CAM θ_bot')
    ax.plot(t_c3,  results_c3['theta_top'],  'r-',  lw=1.5, label='C3 θ_top')
    ax.plot(t_c3,  results_c3['theta_bot'],  'r--', lw=1.0, label='C3 θ_bot')
    ax.set_title('(d) Volumetric Water Content')
    ax.set_ylabel('VWC (m³/m³)'); ax.set_xlabel('Waktu (jam)')
    ax.legend(fontsize=7); ax.grid(True, alpha=0.3)

    # (e) Evapotranspirasi
    ax = fig.add_subplot(gs[2, 0])
    ax.plot(t_cam, np.array(results_cam['j_eva'])*3600, 'b-', lw=1.5,
            label='CAM')
    ax.plot(t_c3,  np.array(results_c3['j_eva'])*3600,  'r-', lw=1.5,
            label='C3')
    ax.set_title('(e) Evapotranspirasi')
    ax.set_ylabel('ET (kg/m²jam)'); ax.set_xlabel('Waktu (jam)')
    ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # (f) [C1] T_in aktual vs asumsi 25°C
    ax = fig.add_subplot(gs[2, 1])
    ax.plot(t_cam, results_cam['T_in_used'], 'b-', lw=1.5, label='T_in CAM (T1Ka)')
    ax.plot(t_c3,  results_c3['T_in_used'],  'r-', lw=1.5, label='T_in C3 (T3Kd)')
    ax.axhline(25.0, color='gray', ls='--', alpha=0.7, label='Asumsi paper (25°C)')
    ax.set_title('[C1] T_in Aktual vs Asumsi Paper')
    ax.set_ylabel('T_in (°C)'); ax.set_xlabel('Waktu (jam)')
    ax.legend(fontsize=7); ax.grid(True, alpha=0.3)
    ax.text(0.02, 0.97, 'Bukan 25°C!', transform=ax.transAxes,
            va='top', color='red', fontsize=9, fontweight='bold')

    plt.tight_layout()

    if save_path:
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"Plot disimpan: {save_path}")

    plt.show()


# ==============================================================================
# SECTION 10: MAIN — CARA PEMAKAIAN
# ==============================================================================

if __name__ == "__main__":

    print("=" * 60)
    print("GREEN ROOF SIMULATION — SAME-STRUCTURE UPDATED VERSION")
    print("=" * 60)

    print("\n[WINDOW VALIDASI]")
    for k, (a, b) in VALIDATION_WINDOWS.items():
        print(f"  {k}: {a} → {b} ({(b-a).total_seconds()/3600:.1f} jam)")

    print("\n[PARAMETER SEMENTARA / SCIENTIFIC GUESS]")
    apply_scientific_guess_parameters()
    print(f"  rho_g       = {substrat.rho_g} kg/m³")
    print(f"  H_g         = {geom.H_g} m")
    print(f"  H_slab      = {slab.H_slab} m")
    print(f"  A_roof      = {geom.A_roof} m²")
    print(f"  theta_sat   = {substrat.theta_sat}")
    print(f"  k_theta_sat = {substrat.k_theta_sat:.2e} m/s")
    print(f"  lambda_dry  = {substrat.lambda_dry} W/mK")
    print(f"  lambda_sat  = {substrat.lambda_sat:.3f} W/mK")

    print("\n[CARA RUN]")
    print("""
# CAM tanpa kalibrasi LAI
res_cam, m_cam = run_validation_case('CAM', calibrate_lai=False)
print(m_cam)

# C3 tanpa kalibrasi LAI
res_c3, m_c3 = run_validation_case('C3', calibrate_lai=False)
print(m_c3)

# Kalau mau coba kalibrasi LAI kasar:
# res_cam, m_cam = run_validation_case('CAM', calibrate_lai=True)
# res_c3,  m_c3  = run_validation_case('C3',  calibrate_lai=True)
""")
